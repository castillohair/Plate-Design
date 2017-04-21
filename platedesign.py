"""
Tools for designing plate experiments

"""
import copy
import os
import random

import numpy
import openpyxl
import pandas

# Versions should comply with PEP440.  For a discussion on single-sourcing
# the version across setup.py and the project code, see
# https://packaging.python.org/en/latest/single_source_version.html
__version__ = '0.1.0'

def _ceil_log(x):
    """
    Return the ceiling of the input at the corresponding order of magnitude.

    If ``x = a*10^b``, where ``1 <= a < 10``, then this function returns
    ``ceil(a)*10^b``, where ``ceil()`` is a standard ceiling function.

    Parameters
    ----------
    x : float
        Number to round.

    Returns
    -------
    float
        The log ceiling of `x`.

    Examples
    --------

    >>> _ceil_log(932)
    1000.0

    >>> _ceil_log(70)
    70.0

    >>> _ceil_log(71)
    80.0

    >>> _ceil_log(69)
    70.0

    >>> _ceil_log(6.9)
    7.0

    >>> _ceil_log(9.9)
    10.0

    """
    xom = 10.**numpy.floor(numpy.log10(x))
    return numpy.ceil(x/xom)*xom

class Experiment(object):
    """
    Object that represents a plate experiment.

    """
    def __init__(self):
        # Initialize properties
        self.n_replicates = 3
        self.randomize = False
        self.sample_vol = 500.
        self.metadata = {}
        # Initialize containers of plates and inducers
        self.plates = []
        self.inducers = []
        # to_apply contains dictionaries with {plate, inducer, across} that
        # should be applied during generation
        self.to_apply = []
        # Other properties used for calculations
        self.inducer_safety_factor = 1.2
        # Spreadsheet to append to replicate spreadsheets: Sheets other than
        # "Samples" will be copied unmodified. If a "Samples" sheet is present,
        # all columns will be added to the final "Samples" sheet, and filled
        # with the values present in the first row.
        self.spreadsheet_to_append = None

    def add_plate(self, plate):
        self.plates.append(plate)

    def add_inducer(self, inducer):
        self.inducers.append(inducer)

    def apply_inducer(self, plate, inducer, across='wells'):
        # Check across argument
        if across not in ['rows', 'cols', 'wells', 'all']:
            raise ValueError('cannot apply inducer across "{}"'.format(across))

        # Add plate if necessary
        if plate not in self.plates:
            self.plates.append(plate)

        # Add inducer if necessary
        if inducer not in self.inducers:
            self.inducers.append(inducer)

        # Add pair
        self.to_apply.append({'plate': plate,
                              'inducer': inducer,
                              'across': across})

    def generate_files(self):
        """
        """
        # Create folders for each replicate
        replicate_folders = ['replicate_{:03d}'.format(i + 1)
                             for i in range(self.n_replicates)]
        for folder in replicate_folders:
            if not os.path.exists(folder):
                os.makedirs(folder)

        # Run inducer calculations
        for inducer in self.inducers:
            # Get all applications of this inducer
            to_apply = [a for a in self.to_apply if a['inducer'] is inducer]

            # Iterate over applications and calculate required volume per
            # replicate
            n_samples = 0
            for a in to_apply:
                n_samples += a['plate'].n_samples_to_apply(a['across'])

            # Data needed for calculations downstream
            inducer.sample_vol = self.sample_vol

            # Calculate volumes
            inducer.calculate_vol(n_samples=n_samples,
                                  n_replicates=self.n_replicates,
                                  safety_factor=self.inducer_safety_factor)

            # Calculate recipe for inducer
            inducer_df = inducer.generate_recipe(
                file_name='inducer_recipes.xlsx')

            # Generate random shufflings of the data
            if self.randomize:
                inducer.generate_shufflings(self.n_replicates)

        # Iterate over replicates
        for replicate_idx in range(self.n_replicates):

            # Get output file name
            rep_file_name = 'experiment_rep{:03d}.xlsx'.format(
                replicate_idx + 1)

            # Start empty dataframe
            samples_table = pandas.DataFrame()
            samples_table_columns = []

            # Run all plate calculations
            for plate in self.plates:
                # Get all applications of this plate
                to_apply = [a for a in self.to_apply if a['plate'] is plate]

                # Set id offset for plate
                plate.id_offset = len(samples_table)
                # Reset plate
                plate.reset()

                # Apply each inducer to plate
                for a in to_apply:
                    # Randomize inducer
                    a['inducer'].randomize(replicate=replicate_idx)
                    # Apply inducer to plate
                    plate.apply_inducer(a['inducer'], a['across'])

                # Generate pipetting sheet
                plate.generate_pipetting_layout(
                    path=replicate_folders[replicate_idx],
                    file_name=rep_file_name)

                # Generate samples table
                plate.generate_samples_table()

                # The following is necessary to preserve the order of the
                # columns when appending
                for col in plate.samples.columns:
                    if col not in samples_table_columns:
                        samples_table_columns.append(col)
                # Append samples table to samples table
                samples_table = samples_table.append(plate.samples)
                # Sort columns
                samples_table = samples_table[samples_table_columns]


            # Generate full file name
            full_file_name = os.path.join(replicate_folders[replicate_idx],
                                          rep_file_name)
            # Generate pandas writer
            writer = pandas.ExcelWriter(full_file_name, engine='openpyxl')
            # If file already exists, open and copy all previous data
            if os.path.isfile(full_file_name):
                workbook = openpyxl.load_workbook(full_file_name)
                writer.book = workbook
                writer.sheets = dict((ws.title, ws)
                                     for ws in workbook.worksheets)

            # Open spreadsheet to append and copy sheets
            if self.spreadsheet_to_append is not None:
                workbook_append = openpyxl.load_workbook(
                    self.spreadsheet_to_append)
                for ws in workbook_append.worksheets:
                    if ws.title != "Samples":
                        # Create new sheet
                        ws_new = writer.book.create_sheet(ws.title)
                        # Copy all cells' content and style
                        for row in ws.iter_rows():
                            for cell in row:
                                new_cell = ws_new.cell(row=cell.row,
                                                       column=cell.col_idx)
                                new_cell.value = cell.value
                                if cell.has_style:
                                    new_cell.font = \
                                        copy.copy(cell.font)
                                    new_cell.border = \
                                        copy.copy(cell.border)
                                    new_cell.fill = \
                                        copy.copy(cell.fill)
                                    new_cell.number_format = \
                                        copy.copy(cell.number_format)
                                    new_cell.protection = \
                                        copy.copy(cell.protection)
                                    new_cell.alignment = \
                                        copy.copy(cell.alignment)
                    else:
                        # Read with pandas
                        samples_extra = pandas.read_excel(
                            self.spreadsheet_to_append,
                            sheetname="Samples")
                        # Extract first row
                        samples_extra = samples_extra.iloc[0]
                        # Add columns to samples_table
                        for index, value in samples_extra.iteritems():
                            try:
                                value = [value.format(i + 1)
                                         for i in range(len(samples_table))]
                            except AttributeError as e:
                                pass
                            samples_table[index] = value

            # Convert pandas table to sheet and save
            samples_table.to_excel(writer, sheet_name='Samples')
            writer.save()

class ChemicalInducer(object):
    """
    Object that generates recipes for dilutions of chemical inducers.

    Examples
    --------
    Declare an inducer called iptg

    >>> iptg = ChemicalInducer(name='IPTG', units='uM', stock_conc=1e6)

    Generate a gradient from 0.1uM to 500uM, log scale, include zero.

    >>> iptg.generate_gradient(min=1,
    ...                        max=500,
    ...                        n=12,
    ...                        scale='log',
    ...                        use_zero=True)

    >>> iptg.calculate_recipe(inoculation_vol=5.,
    ...                       sample_vol=500.,
    ...                       n_samples=6,
    ...                       min_total_vol=100)

    """
    def __init__(self, name, units, id_prefix=None, id_offset=0):
        # Store name and units
        self.name = name
        self.units = units

        # Store ID modifiers for dilutions table
        if id_prefix is None:
            id_prefix=name[0]
        self.id_prefix=id_prefix
        self.id_offset=id_offset

        # Initialize main properties used for calculations
        # Stock concentration of inducer
        self.stock_conc = None
        # Volume of cell sample
        self.sample_vol = None
        # Volume of inducer dilution that will go into each cell sample
        self.inoculation_vol = None

        # Total volume of inducer to make
        self.total_vol = None
        # Volume to aliquot
        self.aliquot_vol = None

        # Initialize secondary properties used for calculations
        self.min_inducer_vol = 1.5
        self.max_inducer_vol = 20.
        self.dilution_step = 10.
        self.inducer_decimals = 2
        self.water_decimals = 1

        # Reset all structures
        self.reset()

    def __str__(self):
        return "{}, {} different concentration(s)".format(
            self.name,
            len(self.dilutions))

    def reset(self):
        # The following initializes an empty dilutions table
        self.conc = [0]
        # Randomization indices
        self.rand_idx = None

    @property
    def conc_header(self):
        return "Concentration ({})".format(self.units)

    @property
    def conc(self):
        return self.dilutions[self.conc_header].values

    @conc.setter
    def conc(self, value):
        # Initialize dataframe with dilutions info
        ids = ['{}{:03d}'.format(self.id_prefix, i)
               for i in range(self.id_offset + 1,
                              len(value) + self.id_offset + 1)]
        self.dilutions = pandas.DataFrame({'ID': ids})
        self.dilutions.set_index('ID', inplace=True)
        self.dilutions[self.conc_header] = value

    def set_gradient(self,
                     min,
                     max,
                     n,
                     n_repeat=1,
                     scale='linear',
                     use_zero=False):
        """
        """

        # Calculate gradient
        if scale == 'linear':
            self.conc = numpy.linspace(min, max, n/n_repeat)
            self.conc = numpy.repeat(self.conc, n_repeat)
        elif scale == 'log':
            if use_zero:
                self.conc = numpy.logspace(numpy.log10(min),
                                           numpy.log10(max),
                                           (n/n_repeat - 1))
                self.conc = numpy.concatenate(([0], self.conc))
            else:
                self.conc = numpy.logspace(numpy.log10(min),
                                           numpy.log10(max),
                                           n/n_repeat)
            self.conc = numpy.repeat(self.conc, n_repeat)
        else:
            raise ValueError("scale {} not recognized".format(scale))

    def calculate_vol(self,
                      n_samples,
                      n_replicates=1,
                      safety_factor=1.2):

        # Calculate volume with safety factor for a replicate
        inducer_rep_vol = n_samples*self.inoculation_vol
        inducer_rep_vol = inducer_rep_vol*safety_factor
        inducer_rep_vol = _ceil_log(inducer_rep_vol)

        if n_replicates > 1:
            # Calculate total amount of inducer
            self.total_vol = inducer_rep_vol*n_replicates
            # Apply safety factor and round again
            self.total_vol = _ceil_log(self.total_vol*safety_factor)
            # Store aliquot volume
            self.aliquot_vol = inducer_rep_vol
        else:
            # Store replicate volume as total volume
            self.total_vol = inducer_rep_vol
            self.aliquot_vol = None

    def generate_recipe(self,
                        path='.',
                        file_name=None,
                        sheet_name=None,
                        ):
        """
        """

        # Initialize relevant arrays
        inducer_vols = numpy.zeros_like(self.conc)
        inducer_dils = numpy.zeros_like(self.conc)
        water_vols = numpy.zeros_like(self.conc)
        actual_concs = numpy.zeros_like(self.conc)

        # Iterate over concentrations
        for i, conc in enumerate(self.conc):
            # If concentration is zero, skip
            if conc == 0:
                inducer_dils[i] = 0
                inducer_vols[i] = 0
                water_vols[i] = self.total_vol
                continue
            # Determine the appropriate dilution to use
            # We start with a high dilution, and scale down until we reach a
            # volume that is acceptable (lower than max_inducer_volume)
            dilution = self.dilution_step**10
            while True:
                inducer_vol = (conc*self.sample_vol/self.inoculation_vol) * \
                    self.total_vol / (self.stock_conc/dilution)
                if (inducer_vol < self.max_inducer_vol):
                    break
                if dilution/self.dilution_step < 1:
                    break
                dilution = dilution/self.dilution_step
            # Round inducer volume to the specified precision
            inducer_vol = numpy.around(inducer_vol,
                                       decimals=self.inducer_decimals)
            # Water volume is the remaining volume
            water_vol = numpy.around(self.total_vol - inducer_vol,
                                     decimals=self.water_decimals)
            # Actual concentration achieved
            actual_conc = self.stock_conc/dilution * \
                          inducer_vol/(inducer_vol + water_vol) * \
                          (self.inoculation_vol/self.sample_vol)
            # Accumulate
            inducer_dils[i] = dilution
            inducer_vols[i] = inducer_vol
            water_vols[i] = water_vol
            actual_concs[i] = actual_conc

        # Make dataframe 
        self.dilutions[self.conc_header] = actual_concs
        self.dilutions['Stock dilution'] = inducer_dils
        self.dilutions['Inducer volume (uL)'] = inducer_vols
        self.dilutions['Water volume (uL)'] = water_vols

        if file_name is not None:
            # Autogenerate sheet name if necessary
            if sheet_name is None:
                sheet_name = self.name
            # Generate full file name
            full_file_name = os.path.join(path, file_name)
            # Generate pandas writer
            writer = pandas.ExcelWriter(full_file_name, engine='openpyxl')
            # If file already exists, open and copy all previous data
            if os.path.isfile(full_file_name):
                workbook = openpyxl.load_workbook(full_file_name)
                writer.book = workbook
                writer.sheets = dict((ws.title, ws)
                                     for ws in workbook.worksheets)
            # Convert dilutions table to excel sheet
            self.dilutions.to_excel(writer, sheet_name=sheet_name)
            # Add message about aliquots
            if self.aliquot_vol is not None:
                message = "Distribute in aliquots of {}uL." \
                    .format(self.aliquot_vol)
                worksheet = writer.sheets[sheet_name]
                worksheet.cell(row=len(self.dilutions) + 3,
                               column=1,
                               value=message)
            # Actually save
            writer.save()

    def generate_shufflings(self, n_replicates):
        # Generate n_replicates random shufflings
        self.rand_idx = []
        for i in range(n_replicates):
            l = range(len(self.dilutions))
            random.shuffle(l)
            self.rand_idx.append(l)
        # Save original dilutions table
        self.dilutions_original = self.dilutions.copy()

    def randomize(self, replicate=-1):
        # If rand_idx has not been initialized, randomization is not
        # possible. Return.
        if self.rand_idx is None:
            return
        # replicate == -1 indicates to restore the table to its original form
        if replicate==-1:
            self.dilutions = self.dilutions_original.copy()
        else:
            self.dilutions = self.dilutions_original.\
                iloc[self.rand_idx[replicate]]

    def split(self, n_splits):
        # Trivial case: one split
        if n_splits==1:
            return self

        # Initialize list to accumulat splits
        splits = []
        # length of dilutions table per split
        split_length = len(self.dilutions)/n_splits
        for i in range(n_splits):
            # Copy inducer object
            split = copy.deepcopy(self)
            # Slice dilutions table
            split.dilutions = \
                split.dilutions.iloc[i*split_length:(i+1)*split_length]
            # Modify id offset
            split.id_offset += i*split_length
            # Accumulate
            splits.append(split)

        return splits

class Plate(object):
    """
    """
    def __init__(self,
                 name,
                 nrows=4,
                 ncols=6,
                 id_prefix='S',
                 id_offset=0,
                 n_samples=None):
        # Store name
        self.name = name
        # Store dimensions
        self.nrows=nrows
        self.ncols=ncols
        # Store number of samples to measure
        if n_samples is None:
            self.n_samples = nrows*ncols
        else:
            self.n_samples = n_samples
        # Store ID modifiers for samples table
        self.id_prefix=id_prefix
        self.id_offset=id_offset

        # Initialize metadata structure: a list of 2-element tuples with column-
        # value pairs
        self.metadata=[]

        # Reset samples table
        self.reset()

    def reset(self):
        # Initialize dataframe with sample IDs
        ids = ['{}{:03d}'.format(self.id_prefix, i)
               for i in range(self.id_offset + 1,
                              self.nrows*self.ncols + self.id_offset + 1)]
        self.samples = pandas.DataFrame({'ID': ids})
        self.samples.set_index('ID', inplace=True)
        # Add metadata
        for m in self.metadata:
            self.samples[m[0]] = m[1]
        # Add plate name
        self.samples['Plate'] = self.name
        # Add row and column numbers
        self.samples['Row'] = numpy.nan
        self.samples['Column'] = numpy.nan
        for i in range(self.nrows):
            for j in range(self.ncols):
                self.samples.set_value(self.samples.index[i*self.ncols + j],
                                       'Row',
                                       i + 1)
                self.samples.set_value(self.samples.index[i*self.ncols + j],
                                       'Column',
                                       j + 1)
        # Only preserve table rows that will be measured
        self.samples = self.samples.iloc[:self.n_samples]
        # Initialize layout map
        self.layout = []
        # to_apply contains dictionaries with {inducer, across}
        self.to_apply = []

    def n_samples_to_apply(self, across='wells'):
        # Check "across" input
        if across not in ['rows', 'cols', 'wells', 'all']:
            raise ValueError('"{}"" not recognized'.format(across))
        # Only "wells" or "all" supported if not measuring full plate
        if (self.n_samples != self.ncols*self.nrows) and\
                (across not in ['wells', 'all']):
            raise ValueError('"{}"" not possible for a non-full plate'.\
                format(across))

        # Calculate number of samples to apply
        if across=='rows':
            n_samples = self.nrows
        elif across=='cols':
            n_samples = self.ncols
        elif across=='wells':
            n_samples = 1
        elif across=='all':
            n_samples = self.n_samples

        return n_samples

    def apply_inducer(self, inducer, across='wells'):
        """
        Apply the inducer to either rows, columns, or individual wells.

        Parameters
        ----------
        inducer : Inducer object
            The inducer to apply to the plate.

        across : str
            "rows" applies the specified inducer to all rows equally.
            "cols" applies to all columns equally. "wells" applies to each
            well individually. 'all' applies inducer to all wells, and
            should be inoculated in the media directly.

        """
        # Check "across" input
        if across not in ['rows', 'cols', 'wells', 'all']:
            raise ValueError('"{}"" not recognized'.format(across))
        # Only "wells" or "all" supported if not measuring full plate
        if (self.n_samples != self.ncols*self.nrows) and\
                (across not in ['wells', 'all']):
            raise ValueError('"{}"" not possible for a non-full plate'.\
                format(across))

        # Check that the inducer has the appropriate number of samples
        nwells = self.n_samples
        if (across=='rows' and (len(inducer.conc)!=self.ncols)) or \
           (across=='cols' and (len(inducer.conc)!=self.nrows)) or \
           (across=='wells' and (len(inducer.conc)!=nwells)) or \
           (across=='all' and (len(inducer.conc)!=1)):
                raise ValueError('inducer does not have the appropriate' + \
                    ' number of samples')

        # Store inducer to apply
        self.to_apply.append({'inducer': inducer,
                              'across': across})

    def generate_pipetting_layout(self,
                                  path='.',
                                  file_name=None,
                                  sheet_name=None):
        # Separate inducers based on "across" attribute
        inducers_rows = [a['inducer']
                         for a in self.to_apply
                         if a['across']=='rows']
        inducers_cols = [a['inducer']
                         for a in self.to_apply
                         if a['across']=='cols']
        inducers_wells = [a['inducer']
                         for a in self.to_apply
                          if a['across']=='wells']
        inducers_all = [a['inducer']
                         for a in self.to_apply
                        if a['across']=='all']

        # Only "wells" or "all" supported if not measuring full plate
        if (self.n_samples != self.ncols*self.nrows) and inducers_rows:
            raise ValueError('"rows"" not possible for a non-full plate')
        if (self.n_samples != self.ncols*self.nrows) and inducers_cols:
            raise ValueError('"cols"" not possible for a non-full plate')

        # Make layout table
        layout = []
        for i in range(self.nrows + len(inducers_rows)):
            layout.append(['']*(self.ncols + len(inducers_cols)))

        # Add well coordinates
        for i in range(self.nrows):
            for j in range(self.ncols):
                layout[i + len(inducers_rows)][j + len(inducers_cols)] += \
                    "({}, {})".format(i + 1, j + 1)
        # Add row inducer information
        for i, inducer in enumerate(inducers_rows):
            for j, val in enumerate(inducer.dilutions.index):
                layout[i][j + len(inducers_cols)] = str(val)
        # Add column inducer information
        for j, inducer in enumerate(inducers_cols):
            for i, val in enumerate(inducer.dilutions.index):
                layout[i + len(inducers_rows)][j] = str(val)
        # Add individual well inducer information
        for k, inducer in enumerate(inducers_wells):
            for i in range(self.nrows):
                for j in range(self.ncols):
                    if (i*self.ncols + j) >= self.n_samples:
                        break
                    layout[i + len(inducers_rows)][j + len(inducers_cols)] += \
                        "\n{}".format(inducer.dilutions.index[i*self.ncols + j])

        if file_name is not None:
            # Autogenerate sheet name if necessary
            if sheet_name is None:
                sheet_name = "Layout {}".format(self.name)
            # Generate file name
            full_file_name = os.path.join(path, file_name)
            # If file already exists, open
            if os.path.isfile(full_file_name):
                workbook = openpyxl.load_workbook(full_file_name)
            else:
                workbook = openpyxl.Workbook()
                # Remove sheet created by default
                worksheet = workbook.get_sheet_by_name(
                    workbook.get_sheet_names()[0])
                workbook.remove_sheet(worksheet)
            # Create cell styles for the plate area
            plate_fill = openpyxl.styles.PatternFill(fill_type='solid',
                                                     start_color='FFDCE6F1',
                                                     end_color='FFDCE6F1')
            plate_alignment = openpyxl.styles.Alignment(wrapText=True,
                                                        horizontal='center')
            plate_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin', color='FF4F81BD'), 
                right=openpyxl.styles.Side(style='thin', color='FF4F81BD'), 
                top=openpyxl.styles.Side(style='thin', color='FF4F81BD'), 
                bottom=openpyxl.styles.Side(style='thin', color='FF4F81BD'))

            # Create and populate worksheet
            worksheet = workbook.create_sheet(title=sheet_name)
            for i in range(self.nrows + len(inducers_rows)):
                for j in range(self.ncols + len(inducers_cols)):
                    cell = worksheet.cell(row=i + 1,
                                          column=j + 1)
                    # Write value
                    cell.value = layout[i][j]
                    # Apply styles
                    if (i >= len(inducers_rows)) and (j >= len(inducers_cols)):
                        cell.fill = plate_fill
                        cell.border = plate_border
                        cell.alignment = plate_alignment
            # Save
            workbook.save(filename=full_file_name)

    def generate_samples_table(self,
                               path='.',
                               file_name=None,
                               sheet_name=None):
        for a in self.to_apply:
            # Only "wells" or "all" supported if not measuring full plate
            if (self.n_samples != self.ncols*self.nrows) and\
                    (a['across'] not in ['wells', 'all']):
                raise ValueError('"{}"" not possible for a non-full plate'.\
                    format(a['across']))

            # Create new empty columns
            title_id = '{} ID'.format(a['inducer'].name)
            title_conc = '{} ({})'.format(a['inducer'].name, a['inducer'].units)
            self.samples[title_id] = ''
            self.samples[title_conc] = numpy.nan

            # Fill well info
            if a['across']=='rows':
                for i in range(self.nrows):
                    for j in range(self.ncols):
                        self.samples.set_value(self.samples.index[i*self.ncols + j],
                                             title_id,
                                             a['inducer'].dilutions.index[j])
                        self.samples.set_value(self.samples.index[i*self.ncols + j],
                                             title_conc,
                                             a['inducer'].conc[j])
            elif a['across']=='cols':
                for i in range(self.nrows):
                    for j in range(self.ncols):
                        self.samples.set_value(self.samples.index[i*self.ncols + j],
                                             title_id,
                                             a['inducer'].dilutions.index[i])
                        self.samples.set_value(self.samples.index[i*self.ncols + j],
                                             title_conc,
                                             a['inducer'].conc[i])
            elif a['across']=='wells':
                self.samples[title_id] = a['inducer'].dilutions.index
                self.samples[title_conc] = a['inducer'].conc

        if file_name is not None:
            # Autogenerate sheet name if necessary
            if sheet_name is None:
                sheet_name = "Samples {}".format(self.name)
            # Get full file name
            full_file_name = os.path.join(path, file_name)
            # Generate pandas writer
            writer = pandas.ExcelWriter(full_file_name, engine='openpyxl')
            # If file already exists, open and copy all previous data
            if os.path.isfile(full_file_name):
                workbook = openpyxl.load_workbook(full_file_name)
                writer.book = workbook
                writer.sheets = dict((ws.title, ws)
                                     for ws in workbook.worksheets)
            # Convert table to excel sheet
            self.samples.to_excel(writer, sheet_name=sheet_name)
            # Actually save
            writer.save()

class PlateArray(object):
    """
    """
    def __init__(self,
                 name,
                 plates,
                 id_prefix='S',
                 id_offset=0):
        # Store name
        self.name = name
        # Store plates
        self.plates = plates
        # plates is a 2D list of Plate objects
        # Check that all plate rows have the same number of plates
        nplates_row = len(plates[0])
        for plate_row in plates:
            if len(plate_row)!=nplates_row:
                raise ValueError("all rows of plates should have the same " + \
                    "number of plates")
        # Check that all plates have the same dimensions
        nrows_plate = plates[0][0].nrows
        ncols_plate = plates[0][0].ncols
        for plate_row in plates:
            for plate in plate_row:
                if plate.nrows!=nrows_plate or plate.ncols!=ncols_plate:
                    raise ValueError("all plates should have the same " + \
                        "dimensions")
        # Dimensions are properties extracted from self.plates
        
        # Store ID modifiers for samples table
        self.id_prefix=id_prefix
        self.id_offset=id_offset

        # Initialize metadata structure: a list of 2-element tuples with column-
        # value pairs
        self.metadata=[]

        # Reset samples table
        self.reset()

    @property
    def array_nrows(self):
        return len(self.plates)

    @property
    def array_ncols(self):
        return len(self.plates[0])

    @property
    def nrows(self):
        return self.array_nrows*self.plates[0][0].nrows

    @property
    def ncols(self):
        return self.array_ncols*self.plates[0][0].ncols

    @property
    def id_prefix(self):
        return self._id_prefix

    @id_prefix.setter
    def id_prefix(self, value):
        # Set own property
        self._id_prefix = value
        # Propagate to plates
        for plate_row in self.plates:
            for plate in plate_row:
                plate.id_prefix = value

    @property
    def id_offset(self):
        return self._id_offset

    @id_offset.setter
    def id_offset(self, value):
        id_offset = value
        # Set own property
        self._id_offset = id_offset
        # Propagate to plates
        for i, plate_row in enumerate(self.plates):
            for j, plate in enumerate(plate_row):
                plate.id_offset = id_offset
                id_offset += plate.nrows*plate.ncols

    def reset(self):
        # Reset all plates
        for plate_row in self.plates:
            for plate in plate_row:
                plate.reset()
        # Build samples table from tables in plates
        self.update_samples_table()

        # Initialize layout map
        self.layout = []
        # to_apply contains dictionaries with {inducer, across}
        self.to_apply = []

    def update_samples_table(self):
        # Initialize table 
        self.samples = pandas.DataFrame()
        # Initializa list of columns
        # This is necessary to maintain the order of the columns
        samples_columns = []
        # Join all samples tables
        for plate_row in self.plates:
            for plate in plate_row:
                # Add columns to column list
                for col in plate.samples.columns:
                    if col not in samples_columns:
                        samples_columns.append(col)
                # Append samples
                self.samples = self.samples.append(plate.samples)
        # Sort columns
        self.samples = self.samples[samples_columns]

    def n_samples_to_apply(self, across='wells'):
        # Check "across" input
        if across not in ['rows', 'cols', 'wells', 'all']:
            raise ValueError('"{}"" not recognized'.format(across))
        if across in ['wells', 'all']:
            raise NotImplementedError

        # Calculate number of samples to apply
        if across=='rows':
            n_samples = self.nrows
        elif across=='cols':
            n_samples = self.ncols
        elif across=='wells':
            n_samples = 1
        elif across=='all':
            n_samples = self.nrows*self.ncols

        return n_samples

    def apply_inducer(self, inducer, across='wells'):
        """
        Apply the inducer to either rows, columns, or individual wells.

        Parameters
        ----------
        inducer : Inducer object
            The inducer to apply to the plate.

        across : str
            "rows" applies the specified inducer to all rows equally.
            "cols" applies to all columns equally. "wells" applies to each
            well individually. 'all' applies inducer to all wells, and
            should be inoculated in the media directly.

        """
        # Check "across" input
        if across not in ['rows', 'cols', 'wells', 'all']:
            raise ValueError('"{}"" not recognized'.format(across))
        if across in ['wells', 'all']:
            raise NotImplementedError

        # Check that the inducer has the appropriate number of samples
        nwells = self.nrows*self.ncols
        if (across=='rows' and (len(inducer.conc)!=self.ncols)) or \
           (across=='cols' and (len(inducer.conc)!=self.nrows)) or \
           (across=='wells' and (len(inducer.conc)!=nwells)) or \
           (across=='all' and (len(inducer.conc)!=1)):
                raise ValueError('inducer does not have the appropriate' + \
                    ' number of samples')

        # Split inducer and apply to individual plates
        if across=='rows':
            inducer_splits = inducer.split(self.array_ncols)
            for i, plate_row in enumerate(self.plates):
                for j, plate in enumerate(plate_row):
                    plate.apply_inducer(inducer_splits[j], across)
        elif across=='cols':
            inducer_splits = inducer.split(self.array_nrows)
            for i, plate_row in enumerate(self.plates):
                for j, plate in enumerate(plate_row):
                    plate.apply_inducer(inducer_splits[i], across)
        elif across=='wells':
            inducer_splits = inducer.split(self.array_nrows*self.array_ncols)
            for i, plate_row in enumerate(self.plates):
                for j, plate in enumerate(plate_row):
                    plate.apply_inducer(inducer_splits[i*self.array_ncols + j],
                                        across)
        elif across=='all':
            raise NotImplementedError


        # Store inducer to apply
        self.to_apply.append({'inducer': inducer,
                              'across': across})

    def generate_pipetting_layout(self,
                                  path='.',
                                  file_name=None,
                                  sheet_name=None):
        # Separate inducers based on "across" attribute
        inducers_rows = [a['inducer']
                         for a in self.to_apply
                         if a['across']=='rows']
        inducers_cols = [a['inducer']
                         for a in self.to_apply
                         if a['across']=='cols']
        inducers_wells = [a['inducer']
                         for a in self.to_apply
                          if a['across']=='wells']
        inducers_all = [a['inducer']
                         for a in self.to_apply
                        if a['across']=='all']

        # Make layout table
        layout = []
        for i in range(self.nrows + len(inducers_rows)):
            layout.append(['']*(self.ncols + len(inducers_cols)))

        # Add well coordinates
        for i in range(self.nrows):
            for j in range(self.ncols):
                array_i = i/(self.nrows/self.array_nrows)
                array_j = j/(self.ncols/self.array_ncols)
                plate_i = i%(self.nrows/self.array_nrows)
                plate_j = j%(self.ncols/self.array_ncols)
                layout[i + len(inducers_rows)][j + len(inducers_cols)] += \
                    "{} ({}, {})".format(self.plates[array_i][array_j].name,
                                       plate_i + 1,
                                       plate_j + 1)
        # Add row inducer information
        for i, inducer in enumerate(inducers_rows):
            for j, val in enumerate(inducer.dilutions.index):
                layout[i][j + len(inducers_cols)] = str(val)
        # Add column inducer information
        for j, inducer in enumerate(inducers_cols):
            for i, val in enumerate(inducer.dilutions.index):
                layout[i + len(inducers_rows)][j] = str(val)
        # Add individual well inducer information
        for k, inducer in enumerate(inducers_wells):
            for i in range(self.nrows):
                for j in range(self.ncols):
                    layout[i + len(inducers_rows)][j + len(inducers_cols)] += \
                        "\n{}".format(inducer.dilutions.index[i*self.ncols + j])

        if file_name is not None:
            # Autogenerate sheet name if necessary
            if sheet_name is None:
                sheet_name = "Layout {}".format(self.name)
            # Generate file name
            full_file_name = os.path.join(path, file_name)
            # If file already exists, open
            if os.path.isfile(full_file_name):
                workbook = openpyxl.load_workbook(full_file_name)
            else:
                workbook = openpyxl.Workbook()
                # Remove sheet created by default
                worksheet = workbook.get_sheet_by_name(
                    workbook.get_sheet_names()[0])
                workbook.remove_sheet(worksheet)
            # Create cell styles for the plate area
            plate_fill = [openpyxl.styles.PatternFill(fill_type='solid',
                                                      start_color='FFDCE6F1',
                                                      end_color='FFDCE6F1'),
                          openpyxl.styles.PatternFill(fill_type='solid',
                                                      start_color='FFB8CCE4',
                                                      end_color='FFB8CCE4'),
                          ]
            plate_alignment = openpyxl.styles.Alignment(wrapText=True,
                                                        horizontal='center')
            plate_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin', color='FF4F81BD'), 
                right=openpyxl.styles.Side(style='thin', color='FF4F81BD'), 
                top=openpyxl.styles.Side(style='thin', color='FF4F81BD'), 
                bottom=openpyxl.styles.Side(style='thin', color='FF4F81BD'))

            # Create and populate worksheet
            worksheet = workbook.create_sheet(title=sheet_name)
            for i in range(self.nrows + len(inducers_rows)):
                for j in range(self.ncols + len(inducers_cols)):
                    cell = worksheet.cell(row=i + 1,
                                          column=j + 1)
                    # Write value
                    cell.value = layout[i][j]
                    # Apply styles
                    if (i >= len(inducers_rows)) and (j >= len(inducers_cols)):
                        array_i = (i - len(inducers_rows)) / \
                            (self.nrows/self.array_nrows)
                        array_j = (j - len(inducers_cols)) / \
                            (self.ncols/self.array_ncols)
                        cell.fill = plate_fill[(array_i + array_j)%2]
                        cell.border = plate_border
                        cell.alignment = plate_alignment

            # Save
            workbook.save(filename=full_file_name)

    def generate_samples_table(self,
                               path='.',
                               file_name=None,
                               sheet_name=None):

        # Generate tables on individual plates
        for plate_row in self.plates:
            for plate in plate_row:
                plate.generate_samples_table()
        # Update samples table from individual plates' tables
        self.update_samples_table()

        if file_name is not None:
            # Autogenerate sheet name if necessary
            if sheet_name is None:
                sheet_name = "Samples {}".format(self.name)
            # Get full file name
            full_file_name = os.path.join(path, file_name)
            # Generate pandas writer
            writer = pandas.ExcelWriter(full_file_name, engine='openpyxl')
            # If file already exists, open and copy all previous data
            if os.path.isfile(full_file_name):
                workbook = openpyxl.load_workbook(full_file_name)
                writer.book = workbook
                writer.sheets = dict((ws.title, ws)
                                     for ws in workbook.worksheets)
            # Convert table to excel sheet
            self.samples.to_excel(writer, sheet_name=sheet_name)
            # Actually save
            writer.save()