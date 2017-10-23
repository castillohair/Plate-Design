# -*- coding: UTF-8 -*-
"""
Module that contains plate classes.

"""

import collections

import numpy
import openpyxl
import pandas

# Create cell styles for the plate area
plate_fill = [openpyxl.styles.PatternFill(fill_type='solid',
                                          start_color='FFDCE6F1',
                                          end_color='FFDCE6F1'),
              openpyxl.styles.PatternFill(fill_type='solid',
                                          start_color='FFB8CCE4',
                                          end_color='FFB8CCE4'),
              ]
plate_alignment = openpyxl.styles.Alignment(wrapText=True,
                                            horizontal='center')
plate_border = openpyxl.styles.Border(
    left=openpyxl.styles.Side(style='thin', color='FF4F81BD'), 
    right=openpyxl.styles.Side(style='thin', color='FF4F81BD'), 
    top=openpyxl.styles.Side(style='thin', color='FF4F81BD'), 
    bottom=openpyxl.styles.Side(style='thin', color='FF4F81BD'))
header_alignment = openpyxl.styles.Alignment(wrapText=True,
                                            horizontal='center')
header_font = openpyxl.styles.Font(bold=True, color="FF1F497D")

class Plate(object):
    """
    Object that represents a plate.

    This class manages inducers that can be applied to individual rows,
    columns, wells, or to all the media in the plate. These are used to
    generate instructions for cell and inducer inoculation during the
    Replicate Setup stage. Finally, this class can generate a ClosedPlate
    instance, which contains final inducer concentrations, initial cell
    densities, and other information about each well in the plate.

    Parameters
    ----------
    name : str
        Name of the plate, to be used in generated files.
    n_rows, n_cols : int, optional
        Number of rows and columns in the plate. Defaults: 4 and 6.

    Attributes
    ----------
    name : str
        Name of the plate, to be used in generated files.
    n_rows, n_cols : int
        Number of rows and columns in the plate.
    samples_to_measure : int
        Number of samples to be measured.
    sample_media_vol : float
        Volume of media per sample (well).
    total_media_vol : float
        Starting total volume of media, to be distributed into wells.
    cell_strain_name : str
        Name of the cell strain to be inoculated in this plate.
    cell_setup_method : str or None
        Method used to determine how much volume of cells to inoculate. Can
        be one of the following: "fixed_od600", "fixed_volume", or
        "fixed_dilution".
    cell_predilution : float
        Dilution factor for the cell preculture before inoculating.
    cell_predilution_vol : float
        Volume of diluted preculture to make in µL.
    cell_initial_od600 : float
        Target initial OD600 for inoculating cells. Only used if
        `cell_setup_method` is "fixed_od600".
    cell_shot_vol : float
        Volume of diluted preculture to inoculate in media. Only used if
        `cell_setup_method` is "fixed_volume".
    cell_total_dilution : float
        Total dilution from preculture to be inoculated in the media. Only
        used if `cell_setup_method` is "fixed_dilution".
    metadata : OrderedDict
        Additional information about the plate, in a ``key: value`` format.
        The ClosedPlate instance returned by ``close_plates()`` will
        include this information in its ``samples_table`` attribute. In it,
        a column with name ``key`` will be created, and all rows will be
        set to ``value``.
    inducers : OrderedDict
        Keys in this dictionary represent how each inducer is applied
        ("rows", "cols", "wells", "media"), and the values are lists of
        inducers to be applied as specified by the key.

    Methods
    -------
    apply_inducer_media_vol
        Get the media volume to which an inducer will be applied.
    apply_inducer_n_shots
        Get number of samples that each inducer dose will be applied to.
    apply_inducer
        Apply an inducer to the plate.
    save_exp_setup_instructions
        Calculate and save instructions for the Experiment Setup stage.
    save_exp_setup_files
        Save additional files required for the Experiment Setup stage.
    save_rep_setup_instructions
        Calculate and save instructions for the Replicate Setup stage.
    add_inducer_setup_instructions
        Add sheet with inducer pipetting instructions to specified workbook.
    add_cell_setup_instructions
        Add sheet with cell inoculation instructions to specified workbook.
    save_rep_setup_files
        Save additional files required for the Replicate Setup stage.
    close_plates
        Generate a ``ClosedPlate`` object using this plate's information.

    """
    def __init__(self,
                 name,
                 n_rows=4,
                 n_cols=6):
        # Store name
        self.name = name
        # Store dimensions
        self.n_rows = n_rows
        self.n_cols = n_cols

        # Measure all samples by default
        self.samples_to_measure = n_rows*n_cols
        # Initialize sample and total media volumes
        self.sample_media_vol = None
        self.total_media_vol = None

        # Initialize parameters for cell setup
        self.cell_strain_name = None
        self.cell_setup_method = None
        self.cell_predilution = 1
        self.cell_predilution_vol = None
        self.cell_initial_od600 = None
        self.cell_shot_vol = None
        self.cell_total_dilution = None

        # Initialize metadata dictionary
        self.metadata = collections.OrderedDict()

        # Initialize list of inducers
        self.inducers = {'rows': [], 'cols': [], 'wells': [], 'media': []}

    def apply_inducer_media_vol(self, apply_to):
        """
        Get the media volume to which an inducer will be applied.

        Parameters
        ----------
        apply_to : {'rows', 'cols', 'wells', 'media'}
            "rows" applies the specified inducer to all rows equally.
            "cols" applies to all columns equally. "wells" applies to each
            well individually. 'media' applies inducer to the media at the
            beginning of the replicate setup stage.

        """
        # Check for the presence of required attributes
        if self.sample_media_vol is None:
            raise AttributeError("sample_media_vol should be set")
        if self.total_media_vol is None:
            raise AttributeError("total_media_vol should be set")

        # Check "apply_to" input
        if apply_to not in ['rows', 'cols', 'wells', 'media']:
            raise ValueError('"{}"" not recognized'.format(apply_to))
        # Only "wells" or "media" supported if not measuring full plate
        if (self.samples_to_measure != self.n_cols*self.n_rows) and\
                (apply_to not in ['wells', 'media']):
            raise ValueError('"{}" not possible if not measuring all wells'.\
                format(apply_to))

        # Calculate number of samples to apply
        if apply_to in ['rows', 'cols', 'wells']:
            return self.sample_media_vol
        elif apply_to=='media':
            return self.total_media_vol

    def apply_inducer_n_shots(self, apply_to):
        """
        Get number of samples that each inducer dose will be applied to.

        Parameters
        ----------
        apply_to : {'rows', 'cols', 'wells', 'media'}
            "rows" applies the specified inducer to all rows equally.
            "cols" applies to all columns equally. "wells" applies to each
            well individually. 'media' applies inducer to the media at the
            beginning of the replicate setup stage.

        """
        # Check "apply_to" input
        if apply_to not in ['rows', 'cols', 'wells', 'media']:
            raise ValueError('"{}"" not recognized'.format(apply_to))
        # Only "wells" or "media" supported if not measuring full plate
        if (self.samples_to_measure != self.n_cols*self.n_rows) and\
                (apply_to not in ['wells', 'media']):
            raise ValueError('"{}" not possible if not measuring all wells'.\
                format(apply_to))

        # Calculate number of samples to apply
        if apply_to=='rows':
            return self.n_rows
        elif apply_to=='cols':
            return self.n_cols
        elif apply_to=='wells':
            return 1
        elif apply_to=='media':
            return 1

    def apply_inducer(self, inducer, apply_to='wells'):
        """
        Apply an inducer to the plate.

        This function stores the specified inducer in the `inducers`
        attribute, after verifying consistency.

        Parameters
        ----------
        inducer : Inducer object
            The inducer to apply to the plate.
        apply_to : {'rows', 'cols', 'wells', 'media'}
            "rows" applies the specified inducer to all rows equally.
            "cols" applies to all columns equally. "wells" applies to each
            well individually. 'media' applies inducer to the media at the
            beginning of the replicate setup stage.

        """
        # Check "apply_to" input
        if apply_to not in ['rows', 'cols', 'wells', 'media']:
            raise ValueError('"{}"" not recognized'.format(apply_to))
        # Only "wells" or "media" supported if not measuring full plate
        if (self.samples_to_measure != self.n_cols*self.n_rows) and\
                (apply_to not in ['wells', 'media']):
            raise ValueError('"{}"" not possible if not measuring all wells'.\
                format(apply_to))

        # Check that the inducer has the appropriate number of samples
        n_wells = self.samples_to_measure
        if (apply_to=='rows' and (len(inducer.doses_table)!=self.n_cols)) or \
           (apply_to=='cols' and (len(inducer.doses_table)!=self.n_rows)) or \
           (apply_to=='wells' and (len(inducer.doses_table)!=n_wells)) or \
           (apply_to=='media' and (len(inducer.doses_table)!=1)):
                raise ValueError('inducer does not have the appropriate' + \
                    ' number of doses')

        # Check that the inducer is not repeated
        if inducer in self.inducers[apply_to]:
            raise ValueError("inducer already in plate's inducer list")

        # Store inducer
        self.inducers[apply_to].append(inducer)

    def save_exp_setup_instructions(self, file_name=None, workbook=None):
        """
        Calculate and save instructions for the Experiment Setup stage.

        Parameters
        ----------
        file_name : str, optional
            Name of the Excel file to save. Either this or `workbook`
            should not be None.
        workbook : Workbook, optional
            If not None, `file_name` is ignored, and a sheet with the
            instructions is directly added to workbook `workbook`. Either
            this or `file_name` should not be None.

        """
        if (file_name is None) and (workbook is None):
            raise ValueError('either file_name or workbook should be specified')

        pass

    def save_exp_setup_files(self, path='.'):
        """
        Save additional files required for the Experiment Setup stage.

        Parameters
        ----------
        path : str
            Folder in which to save files.

        """
        pass

    def save_rep_setup_instructions(self, file_name=None, workbook=None):
        """
        Calculate and save instructions for the Replicate Setup stage.

        Two sheets are generated, named "Inducers for Plate [plate name]"
        and "Cells for Plate [plate name]", with instructions on how to add
        inducers and cells to the plates, respectively.

        Parameters
        ----------
        file_name : str, optional
            Name of the Excel file to save.
        workbook : Workbook, optional
            If not None, `file_name` is ignored, and a sheet with the
            instructions is directly added to workbook `workbook`.

        """
        if (file_name is None) and (workbook is None):
            raise ValueError('either file_name or workbook should be specified')

        # Create workbook if not provided
        if workbook is None:
            # Create and remove empty sheet created by default
            workbook = openpyxl.Workbook()
            workbook.remove_sheet(workbook.active)
            save_workbook = True
        else:
            save_workbook = False

        # Add sheets
        self.add_inducer_setup_instructions(
            workbook,
            "Inducers for Plate {}".format(self.name))
        self.add_cell_setup_instructions(
            workbook,
            "Cells for Plate {}".format(self.name))

        # Save
        if save_workbook:
            if not workbook.sheetnames:
                workbook.create_sheet("Sheet 1")
            workbook.save(filename=file_name)

    def add_inducer_setup_instructions(self, workbook, sheet_name):
        """
        Add sheet with inducer pipetting instructions to specified workbook.

        Only add instructions for an inducer if it has the attribute
        ``shot_vol``.

        Parameters
        ----------
        workbook : Workbook
            Workbook where to add sheet.
        sheet_name : str
            Name to give to the new sheet.

        """
        # Check that a sheet with the specified name doesn't exist
        if sheet_name in [ws.title for ws in workbook.worksheets]:
            raise ValueError("sheet \"{}\"already present in workbook".\
                format(sheet_name))

        # Get only inducers that have ``shot_vol``
        inducers_rows = [ind
                         for ind in self.inducers['rows']
                         if hasattr(ind, 'shot_vol')]
        inducers_cols = [ind
                         for ind in self.inducers['cols']
                         if hasattr(ind, 'shot_vol')]
        inducers_wells = [ind
                          for ind in self.inducers['wells']
                          if hasattr(ind, 'shot_vol')]
        inducers_media = [ind
                          for ind in self.inducers['media']
                          if hasattr(ind, 'shot_vol')]

        # Skip if no inducers are present
        if not(inducers_rows or inducers_cols or \
                inducers_wells or inducers_media):
            return

        # Only "wells" or "media" supported if not measuring full plate
        if (self.samples_to_measure != self.n_cols*self.n_rows) and \
                inducers_rows:
            raise ValueError('"rows" not possible for a non-full plate')
        if (self.samples_to_measure != self.n_cols*self.n_rows) and \
                inducers_cols:
            raise ValueError('"cols" not possible for a non-full plate')

        # Initialize inducer instructions table
        ind_layout = []
        ind_layout_rows = self.n_rows + len(inducers_rows) + \
            len(inducers_rows) + len(inducers_cols) + len(inducers_wells) + \
            len(inducers_media) + 1
        ind_layout_cols = self.n_cols + len(inducers_cols)
        for i in range(ind_layout_rows):
            ind_layout.append(['']*(ind_layout_cols))

        # Add well coordinates
        for i in range(self.n_rows):
            for j in range(self.n_cols):
                row = i + len(inducers_rows)
                col = j + len(inducers_cols)
                ind_layout[row][col] += "({}, {})".format(i + 1, j + 1)
        # Add row inducer information
        for i, inducer in enumerate(inducers_rows):
            for j, val in enumerate(inducer.doses_table.index):
                row = i
                col = j + len(inducers_cols)
                ind_layout[row][col] = str(val)
        # Add column inducer information
        for j, inducer in enumerate(inducers_cols):
            for i, val in enumerate(inducer.doses_table.index):
                row = i + len(inducers_rows)
                col = j
                ind_layout[row][col] = str(val)
        # Add individual well inducer information
        for k, inducer in enumerate(inducers_wells):
            for i in range(self.n_rows):
                for j in range(self.n_cols):
                    if (i*self.n_cols + j) >= self.samples_to_measure:
                        break
                    row = i + len(inducers_rows)
                    col = j + len(inducers_cols)
                    ind_layout[row][col] += "\n{}".format(
                        inducer.doses_table.index[i*self.n_cols + j])
        # Add volume of each inducer to add
        current_row = self.n_rows + len(inducers_rows) + 1
        for inducer in inducers_rows:
            ind_layout[current_row][0] = "Add {:.2f}µL of {} to each well.".\
            format(inducer.shot_vol, inducer.name)
            current_row += 1
        for inducer in inducers_cols:
            ind_layout[current_row][0] = "Add {:.2f}µL of {} to each well.".\
            format(inducer.shot_vol, inducer.name)
            current_row += 1
        for inducer in inducers_wells:
            ind_layout[current_row][0] = "Add {:.2f}µL of {} to each well.".\
            format(inducer.shot_vol, inducer.name)
            current_row += 1
        for inducer in inducers_media:
            ind_layout[current_row][0] = "Add {:.2f}µL of {} to media.".\
            format(inducer.shot_vol, inducer.name)
            current_row += 1

        # Plate area
        plate_min_row = len(inducers_rows)
        plate_max_row = len(inducers_rows) + self.n_rows
        plate_min_col = len(inducers_cols)
        plate_max_col = len(inducers_cols) + self.n_cols

        # Create and populate worksheet
        worksheet = workbook.create_sheet(title=sheet_name)
        for i in range(len(ind_layout)):
            for j in range(len(ind_layout[i])):
                cell = worksheet.cell(row=i+1, column=j+1)
                # Write value
                cell.value = ind_layout[i][j]
                # Apply styles
                if (i >= plate_min_row) and (i < plate_max_row) and\
                        (j >= plate_min_col) and (j < plate_max_col):
                    cell.fill = plate_fill[0]
                    cell.border = plate_border
                    cell.alignment = plate_alignment

    def add_cell_setup_instructions(self, workbook, sheet_name):
        """
        Add sheet with cell inoculation instructions to specified workbook.

        Parameters
        ----------
        workbook : Workbook
            Workbook where to add sheet.
        sheet_name : str
            Name to give to the new sheet.

        """
        # Check that a sheet with the specified name doesn't exist
        if sheet_name in [ws.title for ws in workbook.worksheets]:
            raise ValueError("sheet \"{}\"already present in workbook".\
                format(sheet_name))

        # Do nothing if cell starting method has not been defined.
        if self.cell_setup_method is None:
            return

        # Create sheet
        worksheet = workbook.create_sheet(title=sheet_name)
        # Add setup instructions
        if self.cell_setup_method=='fixed_od600':
            # Check if initial od600 has been specified
            if self.cell_initial_od600 is None:
                raise ValueError("cell initial OD600 should be specified")
            # Set width
            worksheet.column_dimensions['A'].width = 22
            worksheet.column_dimensions['B'].width = 7
            worksheet.column_dimensions['C'].width = 5
            # Info about strain
            worksheet.cell(row=1, column=1).value = "Strain Name"
            worksheet.cell(row=1, column=2).value = self.cell_strain_name
            if self.cell_predilution != 1:
                # Check if predilution volume has been specified
                if self.cell_predilution_vol is None:
                    raise ValueError("cell predilution volume should be "
                        "specified")
                # Instructions for making predilution
                worksheet.cell(row=2, column=1).value = "Predilution"
                worksheet.cell(row=2, column=1).alignment = header_alignment
                worksheet.cell(row=2, column=1).font = header_font
                worksheet.merge_cells(start_row=2,
                                      end_row=2,
                                      start_column=1,
                                      end_column=3)

                worksheet.cell(row=3, column=1).value = "Predilution factor"
                worksheet.cell(row=3, column=2).value = self.cell_predilution
                worksheet.cell(row=3, column=3).value = "x"

                cell_vol = self.cell_predilution_vol / \
                    float(self.cell_predilution)
                media_vol = self.cell_predilution_vol - cell_vol
                worksheet.cell(row=4, column=1).value = "Media volume"
                worksheet.cell(row=4, column=2).value = media_vol
                worksheet.cell(row=4, column=3).value = "µL"

                worksheet.cell(row=5, column=1).value = "Preculture volume"
                worksheet.cell(row=5, column=2).value = cell_vol
                worksheet.cell(row=5, column=3).value = "µL"

                worksheet.cell(row=6, column=1).value = "Predilution OD600"
                worksheet.cell(row=6, column=2).fill = plate_fill[0]

                # Instructions for inoculating into plate media
                worksheet.cell(row=7, column=1).value = "Inoculation"
                worksheet.cell(row=7, column=1).alignment = header_alignment
                worksheet.cell(row=7, column=1).font = header_font
                worksheet.merge_cells(start_row=7,
                                      end_row=7,
                                      start_column=1,
                                      end_column=3)

                worksheet.cell(row=8, column=1).value = "Target OD600"
                worksheet.cell(row=8, column=2).value = self.cell_initial_od600

                worksheet.cell(row=9, column=1).value = "Predilution volume"
                worksheet.cell(row=9, column=2).value = "={}/B6".format(
                    self.total_media_vol*self.cell_initial_od600)
                worksheet.cell(row=9, column=3).value = "µL"

                worksheet.cell(row=10, column=1).value = \
                    "Add into {:.2f}mL media, ".format(self.total_media_vol/1000.) + \
                    "and distribute into plate wells."
            else:
                worksheet.cell(row=2, column=1).value = "Preculture OD600"
                worksheet.cell(row=2, column=2).fill = plate_fill[0]

                worksheet.cell(row=3, column=1).value = "Target OD600"
                worksheet.cell(row=3, column=2).value = self.cell_initial_od600

                worksheet.cell(row=4, column=1).value = "Preculture volume"
                worksheet.cell(row=4, column=2).value = "={}/B2".format(
                    self.total_media_vol*self.cell_initial_od600)
                worksheet.cell(row=4, column=3).value = "µL"

                worksheet.cell(row=5, column=1).value = \
                    "Add into {:.2f}mL media, ".format(self.total_media_vol/1000.) + \
                    "and distribute into plate wells."

        elif self.cell_setup_method=='fixed_volume':
            # Check if shot volume has been specified
            if self.cell_shot_vol is None:
                raise ValueError("cell shot volume should be specified")
            # Set width
            worksheet.column_dimensions['A'].width = 22
            worksheet.column_dimensions['B'].width = 7
            worksheet.column_dimensions['C'].width = 5
            # Info about strain
            worksheet.cell(row=1, column=1).value = "Strain Name"
            worksheet.cell(row=1, column=2).value = self.cell_strain_name
            if self.cell_predilution != 1:
                # Check if predilution volume has been specified
                if self.cell_predilution_vol is None:
                    raise ValueError("cell predilution volume should be "
                        "specified")
                # Instructions for making predilution
                worksheet.cell(row=2, column=1).value = "Predilution"
                worksheet.cell(row=2, column=1).alignment = header_alignment
                worksheet.cell(row=2, column=1).font = header_font
                worksheet.merge_cells(start_row=2,
                                      end_row=2,
                                      start_column=1,
                                      end_column=3)

                worksheet.cell(row=3, column=1).value = "Predilution factor"
                worksheet.cell(row=3, column=2).value = self.cell_predilution
                worksheet.cell(row=3, column=3).value = "x"

                cell_vol = self.cell_predilution_vol / \
                    float(self.cell_predilution)
                media_vol = self.cell_predilution_vol - cell_vol
                worksheet.cell(row=4, column=1).value = "Media volume"
                worksheet.cell(row=4, column=2).value = media_vol
                worksheet.cell(row=4, column=3).value = "µL"

                worksheet.cell(row=5, column=1).value = "Preculture volume"
                worksheet.cell(row=5, column=2).value = cell_vol
                worksheet.cell(row=5, column=3).value = "µL"

                # Instructions for inoculating into plate media
                worksheet.cell(row=6, column=1).value = "Inoculation"
                worksheet.cell(row=6, column=1).alignment = header_alignment
                worksheet.cell(row=6, column=1).font = header_font
                worksheet.merge_cells(start_row=6,
                                      end_row=6,
                                      start_column=1,
                                      end_column=3)

                worksheet.cell(row=7, column=1).value = "Predilution volume"
                worksheet.cell(row=7, column=2).value = self.cell_shot_vol
                worksheet.cell(row=7, column=3).value = "µL"

                worksheet.cell(row=8, column=1).value = \
                    "Add into {:.2f}mL media, ".format(self.total_media_vol/1000.) + \
                    "and distribute into plate wells."
            else:
                # Instructions for inoculating into plate media
                worksheet.cell(row=2, column=1).value = "Preculture volume"
                worksheet.cell(row=2, column=2).value = self.cell_shot_vol
                worksheet.cell(row=2, column=3).value = "µL"

                worksheet.cell(row=3, column=1).value = \
                    "Add into {:.2f}mL media, ".format(self.total_media_vol/1000.) + \
                    "and distribute into plate wells."

        elif self.cell_setup_method=='fixed_dilution':
            # Check if total dilution has been specified
            if self.cell_total_dilution is None:
                raise ValueError("cell total dilution should be specified")
            # Set width
            worksheet.column_dimensions['A'].width = 22
            worksheet.column_dimensions['B'].width = 7
            worksheet.column_dimensions['C'].width = 5
            # Info about strain
            worksheet.cell(row=1, column=1).value = "Strain Name"
            worksheet.cell(row=1, column=2).value = self.cell_strain_name
            if self.cell_predilution != 1:
                # Check if predilution volume has been specified
                if self.cell_predilution_vol is None:
                    raise ValueError("cell predilution volume should be "
                        "specified")
                # Instructions for making predilution
                worksheet.cell(row=2, column=1).value = "Predilution"
                worksheet.cell(row=2, column=1).alignment = header_alignment
                worksheet.cell(row=2, column=1).font = header_font
                worksheet.merge_cells(start_row=2,
                                      end_row=2,
                                      start_column=1,
                                      end_column=3)

                worksheet.cell(row=3, column=1).value = "Predilution factor"
                worksheet.cell(row=3, column=2).value = self.cell_predilution
                worksheet.cell(row=3, column=3).value = "x"

                cell_vol = self.cell_predilution_vol / \
                    float(self.cell_predilution)
                media_vol = self.cell_predilution_vol - cell_vol
                worksheet.cell(row=4, column=1).value = "Media volume"
                worksheet.cell(row=4, column=2).value = media_vol
                worksheet.cell(row=4, column=3).value = "µL"

                worksheet.cell(row=5, column=1).value = "Preculture volume"
                worksheet.cell(row=5, column=2).value = cell_vol
                worksheet.cell(row=5, column=3).value = "µL"

                # Instructions for inoculating into plate media
                worksheet.cell(row=6, column=1).value = "Inoculation"
                worksheet.cell(row=6, column=1).alignment = header_alignment
                worksheet.cell(row=6, column=1).font = header_font
                worksheet.merge_cells(start_row=6,
                                      end_row=6,
                                      start_column=1,
                                      end_column=3)

                cell_shot_vol = self.total_media_vol / \
                    float(self.cell_total_dilution) * self.cell_predilution
                worksheet.cell(row=7, column=1).value = "Predilution volume"
                worksheet.cell(row=7, column=2).value = cell_shot_vol
                worksheet.cell(row=7, column=3).value = "µL"

                worksheet.cell(row=8, column=1).value = \
                    "Add into {:.2f}mL media, ".format(self.total_media_vol/1000.) + \
                    "and distribute into plate wells."
            else:
                cell_shot_vol = self.total_media_vol / \
                    float(self.cell_total_dilution)
                # Instructions for inoculating into plate media
                worksheet.cell(row=2, column=1).value = "Preculture volume"
                worksheet.cell(row=2, column=2).value = cell_shot_vol
                worksheet.cell(row=2, column=3).value = "µL"

                worksheet.cell(row=3, column=1).value = \
                    "Add into {:.2f}mL media, ".format(self.total_media_vol/1000.) + \
                    "and distribute into plate wells."

        else:
            raise ValueError("cell setup method {} not recognized".format(
                self.cell_setup_method))

    def save_rep_setup_files(self, path='.'):
        """
        Save additional files required for the Replicate Setup stage.

        Parameters
        ----------
        path : str
            Folder in which to save files.

        """
        pass

    def close_plates(self):
        """
        Generate a ``ClosedPlate`` object using this plate's information.

        The individual ``ClosedPlate`` instances contain general plate
        information such as plate name, dimensions, cell inoculation
        conditions, and metadata, as well as well-specific information such
        as inducer concentrations. All this info is generated when calling
        `close_plates()`, and will remain fixed even after modifying
        inducers or other information in the ``Plate`` object.

        Within an experiment workflow, this function is meant to be called
        at the end of the Replicate Setup stage.

        Return
        ------
        list
            ``ClosedPlate`` instances with information about each sample.
            This list will only contain one ``ClosedPlate`` instance, as
            ``Plate`` represents a single plate.

        """
        # Prepare plate info
        plate_info = collections.OrderedDict()

        # Add cell info
        plate_info['Strain'] = self.cell_strain_name
        if self.cell_setup_method=='fixed_od600':
            plate_info['Preculture Dilution'] = self.cell_predilution
            plate_info['Initial OD600'] = self.cell_initial_od600
        elif self.cell_setup_method=='fixed_volume':
            plate_info['Preculture Dilution'] = self.cell_predilution
            plate_info['Cell Inoculated Vol.'] = self.cell_shot_vol
        elif self.cell_setup_method=='fixed_dilution':
            plate_info['Preculture Dilution'] = self.cell_predilution
            plate_info['Total Cell Dilution'] = self.cell_total_dilution

        # Add additional plate metadata
        # The following try-catch block is needed to ensure compatibility with
        # both python2 and python3.
        try:
            items = self.metadata.iteritems()
        except AttributeError:
            items = self.metadata.items()
        for k, v in items:
            plate_info[k] = v

        # Prepare well info
        well_info = pandas.DataFrame(index=range(self.n_rows*self.n_cols))
        # Make boolean array indicating which samples should be measured
        samples_to_measure_bool = (numpy.arange(self.n_rows*self.n_cols) < \
            self.samples_to_measure)

        # Add inducer info
        # The following try-catch block is needed to ensure compatibility with
        # both python2 and python3.
        try:
            items = self.inducers.iteritems()
        except AttributeError:
            items = self.inducers.items()
        for apply_to, inducers in items:
            for inducer in inducers:
                if apply_to=='rows':
                    for column in inducer.doses_table.columns:
                        for i in range(self.n_rows):
                            for j in range(self.n_cols):
                                well_info.set_value(
                                    well_info.index[i*self.n_cols + j],
                                    column,
                                    inducer.doses_table.iloc[j][column])
                elif apply_to=='cols':
                    for column in inducer.doses_table.columns:
                        for i in range(self.n_rows):
                            for j in range(self.n_cols):
                                well_info.set_value(
                                    well_info.index[i*self.n_cols + j],
                                    column,
                                    inducer.doses_table.iloc[i][column])
                elif apply_to=='wells':
                    for column in inducer.doses_table.columns:
                        well_info.loc[samples_to_measure_bool, column] = \
                            inducer.doses_table[column].values
                elif apply_to=='media':
                    for column in inducer.doses_table.columns:
                        well_info.loc[samples_to_measure_bool, column] = \
                            inducer.doses_table[column].values[0]

        # Add which wells should be measured
        well_info['Measure'] = samples_to_measure_bool

        # Create closed plate object
        closed_plate = ClosedPlate(name=self.name,
                                   n_rows=self.n_rows,
                                   n_cols=self.n_cols,
                                   plate_info=plate_info,
                                   well_info=well_info)

        return [closed_plate]

class PlateArray(Plate):
    """
    Object that represents an array of plates.

    This class manages inducers that can be applied to individual rows,
    columns, wells, or to all the media in the array. These are used to
    generate instructions for cell and inducer inoculation during the
    Replicate Setup stage. Finally, this class can generate ClosedPlate
    instances, one for each plate in the array, with a table containing
    fixed information about each sample such as final inducer
    concentration, useful for the Replicate Measurement stage.

    Parameters
    ----------
    name : str
        Name of the plate array, to be used in generated files.
    array_n_rows, array_n_cols : int
        Number of rows and columns in the plate array.
    plate_names : list
        Names of the plates, to be used in generated files.
    plate_n_rows, plate_n_cols : int, optional
        Number of rows and columns in each plate. Defaults: 4 and 6.

    Attributes
    ----------
    name : str
        Name of the plate array, to be used in generated files.
    array_n_rows, array_n_cols : int
        Number of rows and columns in the plate array.
    plate_names : list
        Names of the plates, to be used in generated files.
    plate_n_rows, plate_n_cols : int
        Number of rows and columns in each plate.        
    n_rows, n_cols : int
        Total number of rows and columns in the plate array.
    samples_to_measure : int
        Number of samples to be measured.
    sample_media_vol : float
        Volume of media per sample (well).
    total_media_vol : float
        Starting total volume of media, to be distributed into wells.
    cell_strain_name : str
        Name of the cell strain to be inoculated in this plate.
    cell_setup_method : str or None
        Method used to determine how much volume of cells to inoculate. Can
        be one of the following: "fixed_od600", "fixed_volume", or
        "fixed_dilution".
    cell_predilution : float
        Dilution factor for the cell preculture before inoculating.
    cell_predilution_vol : float
        Volume of diluted preculture to make in µL.
    cell_initial_od600 : float
        Target initial OD600 for inoculating cells. Only used if
        `cell_setup_method` is "fixed_od600".
    cell_shot_vol : float
        Volume of diluted preculture to inoculate in media. Only used if
        `cell_setup_method` is "fixed_volume".
    cell_total_dilution : float
        Total dilution from preculture to be inoculated in the media. Only
        used if `cell_setup_method` is "fixed_dilution".
    metadata : OrderedDict
        Additional information about the array, in a ``key: value`` format.
        ClosedPlate instances returned by ``close_plates()`` will include
        this information in their ``samples_table`` attribute. In them, a
        column with name ``key`` will be created, and all rows will be
        set to ``value``.
    inducers : OrderedDict
        Keys in this dictionary represent how each inducer is applied
        ("rows", "cols", "wells", "media"), and the values are lists of
        inducers to be applied as specified by the key.

    Methods
    -------
    apply_inducer_media_vol
        Get the media volume to which an inducer will be applied.
    apply_inducer_n_shots
        Get number of samples that each inducer dose will be applied to.
    apply_inducer
        Apply an inducer to the plate.
    save_rep_setup_instructions
        Calculate and save instructions for the Replicate Setup stage.
    add_inducer_setup_instructions
        Add sheet with inducer pipetting instructions to specified workbook.
    add_cell_setup_instructions
        Add sheet with cell inoculation instructions to specified workbook.
    close_plates
        Generate ``ClosedPlate`` objects for each plate in the array.

    """
    def __init__(self,
                 name,
                 array_n_rows,
                 array_n_cols,
                 plate_names,
                 plate_n_rows=4,
                 plate_n_cols=6):
        # Store name
        self.name = name
        # Store dimensions
        self.array_n_rows = array_n_rows
        self.array_n_cols = array_n_cols
        self.plate_n_rows = plate_n_rows
        self.plate_n_cols = plate_n_cols
        # Verify that we have the appropriate amount of plate names, and store
        if len(plate_names) != array_n_rows*array_n_cols:
            raise ValueError("plate_names should have {} elements".format(
                array_n_rows*array_n_cols))
        self.plate_names = plate_names

        # Measure all samples by default
        self.samples_to_measure = self.n_rows*self.n_cols
        # Initialize sample and total media volumes
        self.sample_media_vol = None
        self.total_media_vol = None

        # Initialize parameters for cell setup
        self.cell_strain_name = None
        self.cell_setup_method = None
        self.cell_predilution = 1
        self.cell_predilution_vol = None
        self.cell_initial_od600 = None
        self.cell_shot_vol = None
        self.cell_total_dilution = None

        # Initialize metadata dictionary
        self.metadata = collections.OrderedDict()

        # Initialize list of inducers
        self.inducers = {'rows': [], 'cols': [], 'wells': [], 'media': []}

    @property
    def n_rows(self):
        """
        Total number of rows in the plate array.

        """
        return self.array_n_rows*self.plate_n_rows

    @property
    def n_cols(self):
        """
        Total number of columns in the plate array.
        
        """
        return self.array_n_cols*self.plate_n_cols

    def save_rep_setup_instructions(self, file_name=None, workbook=None):
        """
        Calculate and save instructions for the Replicate Setup stage.

        Two sheets are generated, named "Inducers for Plate Array [plate
        array name]" and "Cells for Plate Array [plate array name]", with
        instructions on how to add inducers and cells to the plates,
        respectively.

        Parameters
        ----------
        file_name : str, optional
            Name of the Excel file to save.
        workbook : Workbook, optional
            If not None, `file_name` is ignored, and a sheet with the
            instructions is directly added to workbook `workbook`.

        """
        if (file_name is None) and (workbook is None):
            raise ValueError('either file_name or workbook should be specified')

        # Create workbook if not provided
        if workbook is None:
            # Create and remove empty sheet created by default
            workbook = openpyxl.Workbook()
            workbook.remove_sheet(workbook.active)
            save_workbook = True
        else:
            save_workbook = False

        # Add sheets
        self.add_inducer_setup_instructions(
            workbook,
            "Inducers for Plate Array {}".format(self.name))
        self.add_cell_setup_instructions(
            workbook,
            "Cells for Plate Array {}".format(self.name))

        # Save
        if save_workbook:
            if not workbook.sheetnames:
                workbook.create_sheet("Sheet 1")
            workbook.save(filename=file_name)

    def add_inducer_setup_instructions(self, workbook, sheet_name):
        """
        Add sheet with inducer pipetting instructions to specified workbook.

        Only add instructions for an inducer if it has the attribute
        ``shot_vol``.

        Parameters
        ----------
        workbook : Workbook
            Workbook where to add sheet.
        sheet_name : str
            Name to give to the new sheet.

        """
        # Check that a sheet with the specified name doesn't exist
        if sheet_name in [ws.title for ws in workbook.worksheets]:
            raise ValueError("sheet \"{}\"already present in workbook".\
                format(sheet_name))

        # Get only inducers that have ``shot_vol``
        inducers_rows = [ind
                         for ind in self.inducers['rows']
                         if hasattr(ind, 'shot_vol')]
        inducers_cols = [ind
                         for ind in self.inducers['cols']
                         if hasattr(ind, 'shot_vol')]
        inducers_wells = [ind
                          for ind in self.inducers['wells']
                          if hasattr(ind, 'shot_vol')]
        inducers_media = [ind
                          for ind in self.inducers['media']
                          if hasattr(ind, 'shot_vol')]

        # Skip if no inducers are present
        if not(inducers_rows or inducers_cols or \
                inducers_wells or inducers_media):
            return

        # Check that a sheet with the specified name doesn't exist
        if sheet_name in [ws.title for ws in workbook.worksheets]:
            raise ValueError("sheet \"{}\"already present in workbook".\
                format(sheet_name))

        # Only "wells" or "media" supported if not measuring full plate
        if (self.samples_to_measure != self.n_cols*self.n_rows) and \
                inducers_rows:
            raise ValueError('"rows" not possible for a non-full plate')
        if (self.samples_to_measure != self.n_cols*self.n_rows) and \
                inducers_cols:
            raise ValueError('"cols" not possible for a non-full plate')

        # Initialize inducer instructions table
        ind_layout = []
        ind_layout_rows = self.n_rows + len(inducers_rows) + \
            len(inducers_rows) + len(inducers_cols) + len(inducers_wells) + \
            len(inducers_media) + 1
        ind_layout_cols = self.n_cols + len(inducers_cols)
        for i in range(ind_layout_rows):
            ind_layout.append(['']*(ind_layout_cols))

        # Add well coordinates
        for i in range(self.n_rows):
            for j in range(self.n_cols):
                array_i = i/(self.plate_n_rows)
                array_j = j/(self.plate_n_cols)
                plate_i = i%(self.plate_n_rows)
                plate_j = j%(self.plate_n_cols)
                row = i + len(inducers_rows)
                col = j + len(inducers_cols)
                ind_layout[row][col] += "{} ({}, {})".format(
                    self.plate_names[array_i*self.array_n_cols + array_j],
                    plate_i + 1,
                    plate_j + 1)
        # Add row inducer information
        for i, inducer in enumerate(inducers_rows):
            for j, val in enumerate(inducer.doses_table.index):
                row = i
                col = j + len(inducers_cols)
                ind_layout[row][col] = str(val)
        # Add column inducer information
        for j, inducer in enumerate(inducers_cols):
            for i, val in enumerate(inducer.doses_table.index):
                row = i + len(inducers_rows)
                col = j
                ind_layout[row][col] = str(val)
        # Add individual well inducer information
        for k, inducer in enumerate(inducers_wells):
            for i in range(self.n_rows):
                for j in range(self.n_cols):
                    if (i*self.n_cols + j) >= self.samples_to_measure:
                        break
                    row = i + len(inducers_rows)
                    col = j + len(inducers_cols)
                    ind_layout[row][col] += "\n{}".format(
                        inducer.doses_table.index[i*self.n_cols + j])
        # Add volume of each inducer to add
        current_row = self.n_rows + len(inducers_rows) + 1
        for inducer in inducers_rows:
            ind_layout[current_row][0] = "Add {:.2f}µL of {} to each well.".\
            format(inducer.shot_vol, inducer.name)
            current_row += 1
        for inducer in inducers_cols:
            ind_layout[current_row][0] = "Add {:.2f}µL of {} to each well.".\
            format(inducer.shot_vol, inducer.name)
            current_row += 1
        for inducer in inducers_wells:
            ind_layout[current_row][0] = "Add {:.2f}µL of {} to each well.".\
            format(inducer.shot_vol, inducer.name)
            current_row += 1
        for inducer in inducers_media:
            ind_layout[current_row][0] = "Add {:.2f}µL of {} to media.".\
            format(inducer.shot_vol, inducer.name)
            current_row += 1

        # Plate area
        plate_min_row = len(inducers_rows)
        plate_max_row = len(inducers_rows) + self.n_rows
        plate_min_col = len(inducers_cols)
        plate_max_col = len(inducers_cols) + self.n_cols

        # Create and populate worksheet
        worksheet = workbook.create_sheet(title=sheet_name)
        for i in range(len(ind_layout)):
            for j in range(len(ind_layout[i])):
                cell = worksheet.cell(row=i+1, column=j+1)
                # Write value
                cell.value = ind_layout[i][j]
                # Apply styles
                if (i >= plate_min_row) and (i < plate_max_row) and\
                        (j >= plate_min_col) and (j < plate_max_col):
                    array_i = (i - plate_min_row) / self.plate_n_rows
                    array_j = (j - plate_min_col) / self.plate_n_cols
                    cell.fill = plate_fill[(array_i + array_j)%2]
                    cell.border = plate_border
                    cell.alignment = plate_alignment

    def close_plates(self):
        """
        Generate ``ClosedPlate`` objects for each plate in the array.

        The individual ``ClosedPlate`` instances contain general plate
        information such as plate name, dimensions, cell inoculation
        conditions, and metadata, as well as well-specific information such
        as inducer concentrations. All this info is generated when calling
        `close_plates()`, and will remain fixed even after modifying
        inducers or other information in the ``PlateArray`` object.

        Within an experiment workflow, this function is meant to be called
        at the end of the Replicate Setup stage.

        Return
        ------
        list
            ``ClosedPlate`` instances with information about each sample.
            The number of closed plates in this list is the number of
            plates in the array, i.e., ``array_n_rows * array_n_cols``.

        """
        # Prepare plate info
        # Plate info will be the same on all plates, so it can be made once.
        plate_info = collections.OrderedDict()

        # Add plate array name
        plate_info["Plate Array"] = self.name

        # Add cell info
        plate_info['Strain'] = self.cell_strain_name
        if self.cell_setup_method=='fixed_od600':
            plate_info['Preculture Dilution'] = self.cell_predilution
            plate_info['Initial OD600'] = self.cell_initial_od600
        elif self.cell_setup_method=='fixed_volume':
            plate_info['Preculture Dilution'] = self.cell_predilution
            plate_info['Cell Inoculated Vol.'] = self.cell_shot_vol
        elif self.cell_setup_method=='fixed_dilution':
            plate_info['Preculture Dilution'] = self.cell_predilution
            plate_info['Total Cell Dilution'] = self.cell_total_dilution

        # Add additional plate metadata
        # The following try-catch block is needed to ensure compatibility with
        # both python2 and python3.
        try:
            items = self.metadata.iteritems()
        except AttributeError:
            items = self.metadata.items()
        for k, v in items:
            plate_info[k] = v

        # Prepare well info
        # Well info will be prepared for all the wells in the array. Later,
        # wells corresponding to each individual plate will be separated.
        # For this to be possible, two additional columns with the plate
        # position in the array will be added.
        well_info_array = pandas.DataFrame(index=range(self.n_rows*self.n_cols))
        well_info_array['Array Row'] = (numpy.repeat(numpy.arange(self.n_rows),
            self.n_cols) // self.plate_n_rows) + 1
        well_info_array['Array Column'] = (numpy.tile(numpy.arange(self.n_cols),
            self.n_rows) // self.plate_n_cols) + 1

        # Make boolean array indicating which samples should be measured
        samples_to_measure_bool = (numpy.arange(self.n_rows*self.n_cols) < \
            self.samples_to_measure)

        # Add inducer info
        # The following try-catch block is needed to ensure compatibility with
        # both python2 and python3.
        try:
            items = self.inducers.iteritems()
        except AttributeError:
            items = self.inducers.items()
        for apply_to, inducers in items:
            for inducer in inducers:
                if apply_to=='rows':
                    for column in inducer.doses_table.columns:
                        for i in range(self.n_rows):
                            for j in range(self.n_cols):
                                well_info_array.set_value(
                                    well_info_array.index[i*self.n_cols + j],
                                    column,
                                    inducer.doses_table.iloc[j][column])
                elif apply_to=='cols':
                    for column in inducer.doses_table.columns:
                        for i in range(self.n_rows):
                            for j in range(self.n_cols):
                                well_info_array.set_value(
                                    well_info_array.index[i*self.n_cols + j],
                                    column,
                                    inducer.doses_table.iloc[i][column])
                elif apply_to=='wells':
                    for column in inducer.doses_table.columns:
                        well_info_array.loc[samples_to_measure_bool, column] = \
                            inducer.doses_table[column].values
                elif apply_to=='media':
                    for column in inducer.doses_table.columns:
                        well_info_array.loc[samples_to_measure_bool, column] = \
                            inducer.doses_table[column].values[0]

        # Add which wells should be measured
        well_info_array['Measure'] = samples_to_measure_bool

        # Create closed plate objects
        closed_plates = []
        for i in range(self.array_n_rows):
            for j in range(self.array_n_cols):
                # Get plate name
                plate_name = self.plate_names[i*self.array_n_cols + j]
                # Filter well info
                well_info = well_info_array[(well_info_array['Array Row']==i+1)\
                                       & (well_info_array['Array Column']==j+1)]
                # Drop columns "Array Row" and "Array Column"
                well_info = well_info.drop('Array Row', axis=1)
                well_info = well_info.drop('Array Column', axis=1)
                # Reset index
                well_info.reset_index(drop=True, inplace=True)
                # Create closed plate
                closed_plate = ClosedPlate(name=plate_name,
                                           n_rows=self.plate_n_rows,
                                           n_cols=self.plate_n_cols,
                                           plate_info=plate_info.copy(),
                                           well_info=well_info)
                closed_plates.append(closed_plate)

        return closed_plates

class ClosedPlate(object):
    """
    Object that represents a closed plate.

    This class represents a physical plate in which samples have been
    inoculated with cells and inducers at the end of the Replicate Setup
    stage. General information about the plate (e.g. name, dimensions,
    other metadata) and well-specific information (e.g. inducer
    concentrations) are retained here. Because a `ClosedPlate` object
    represents a plate after all pipetting has taken place, its contents
    are not meant to be modified.

    A `ClosedPlate` object is not meant to be created by a user. Rather,
    the user is expected to design a plate experiment using classes
    `Plate`, `PlateArray`, or any derived class.

    Parameters
    ----------
    name : str
        Name of the plate.
    n_rows, n_cols : int, optional
        Number of rows and columns in the plate. Defaults: 4 and 6.
    plate_info : dict, optional
        General information about the plate, in a ``key:value`` format.
    well_info : DataFrame, optional
        Information about each well in the plate. Should have ``n_rows *
        n_cols`` rows.

    Attributes
    ----------
    name : str
        Name of the plate.
    n_rows, n_cols : int
        Number of rows and columns in the plate
    plate_info : dict
        General information about the plate, in a ``key:value`` format.
    well_info : DataFrame
        Information about each well in the plate
    samples_table : DataFrame
        Table with full information about each well in the plate, to be
        used in the Replicate Measurement stage. This table is calculated
        from all the other attributes upon object creation. If any
        attribute is modified, changes will not be reflected in
        `samples_table` unless the method `update_samples_table()` is
        called. `samples_table` contains one row per well in the plate. In
        it, the well position is indicated, and the plate name and all
        ``key:value`` pairs in `plate_info` are repeated in all rows. All
        the contents in `well_info` are copied into `samples_table` without
        modification.

    Methods
    -------
    update_samples_table
        Updates `samples_table` from all other attributes.

    """
    def __init__(self,
                 name,
                 n_rows=4,
                 n_cols=6,
                 plate_info=None,
                 well_info=None):
        # Store name
        self.name = name
        # Store dimensions
        self.n_rows = n_rows
        self.n_cols = n_cols

        # Check that well info has proper length
        if (well_info is not None) and \
                (len(well_info)!=self.n_rows*self.n_cols):
            raise ValueError('number of rows in well_info does not match plate '
                'dimensions')

        # Store well and plate info
        self.well_info = well_info
        self.plate_info = plate_info

        # The following initializes the samples table from current information
        self.update_samples_table()

    def update_samples_table(self):
        """
        Updates `samples_table` from all other attributes.

        """
        # Initialize samples table as a DataFrame
        samples_table = pandas.DataFrame(index=range(self.n_rows*self.n_cols))

        # Add plate name
        samples_table['Plate'] = self.name

        # Add plate information
        if self.plate_info is not None:
            # The following try-catch block is needed to ensure compatibility
            # with both python2 and python3.
            try:
                items = self.plate_info.iteritems()
            except AttributeError:
                items = self.plate_info.items()
            for k, v in items:
                samples_table[k] = v

        # Add row and column numbers
        samples_table['Row'] = numpy.repeat(numpy.arange(self.n_rows) + 1,
            self.n_cols)
        samples_table['Column'] = numpy.tile(numpy.arange(self.n_cols) + 1,
            self.n_rows)

        # Add well information
        if self.well_info is not None:
            # Check proper dimensions first
            if (len(self.well_info) != self.n_rows*self.n_cols):
                raise ValueError('number of rows in well_info does not match '
                    'plate dimensions')
            # Add information
            samples_table = pandas.concat([samples_table, self.well_info],
                                          axis=1)

        # Save samples table
        self.samples_table = samples_table
