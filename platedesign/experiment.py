# -*- coding: UTF-8 -*-
"""
Module that contains the experiment class.

"""

import collections
import copy
import os
import random
import six

import numpy
import openpyxl
import pandas

import platedesign
import platedesign.inducer

class Experiment(object):
    """
    Object that represents a plate experiment.

    A plate experiment can include one or more plates and several inducers
    that are applied to these plates. An experiment is performed in
    replicates: one or more repetitions of the same protocol. Some
    conditions can be randomized in each replicate, such as which samples
    have which inducer concentrations.

    Setting up a plate experiment is performed in several stages:
        - Experiment setup: Inducers and plates are prepared and stored.
          No cells are cultured during this stage.
        - Replicate setup: Plates are set up with the indicated amount of
          of inducer, inoculated with cells, and placed under growth
          conditions for a specified amount of time.
        - Replicate measurement: Plates are placed in growth-arresting
          conditions, and measurements of each sample are conducted. These
          can be absorbance or fluorescence from plate readers,
          fluorescence from flow cytometers, etc.

    The major function of this class is to generate instruction files for
    each one of these stages, given specifications on the inducers and
    plates. For example:
        - Experiment setup: Instructions for preparing inducer dilutions
          and media.
        - Replicate setup: Instructions for inoculating inducers and cells.
        - Replicate measurement: Table with list of samples to measure.

    Attributes
    ----------
    plates : list
        Plates or plate arrays in the experiment.
    inducers : list
        Inducers in the experiment.
    n_replicates : int
        Number of replicates.
    randomize_inducers : bool
        Whether to randomize inducer concentrations for each replicate.
    n_replicates_extra_inducer : int
        Number of extra replicates for which to prepare inducer. For
        example, if `n_replicates` is 3 and `n_replicates_extra_inducer` is
        2, inducer will be prepared for 5 replicates.
    plate_resources : dict
        Names of per-plate resources (e.g. slots in an incubator,
        temperature sensors, optogenetic devices, etc), in a ``key: value``
        format, where ``value`` is a list with as many elements as plates
        are going to be handled in the experiment. Resources are assigned
        to each closed plate at the end of the Replicate Setup Stage.
    randomize_plate_resources : bool
        Whether to randomize assignment of resources to each plate. Note
        that each resource is randomized independently.
    measurement_order : str
        Order in which to measure plates. Can be in the order that plates
        were added to the experiment ("Plate"), random ("Random"), or in
        the specified order of one of the plate resources (by specifying
        the resource name). Note that the latter will only be different
        from the original plate order if `randomize_plate_resources` is
        ``True``.
    measurement_template : str
        Name of a file to be used as template for the replicate measurement
        table. Sheets other than "Samples" will be copied unmodified. If a
        "Samples" sheet is present, all columns will be added to the final
        "Samples" sheet, and filled with the values present in the first
        row.
    plate_measurements : list of str
        Each element of this list is the name of a measurement to record
        at the end of the experiment for each plate. An empty table in
        which to record these values will be created in the replicate
        measurement file, sheet "Plate Measurements".
    replicate_measurements : list of str
        Each element of this list is the name of a measurement to record
        at the end of the experiment for the whole replicate. An empty table
        in which to record these values will be created in the replicate
        measurement file, sheet "Replicate Measurements".

    """
    def __init__(self):

        # Initialize containers of plates and inducers.
        self.plates = []
        self.inducers = []

        # Initialize properties
        self.n_replicates = 3
        self.randomize_inducers = False
        self.n_replicates_extra_inducer = 0
        self.plate_resources = collections.OrderedDict()
        self.randomize_plate_resources = False
        self.measurement_order = 'Plate'
        self.measurement_template = None
        self.plate_measurements = []
        self.replicate_measurements = []

    def add_plate(self, plate):
        """
        Add plate to experiment.

        Parameters
        ----------
        plate : Plate
            Plate to add.

        """
        self.plates.append(plate)

    def add_inducer(self, inducer):
        """
        Add inducer to experiment.

        Parameters
        ----------
        inducer : Inducer
            Inducer to add.

        """
        self.inducers.append(inducer)

    def generate(self, path='.'):
        """
        Generate instruction files for all stages of the experiment.

        This function creates files with instructions for all the stages of
        an experiment. Experiment setup files are stored in `path`. Then,
        separate folders are created for each one of the replicates, and
        files for the replicate setup and replicate measurement stages are
        created therein.

        Parameters
        ----------
        path : str, optional
            Folder in which to create all experiment files.

        """
        # Obtain total number of closed plates
        n_closed_plates = 0
        for plate in self.plates:
            n_closed_plates += plate.n_plates

        # Check that enough plate resources have been specified
        for k, v in six.iteritems(self.plate_resources):
            if len(v) < n_closed_plates:
                raise ValueError(
                    "{} resources of type {} specified, should be {} or more".\
                        format(len(v), k, n_closed_plates))

        # Create folders for each replicate, if necessary
        if self.n_replicates > 1:
            replicate_folders = [os.path.join(path,
                                              'replicate_{:03d}'.format(i + 1))
                                 for i in range(self.n_replicates)]
            # Check if folders already exist. If so, abort.
            for folder in replicate_folders:
                if os.path.exists(folder):
                    raise IOError("folder {} already exists".format(folder))
            # Create folders
            for folder in replicate_folders:
                os.makedirs(folder)
        else:
            replicate_folders = [path]

        ###
        # Experiment Setup Stage
        ###
        # Check if spreadsheet already exists
        if self.n_replicates > 1:
            wb_exp_setup_filename = os.path.join(path, 'experiment_setup.xlsx')
        else:
            wb_exp_setup_filename = os.path.join(path, 'setup.xlsx')
        if os.path.exists(wb_exp_setup_filename):
            raise IOError("file {} already exists".format(
                wb_exp_setup_filename))
        # Create single spreadsheet for all experiment setup instructions
        wb_exp_setup = openpyxl.Workbook()
        # Remove sheet created by default
        wb_exp_setup.remove(wb_exp_setup.active)
        # Run Experiment Setup for inducers
        for inducer in self.inducers:
            # Get inducer applications on all plates
            ind_applications = []
            for plate in self.plates:
                for apply_to, plate_inducers in six.iteritems(plate.inducers):
                    if inducer in plate_inducers:
                        ind_applications.append({'apply_to': apply_to,
                                                 'plate': plate})

            # Consistency check: inducers should be applied to all plates
            # identically.
            apply_to_all = [a['apply_to'] for a in ind_applications]
            if not all([a==apply_to_all[0] for a in apply_to_all]):
                raise ValueError("inducer can only be applied to the same"
                    " dimension on all plates")
            apply_to = apply_to_all[0]

            # The following only applies to chemical inducers
            if isinstance(inducer, platedesign.inducer.ChemicalInducerBase):
                # Consistency check: inducers should be applied to samples with
                # identical volumes in all plates
                media_vol_all = [a['plate'].apply_inducer_media_vol(apply_to)
                                 for a in ind_applications]
                if not all([m==media_vol_all[0] for m in media_vol_all]):
                    raise ValueError("inducer can only be applied to the same"
                        " media volume on all plates")
                # Set media volume in inducer object
                inducer.media_vol = media_vol_all[0]

                # Calculate total amount of inducer to make from number of shots
                # and replicates
                n_shots = sum([a['plate'].apply_inducer_n_shots(apply_to)
                               for a in ind_applications])
                inducer.set_vol_from_shots(
                    n_shots=n_shots,
                    n_replicates=self.n_replicates + \
                        self.n_replicates_extra_inducer)

            # Save files
            inducer.save_exp_setup_instructions(workbook=wb_exp_setup)
            inducer.save_exp_setup_files(path=path)

        # Run Experiment Setup for plates
        for plate in self.plates:
            plate.save_exp_setup_instructions(workbook=wb_exp_setup)
            plate.save_exp_setup_files(path=path)

        # Save spreadsheet
        if self.n_replicates > 1 and len(wb_exp_setup.worksheets) > 0:
            wb_exp_setup.save(filename=wb_exp_setup_filename)

        # Iterate over replicates
        for replicate_idx in range(self.n_replicates):
            ###
            # Replicate Setup Stage
            ###
            # Create new spreadsheet for replicate setup if more than one
            # replicate, else keep using the experiment setup spreadsheet.
            if self.n_replicates > 1:
                wb_rep_setup = openpyxl.Workbook()
                wb_rep_setup.remove(wb_rep_setup.active)
            else:
                wb_rep_setup = wb_exp_setup

            # Make summary sheet
            summary_table = pandas.DataFrame(columns=['Plate Array',
                                                      'Plate',
                                                      'Strain'])
            # Add plate and plate array names
            for p in self.plates:
                if hasattr(p, 'plate_names'):
                    # Plate array
                    for plate in p.plate_names:
                        summary_table.loc[len(summary_table)] = \
                            [p.name, plate, p.cell_strain_name]
                else:
                    # Single plate
                    summary_table.loc[len(summary_table)] = \
                        [None, p.name, p.cell_strain_name]
            # Discard "Plate Array" column if empty
            if summary_table['Plate Array'].isnull().values.all():
                summary_table.drop('Plate Array', axis=1, inplace=True)
            # Save
            writer = pandas.ExcelWriter('temp', engine='openpyxl')
            writer.book = wb_rep_setup
            summary_table.to_excel(writer,
                                   sheet_name='Summary',
                                   index=False)

            # Shuffle inducer and save replicate setup files
            for inducer in self.inducers:
                if self.randomize_inducers:
                    inducer.shuffle()
                inducer.save_rep_setup_instructions(workbook=wb_rep_setup)
                inducer.save_rep_setup_files(
                    path=replicate_folders[replicate_idx])

            # Randomize plate resources
            # We will not modify the elements of ``plate_resources``. Instead,
            # we will maintain a set of indices that will be shuffled.
            # This will keep the original ``plate_resources`` intact for the
            # next replicate. In addition, this will allow to sort closed plates
            # based on a resource's original order later on.
            plate_resources_ind = {
                k: list(range(len(v)))
                for k, v in six.iteritems(self.plate_resources)}
            if self.randomize_plate_resources:
                for k, v in six.iteritems(plate_resources_ind):
                    random.shuffle(v)

            # Assign resources to plates
            resource_shift = 0
            for plate in self.plates:
                for k, v in six.iteritems(self.plate_resources):
                    # Make a copy of the resource list, and reorganize it
                    # using the shuffled indices
                    v_rep = [v[i] for i in plate_resources_ind[k]]
                    # Copy resources
                    plate.resources[k] = v_rep[resource_shift: \
                                               resource_shift + plate.n_plates]
                resource_shift += plate.n_plates

            # Generate and save replicate setup information
            for plate in self.plates:
                # Save files
                plate.save_rep_setup_instructions(workbook=wb_rep_setup)
                plate.save_rep_setup_files(
                    path=replicate_folders[replicate_idx])

            # Get closed plates from plates and plate arrays.
            closed_plates = []
            for plate in self.plates:
                closed_plates.extend(plate.close_plates())

            # Reorganize closed plates
            if self.measurement_order == "Plate":
                # Don't do anything
                pass
            elif self.measurement_order == "Random":
                random.shuffle(closed_plates)
            elif self.measurement_order in self.plate_resources:
                # Reorganize based on original order of a plate resource
                # To do this, we create a new ``closed_plates_temp`` list, and
                # we add closed plates to positions determined by the shuffled
                # resource indices
                resources_idx = plate_resources_ind[self.measurement_order]
                closed_plates_temp = [None]*len(resources_idx)
                for i, closed_plate in zip(resources_idx, closed_plates):
                    closed_plates_temp[i] = closed_plate
                closed_plates = [c for c in closed_plates_temp if c is not None]
            else:
                raise ValueError("measurement order {} not supported".format(
                    self.measurement_order))

            # Add resources sheet to replicate setup instructions
            if self.plate_resources:
                # Generate table
                resources_table = pandas.DataFrame()
                resources_table['Plate'] = [p.name for p in closed_plates]
                for k, v in six.iteritems(self.plate_resources):
                    resources_table[k] = [p.plate_info[k]
                                          for p in closed_plates]
                # Generate pandas writer and reassign workbook
                writer = pandas.ExcelWriter('temp', engine='openpyxl')
                writer.book = wb_rep_setup
                resources_table.to_excel(writer,
                                         sheet_name='Plate Resources',
                                         index=False)

            # Save spreadsheet
            if self.n_replicates > 1:
                wb_rep_setup_filename = os.path.join(
                    replicate_folders[replicate_idx],
                    'replicate_{:03d}_setup.xlsx'.format(replicate_idx + 1))
            else:
                wb_rep_setup_filename = os.path.join(
                    replicate_folders[replicate_idx],
                    'setup.xlsx')

            if len(wb_rep_setup.worksheets) > 0:
                wb_rep_setup.save(filename=wb_rep_setup_filename)

            ###
            # Replicate Measurement Stage
            ###

            # Plate measurements table
            plate_measurements_table = pandas.DataFrame()
            plate_measurements_table['Plate'] = [p.name for p in closed_plates]
            for m in self.plate_measurements:
                plate_measurements_table[m] = numpy.nan
            # Plate column should be the index
            plate_measurements_table.set_index('Plate', inplace=True)

            # Replicate measurements table
            replicate_measurements_table = pandas.DataFrame()
            replicate_measurements_table['Key'] = self.replicate_measurements
            replicate_measurements_table['Value'] = numpy.nan
            replicate_measurements_table.set_index('Key', inplace=True)

            # Samples table
            samples_table = pandas.DataFrame()
            samples_table_columns = []
            for closed_plate in closed_plates:
                # Update and extract samples table from plate, and eliminate
                # samples that should not be measured
                closed_plate.update_samples_table()
                plate_table = closed_plate.samples_table.copy()
                if 'Measure' in plate_table.columns:
                    plate_table = plate_table[plate_table['Measure']]
                    plate_table.drop('Measure', axis=1, inplace=True)
                # The following is necessary to preserve the order of the
                # columns when appending
                for column in plate_table.columns:
                    if column not in samples_table_columns:
                        samples_table_columns.append(column)
                # Append plate's samples table to samples table
                samples_table = samples_table.append(plate_table)
                # Reorganize columns
                samples_table = samples_table[samples_table_columns]

            # Create ID column and set as the index
            samples_table['ID'] = ['S{:04d}'.format(i+1)
                                   for i in range(len(samples_table))]
            samples_table.set_index('ID', inplace=True)

            # File name for replicate measurement file
            if self.n_replicates > 1:
                wb_rep_measurement_filename = os.path.join(
                    replicate_folders[replicate_idx],
                    'replicate_{:03d}_measurement.xlsx'.format(replicate_idx+1))
            else:
                wb_rep_measurement_filename = os.path.join(
                    replicate_folders[replicate_idx],
                    'measurement.xlsx')
            # Generate pandas writer
            writer = pandas.ExcelWriter(wb_rep_measurement_filename,
                                        engine='openpyxl')
            
            # Add information from template file, if specified
            if self.measurement_template is not None:
                workbook_template = openpyxl.load_workbook(
                    self.measurement_template)
                for ws in workbook_template.worksheets:
                    if ws.title != "Samples":
                        # Create new sheet
                        ws_new = writer.book.create_sheet(ws.title)
                        # Copy all cells' content and style
                        for row in ws.iter_rows():
                            for cell in row:
                                new_cell = ws_new.cell(row=cell.row,
                                                       column=cell.col_idx)
                                new_cell.value = cell.value
                                if cell.has_style:
                                    new_cell.font = \
                                        copy.copy(cell.font)
                                    new_cell.border = \
                                        copy.copy(cell.border)
                                    new_cell.fill = \
                                        copy.copy(cell.fill)
                                    new_cell.number_format = \
                                        copy.copy(cell.number_format)
                                    new_cell.protection = \
                                        copy.copy(cell.protection)
                                    new_cell.alignment = \
                                        copy.copy(cell.alignment)
                    else:
                        # Read with pandas
                        samples_extra = pandas.read_excel(
                            self.measurement_template,
                            sheetname="Samples")
                        # Extract first row
                        samples_extra = samples_extra.iloc[0]
                        # Add columns to samples_table
                        for index, value in six.iteritems(samples_extra):
                            try:
                                value = [value.format(i + 1)
                                         for i in range(len(samples_table))]
                            except AttributeError as e:
                                pass
                            samples_table[index] = value

            # Convert pandas tables to sheet and save
            samples_table.to_excel(writer, sheet_name='Samples')
            if len(plate_measurements_table.columns):
                plate_measurements_table.to_excel(
                    writer,
                    sheet_name='Plate Measurements')
            if len(self.replicate_measurements):
                replicate_measurements_table.to_excel(
                    writer,
                    sheet_name='Replicate Measurements',
                    header=False)
            writer.save()
