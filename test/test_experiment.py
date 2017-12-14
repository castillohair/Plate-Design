# -*- coding: UTF-8 -*-
"""
Unit tests for experiment module

"""

import collections
import itertools
import os
import random
import six
import shutil
import unittest

import numpy
import openpyxl
import pandas

import platedesign

def assert_worksheets_equal(test_case,
                            wb_name_1=None,
                            ws_name_1=None,
                            wb_name_2=None,
                            ws_name_2=None,
                            ws_1=None,
                            ws_2=None):
    # Get worksheets
    if ws_1 is None:
        wb_1 = openpyxl.load_workbook(filename=wb_name_1)
        ws_1 = wb_1[ws_name_1]
    if ws_2 is None:
        wb_2 = openpyxl.load_workbook(filename=wb_name_2)
        ws_2 = wb_2[ws_name_2]
    # Verify number of rows and columns
    test_case.assertEqual(len(list(ws_1.rows)), len(list(ws_2.rows)))
    test_case.assertEqual(len(list(ws_1.columns)), len(list(ws_2.columns)))
    # Verify contents
    for row_1, row_2 in zip(ws_1.rows, ws_2.rows):
        for cell_1, cell_2 in zip(row_1, row_2):
            test_case.assertEqual(cell_1.value, cell_2.value)

def get_dataframe_from_worksheet(ws):
    ws_values = ws.values
    cols = next(ws_values)
    ws_values = list(ws_values)
    return pandas.DataFrame(ws_values, columns=cols)

class TestExperiment(unittest.TestCase):
    """
    Tests for the Experiment class

    """
    def setUp(self):
        # Directory where to save temporary files
        self.temp_dir = "test/temp_experiment"
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)
        # Set random seed
        random.seed(1)

    def tearDown(self):
        # Delete temporary directory
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def test_create(self):
        exp = platedesign.experiment.Experiment()

    def test_default_attributes(self):
        exp = platedesign.experiment.Experiment()
        # Containers of plates and inducers.
        self.assertEqual(exp.plates, [])
        self.assertEqual(exp.inducers, [])
        # Other attributes
        self.assertEqual(exp.n_replicates, 3)
        self.assertFalse(exp.randomize_inducers)
        self.assertEqual(exp.plate_resources, collections.OrderedDict())
        self.assertFalse(exp.randomize_plate_resources)
        self.assertEqual(exp.measurement_order, 'Plate')
        self.assertIsNone(exp.measurement_template)
        self.assertEqual(exp.plate_measurements, [])
        self.assertEqual(exp.replicate_measurements, [])

    def test_add_plate(self):
        exp = platedesign.experiment.Experiment()
        # Verify that plate and inducer lists are empty
        self.assertEqual(exp.plates, [])
        self.assertEqual(exp.inducers, [])
        # Create plate and add
        p1 = platedesign.plate.Plate('P1', n_rows=4, n_cols=6)
        exp.add_plate(p1)
        # Verify that plate list contains new plate
        self.assertEqual(exp.plates, [p1])
        self.assertEqual(exp.inducers, [])
        # Add a second plate
        p2 = platedesign.plate.Plate('P2', n_rows=4, n_cols=6)
        exp.add_plate(p2)
        # Verify that plate list contains new plate
        self.assertEqual(exp.plates, [p1, p2])
        self.assertEqual(exp.inducers, [])

    def test_add_inducer(self):
        exp = platedesign.experiment.Experiment()
        # Verify that plate and inducer lists are empty
        self.assertEqual(exp.plates, [])
        self.assertEqual(exp.inducers, [])
        # Create inducer and add
        ind1 = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        exp.add_inducer(ind1)
        # Verify that inducer list contains new inducer
        self.assertEqual(exp.plates, [])
        self.assertEqual(exp.inducers, [ind1])
        # Add a second inducer
        ind2 = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        exp.add_inducer(ind2)
        # Verify that plate list contains new plate
        self.assertEqual(exp.plates, [])
        self.assertEqual(exp.inducers, [ind1, ind2])

    def test_one_inducer_one_replicate(self):
        # Experiment with one inducer and one replicate
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 1
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [0, 2, 8, 16, 64, 500]
        exp.add_inducer(iptg)
        # Plate
        plate = platedesign.plate.Plate('P1', n_rows=4, n_cols=6)
        plate.cell_strain_name = 'Test Strain 1'
        plate.total_media_vol = 16000.
        plate.sample_media_vol = 500.
        plate.cell_setup_method = 'fixed_volume'
        plate.cell_predilution = 100
        plate.cell_predilution_vol = 1000
        plate.cell_shot_vol = 5
        plate.apply_inducer(inducer=iptg, apply_to='rows')
        exp.add_plate(plate)
        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertIsNone(iptg.replicate_vol)
        self.assertEqual(iptg.total_vol, 30)

        # Check that appropriate excel files exist
        self.assertTrue(os.path.isfile(os.path.join(self.temp_dir, 
                                                    'setup.xlsx')))
        self.assertTrue(os.path.isfile(os.path.join(self.temp_dir, 
                                                    'measurement.xlsx')))
        self.assertFalse(os.path.isfile(os.path.join(self.temp_dir, 
                                                     'experiment_setup.xlsx')))
        self.assertFalse(os.path.isdir(os.path.join(self.temp_dir, 
                                                    'replicate_001')))
        # Check sheet names of setup file
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG",
                                         "Summary",
                                         "Inducers for Plate P1",
                                         "Cells for Plate P1",
                                         ])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])
        # Check summary table
        summary_table = get_dataframe_from_worksheet(wb["Summary"])
        summary_table_exp = pandas.DataFrame()
        summary_table_exp['Plate Array'] = [None]
        summary_table_exp['Plate'] = ['P1']
        summary_table_exp['Strain'] = ['Test Strain 1']
        pandas.util.testing.assert_frame_equal(summary_table,
                                               summary_table_exp)
        # Check proper plate setup instructions
        # Instructions will be compared to the ones generated by the plate
        # class directly
        file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')
        plate.save_rep_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, 
                                ws_1=wb['Inducers for Plate P1'],
                                ws_2=wb_exp['Inducers for Plate P1'])
        assert_worksheets_equal(self, 
                                ws_1=wb['Cells for Plate P1'],
                                ws_2=wb_exp['Cells for Plate P1'])

        # Check sheet names of measurement file
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'measurement.xlsx'))
        self.assertEqual(wb.sheetnames, ["Samples"])
        samples_table = get_dataframe_from_worksheet(wb["Samples"])
        self.assertEqual(len(samples_table.columns), 8)
        numpy.testing.assert_array_equal(
            samples_table['ID'].values,
            ["S{:04d}".format(i+1) for i in range(24)])
        numpy.testing.assert_array_equal(
            samples_table['Plate'].values,
            ["P1" for i in range(24)])
        numpy.testing.assert_array_equal(
            samples_table['Strain'].values,
            ["Test Strain 1" for i in range(24)])
        numpy.testing.assert_array_equal(
            samples_table['Preculture/Aliquot Dilution'].values,
            [100 for i in range(24)])
        numpy.testing.assert_array_equal(
            samples_table['Cell Inoculated Vol.'].values,
            [5 for i in range(24)])
        numpy.testing.assert_array_equal(
            samples_table['Row'].values,
            [(i//6 + 1) for i in range(24)])
        numpy.testing.assert_array_equal(
            samples_table['Column'].values,
            [(i%6 + 1) for i in range(24)])
        numpy.testing.assert_array_equal(
            samples_table[u'IPTG Concentration (µM)'].values,
            numpy.tile([0, 2, 8, 16, 64, 500], 4))

    def test_one_inducer_many_replicates(self):
        # Experiment with one inducer and many replicates
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [0, 2, 8, 16, 64, 500]
        exp.add_inducer(iptg)
        # Plate
        plate = platedesign.plate.Plate('P1', n_rows=4, n_cols=6)
        plate.cell_strain_name = 'Test Strain 1'
        plate.total_media_vol = 16000.
        plate.sample_media_vol = 500.
        plate.cell_setup_method = 'fixed_volume'
        plate.cell_predilution = 100
        plate.cell_predilution_vol = 1000
        plate.cell_shot_vol = 5
        plate.apply_inducer(inducer=iptg, apply_to='rows')
        exp.add_plate(plate)
        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 30)
        self.assertEqual(iptg.total_vol, 200)

        # Check that appropriate excel files exist
        self.assertTrue(os.path.isfile(os.path.join(self.temp_dir, 
                                                    'experiment_setup.xlsx')))
        for i in range(3):
            # Replicate directory
            self.assertTrue(os.path.isdir(os.path.join(
                self.temp_dir, 
                'replicate_{:03d}'.format(i+1))))
            # Replicate setup file
            self.assertTrue(os.path.isfile(os.path.join(
                self.temp_dir, 
                'replicate_{:03d}'.format(i+1),
                'replicate_{:03d}_setup.xlsx'.format(i+1),)))
            # Replicate measurement file
            self.assertTrue(os.path.isfile(os.path.join(
                self.temp_dir, 
                'replicate_{:03d}'.format(i+1),
                'replicate_{:03d}_measurement.xlsx'.format(i+1),)))
        # No additional replicate directory
        self.assertFalse(os.path.isdir(os.path.join(self.temp_dir, 
                                                    'replicate_004')))
        # No individual setup or measurement file
        self.assertFalse(os.path.isfile(os.path.join(self.temp_dir, 
                                                     'setup.xlsx')))
        self.assertFalse(os.path.isfile(os.path.join(self.temp_dir, 
                                                     'measurement.xlsx')))
        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        # Check sheet names of replicate setup files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate P1",
                                             "Cells for Plate P1",
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = [None]
            summary_table_exp['Plate'] = ['P1']
            summary_table_exp['Strain'] = ['Test Strain 1']
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')
            plate.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate P1'],
                                    ws_2=wb_exp['Inducers for Plate P1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P1'],
                                    ws_2=wb_exp['Cells for Plate P1'])

        # Check sheet names of replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             ])
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 8)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1" for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1" for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100 for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [(i//6 + 1) for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.tile([0, 2, 8, 16, 64, 500], 4))

    def test_one_plate_two_inducers(self):
        # Experiment with two inducers
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [0, 2, 8, 16, 64, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0, 0.005, 0.05, 0.05]
        exp.add_inducer(xyl)
        # Plate
        plate = platedesign.plate.Plate('P1', n_rows=4, n_cols=6)
        plate.cell_strain_name = 'Test Strain 1'
        plate.total_media_vol = 16000.
        plate.sample_media_vol = 500.
        plate.cell_setup_method = 'fixed_volume'
        plate.cell_predilution = 100
        plate.cell_predilution_vol = 1000
        plate.cell_shot_vol = 5
        plate.apply_inducer(inducer=iptg, apply_to='rows')
        plate.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(plate)
        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 30)
        self.assertEqual(iptg.total_vol, 200)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 40)
        self.assertEqual(xyl.total_vol, 200)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate P1",
                                             "Cells for Plate P1",
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = [None]
            summary_table_exp['Plate'] = ['P1']
            summary_table_exp['Strain'] = ['Test Strain 1']
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')
            plate.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate P1'],
                                    ws_2=wb_exp['Inducers for Plate P1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P1'],
                                    ws_2=wb_exp['Cells for Plate P1'])

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             ])
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 9)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1" for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1" for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100 for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [(i//6 + 1) for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(24)])
            numpy.testing.assert_array_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.tile([0, 2, 8, 16, 64, 500], 4))
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                numpy.repeat([0.0, 0.005, 0.05, 0.05], 6))

    def test_three_plates_two_inducers(self):
        # Two inducer, three plate experiment
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [0, 2, 8, 16, 64, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0, 0.005, 0.05, 0.05]
        exp.add_inducer(xyl)
        # Plates
        plate1 = platedesign.plate.Plate('P1', n_rows=4, n_cols=6)
        plate1.cell_strain_name = 'Test Strain 1'
        plate1.total_media_vol = 16000.
        plate1.sample_media_vol = 500.
        plate1.cell_setup_method = 'fixed_volume'
        plate1.cell_predilution = 100
        plate1.cell_predilution_vol = 1000
        plate1.cell_shot_vol = 5
        plate1.apply_inducer(inducer=iptg, apply_to='rows')
        plate1.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(plate1)

        plate2 = platedesign.plate.Plate('P2', n_rows=4, n_cols=6)
        plate2.cell_strain_name = 'Test Strain 2'
        plate2.total_media_vol = 16000.
        plate2.sample_media_vol = 500.
        plate2.cell_setup_method = 'fixed_volume'
        plate2.cell_predilution = 1000
        plate2.cell_predilution_vol = 1000
        plate2.cell_shot_vol = 5
        plate2.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(plate2)

        plate3 = platedesign.plate.Plate('P3', n_rows=4, n_cols=6)
        plate3.samples_to_measure = 4
        plate3.cell_strain_name = 'Test Strain 3'
        plate3.total_media_vol = 16000.
        plate3.sample_media_vol = 500.
        plate3.cell_setup_method = 'fixed_volume'
        plate3.cell_predilution = 10
        plate3.cell_predilution_vol = 1000
        plate3.cell_shot_vol = 5
        exp.add_plate(plate3)

        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 30)
        self.assertEqual(iptg.total_vol, 200)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 80)
        self.assertEqual(xyl.total_vol, 300)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate P1",
                                             "Cells for Plate P1",
                                             "Inducers for Plate P2",
                                             "Cells for Plate P2",
                                             "Cells for Plate P3",
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = [None, None, None]
            summary_table_exp['Plate'] = ['P1', 'P2', 'P3']
            summary_table_exp['Strain'] = ['Test Strain 1',
                                           'Test Strain 2',
                                           'Test Strain 3']
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')

            plate1.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate P1'],
                                    ws_2=wb_exp['Inducers for Plate P1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P1'],
                                    ws_2=wb_exp['Cells for Plate P1'])

            plate2.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate P2'],
                                    ws_2=wb_exp['Inducers for Plate P2'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P2'],
                                    ws_2=wb_exp['Cells for Plate P2'])

            plate3.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P3'],
                                    ws_2=wb_exp['Cells for Plate P3'])

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             ])
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 9)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(52)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1"]*24 + ["P2"]*24 + ["P3"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1"]*24 + ["Test Strain 2"]*24 + ["Test Strain 3"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100]*24 + [1000]*24 + [10]*4)
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(52)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [((i//6)%4 + 1) for i in range(52)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(52)])
            numpy.testing.assert_array_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.concatenate((
                    numpy.tile([0, 2, 8, 16, 64, 500], 4),
                    [numpy.nan]*28)))
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                numpy.concatenate((
                    numpy.repeat([0.0, 0.005, 0.05, 0.05], 6),
                    numpy.repeat([0.0, 0.005, 0.05, 0.05], 6),
                    [numpy.nan]*4)))

    def test_three_plates_two_inducers_randomize(self):
        # Two inducer, three plate experiment
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        exp.randomize_inducers = True
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [0, 2, 8, 16, 64, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0, 0.005, 0.05, 0.05]
        exp.add_inducer(xyl)
        # Plates
        plate1 = platedesign.plate.Plate('P1', n_rows=4, n_cols=6)
        plate1.cell_strain_name = 'Test Strain 1'
        plate1.total_media_vol = 16000.
        plate1.sample_media_vol = 500.
        plate1.cell_setup_method = 'fixed_volume'
        plate1.cell_predilution = 100
        plate1.cell_predilution_vol = 1000
        plate1.cell_shot_vol = 5
        plate1.apply_inducer(inducer=iptg, apply_to='rows')
        plate1.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(plate1)

        plate2 = platedesign.plate.Plate('P2', n_rows=4, n_cols=6)
        plate2.cell_strain_name = 'Test Strain 2'
        plate2.total_media_vol = 16000.
        plate2.sample_media_vol = 500.
        plate2.cell_setup_method = 'fixed_volume'
        plate2.cell_predilution = 1000
        plate2.cell_predilution_vol = 1000
        plate2.cell_shot_vol = 5
        plate2.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(plate2)

        plate3 = platedesign.plate.Plate('P3', n_rows=4, n_cols=6)
        plate3.samples_to_measure = 4
        plate3.cell_strain_name = 'Test Strain 3'
        plate3.total_media_vol = 16000.
        plate3.sample_media_vol = 500.
        plate3.cell_setup_method = 'fixed_volume'
        plate3.cell_predilution = 10
        plate3.cell_predilution_vol = 1000
        plate3.cell_shot_vol = 5
        exp.add_plate(plate3)

        # Generate experiment files
        random.seed(1)
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 30)
        self.assertEqual(iptg.total_vol, 200)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 80)
        self.assertEqual(xyl.total_vol, 300)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        iptg.unshuffle()
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        xyl.unshuffle()
        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        # Shuffling indices are different in python2 and 3
        if six.PY2:
            iptg_shuffled_idx = [[1, 2, 5, 3, 4, 0],
                                 [4, 2, 1, 3, 5, 0],
                                 [5, 2, 0, 3, 4, 1]]
            xyl_shuffled_idx = [[0, 2, 3, 1],
                                [3, 2, 1, 0],
                                [1, 0, 3, 2]]
        elif six.PY3:
            iptg_shuffled_idx = [[2, 3, 5, 0, 4, 1],
                                 [2, 4, 0, 1, 3, 5],
                                 [2, 5, 1, 3, 0, 4]]
            xyl_shuffled_idx = [[0, 2, 1, 3],
                                [3, 2, 1, 0],
                                [2, 3, 1, 0]]
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate P1",
                                             "Cells for Plate P1",
                                             "Inducers for Plate P2",
                                             "Cells for Plate P2",
                                             "Cells for Plate P3",
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = [None, None, None]
            summary_table_exp['Plate'] = ['P1', 'P2', 'P3']
            summary_table_exp['Strain'] = ['Test Strain 1',
                                           'Test Strain 2',
                                           'Test Strain 3']
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            # Shuffling indices should be restored
            iptg.shuffled_idx = iptg_shuffled_idx[rep_idx]
            xyl.shuffled_idx = xyl_shuffled_idx[rep_idx]
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')

            plate1.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate P1'],
                                    ws_2=wb_exp['Inducers for Plate P1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P1'],
                                    ws_2=wb_exp['Cells for Plate P1'])

            plate2.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate P2'],
                                    ws_2=wb_exp['Inducers for Plate P2'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P2'],
                                    ws_2=wb_exp['Cells for Plate P2'])

            plate3.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P3'],
                                    ws_2=wb_exp['Cells for Plate P3'])

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             ])
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 9)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(52)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1"]*24 + ["P2"]*24 + ["P3"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1"]*24 + ["Test Strain 2"]*24 + ["Test Strain 3"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100]*24 + [1000]*24 + [10]*4)
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(52)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [((i//6)%4 + 1) for i in range(52)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(52)])
            iptg_conc = [0, 2, 8, 16, 64, 500]
            iptg_conc = [iptg_conc[idx] for idx in iptg_shuffled_idx[rep_idx]]
            xyl_conc = [0.0, 0.005, 0.05, 0.05]
            xyl_conc = [xyl_conc[idx] for idx in xyl_shuffled_idx[rep_idx]]
            numpy.testing.assert_array_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.concatenate((
                    numpy.tile(iptg_conc, 4),
                    [numpy.nan]*28)))
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                numpy.concatenate((
                    numpy.repeat(xyl_conc, 6),
                    numpy.repeat(xyl_conc, 6),
                    [numpy.nan]*4)))

    def test_plate_array_two_inducers(self):
        # Two inducer, three plate experiment
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [ 0,   0.5,  1,   2,   4,   8,
                               16,  32,   64, 128, 256, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0,  0.0005, 0.001, 0.005,
                              0.01, 0.05,   0.1,   0.5]
        exp.add_inducer(xyl)

        # Plates
        platearray = platedesign.plate.PlateArray(
            'PA1',
            array_n_rows=2,
            array_n_cols=2,
            plate_names=['P{}'.format(i+1) for i in range(4)],
            plate_n_rows=4,
            plate_n_cols=6)
        platearray.cell_strain_name = 'Test Strain 1'
        platearray.total_media_vol = 16000.*4
        platearray.sample_media_vol = 500.
        platearray.cell_setup_method = 'fixed_volume'
        platearray.cell_predilution = 100
        platearray.cell_predilution_vol = 1000
        platearray.cell_shot_vol = 5
        platearray.apply_inducer(inducer=iptg, apply_to='rows')
        platearray.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(platearray)

        plate5 = platedesign.plate.Plate('P5', n_rows=4, n_cols=6)
        plate5.samples_to_measure = 4
        plate5.cell_strain_name = 'Autofluorescence Strain'
        plate5.total_media_vol = 16000.
        plate5.sample_media_vol = 500.
        plate5.cell_setup_method = 'fixed_volume'
        plate5.cell_predilution = 10
        plate5.cell_predilution_vol = 1000
        plate5.cell_shot_vol = 5
        exp.add_plate(plate5)

        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 50)
        self.assertEqual(iptg.total_vol, 200)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 80)
        self.assertEqual(xyl.total_vol, 300)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate Array PA1",
                                             "Cells for Plate Array PA1",
                                             "Cells for Plate P5",
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = ['PA1', 'PA1', 'PA1', 'PA1', None,]
            summary_table_exp['Plate'] = ['P1', 'P2', 'P3', 'P4', 'P5']
            summary_table_exp['Strain'] = ['Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Autofluorescence Strain',]
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')

            platearray.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate Array PA1'],
                                    ws_2=wb_exp['Inducers for Plate Array PA1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate Array PA1'],
                                    ws_2=wb_exp['Cells for Plate Array PA1'])

            plate5.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P5'],
                                    ws_2=wb_exp['Cells for Plate P5'])

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             ])
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 10)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1"]*24 + ["P2"]*24 + ["P3"]*24 + ["P4"]*24 + ["P5"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Plate Array'].values,
                ["PA1"]*96 + [None]*4)
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1"]*96 + ["Autofluorescence Strain"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100]*96 + [10]*4)
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [((i//6)%4 + 1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(100)])
            numpy.testing.assert_almost_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.concatenate((
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    [numpy.nan]*4)))
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                numpy.concatenate((
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    [numpy.nan]*4)))

    def test_plate_array_two_inducers_extra_inducer(self):
        # Two inducer, three plate experiment
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        exp.n_replicates_extra_inducer = 2
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [ 0,   0.5,  1,   2,   4,   8,
                               16,  32,   64, 128, 256, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0,  0.0005, 0.001, 0.005,
                              0.01, 0.05,   0.1,   0.5]
        exp.add_inducer(xyl)

        # Plates
        platearray = platedesign.plate.PlateArray(
            'PA1',
            array_n_rows=2,
            array_n_cols=2,
            plate_names=['P{}'.format(i+1) for i in range(4)],
            plate_n_rows=4,
            plate_n_cols=6)
        platearray.cell_strain_name = 'Test Strain 1'
        platearray.total_media_vol = 16000.*4
        platearray.sample_media_vol = 500.
        platearray.cell_setup_method = 'fixed_volume'
        platearray.cell_predilution = 100
        platearray.cell_predilution_vol = 1000
        platearray.cell_shot_vol = 5
        platearray.apply_inducer(inducer=iptg, apply_to='rows')
        platearray.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(platearray)

        plate5 = platedesign.plate.Plate('P5', n_rows=4, n_cols=6)
        plate5.samples_to_measure = 4
        plate5.cell_strain_name = 'Autofluorescence Strain'
        plate5.total_media_vol = 16000.
        plate5.sample_media_vol = 500.
        plate5.cell_setup_method = 'fixed_volume'
        plate5.cell_predilution = 10
        plate5.cell_predilution_vol = 1000
        plate5.cell_shot_vol = 5
        exp.add_plate(plate5)

        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 50)
        self.assertEqual(iptg.total_vol, 300)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 80)
        self.assertEqual(xyl.total_vol, 500)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate Array PA1",
                                             "Cells for Plate Array PA1",
                                             "Cells for Plate P5",
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = ['PA1', 'PA1', 'PA1', 'PA1', None,]
            summary_table_exp['Plate'] = ['P1', 'P2', 'P3', 'P4', 'P5']
            summary_table_exp['Strain'] = ['Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Autofluorescence Strain',]
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')

            platearray.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self,
                                    ws_1=wb['Inducers for Plate Array PA1'],
                                    ws_2=wb_exp['Inducers for Plate Array PA1'])
            assert_worksheets_equal(self,
                                    ws_1=wb['Cells for Plate Array PA1'],
                                    ws_2=wb_exp['Cells for Plate Array PA1'])

            plate5.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self,
                                    ws_1=wb['Cells for Plate P5'],
                                    ws_2=wb_exp['Cells for Plate P5'])

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             ])
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 10)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1"]*24 + ["P2"]*24 + ["P3"]*24 + ["P4"]*24 + ["P5"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Plate Array'].values,
                ["PA1"]*96 + [None]*4)
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1"]*96 + ["Autofluorescence Strain"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100]*96 + [10]*4)
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [((i//6)%4 + 1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(100)])
            numpy.testing.assert_almost_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.concatenate((
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    [numpy.nan]*4)))
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                numpy.concatenate((
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    [numpy.nan]*4)))

    def test_plate_array_two_inducers_with_resources(self):
        # Two inducer, three plate experiment
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        exp.plate_resources['Location'] = ['Stack 1-1',
                                           'Stack 1-2',
                                           'Stack 1-3',
                                           'Stack 1-4',
                                           'Stack 1-5',
                                           'Stack 1-6']
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [ 0,   0.5,  1,   2,   4,   8,
                               16,  32,   64, 128, 256, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0,  0.0005, 0.001, 0.005,
                              0.01, 0.05,   0.1,   0.5]
        exp.add_inducer(xyl)

        # Plates
        platearray = platedesign.plate.PlateArray(
            'PA1',
            array_n_rows=2,
            array_n_cols=2,
            plate_names=['P{}'.format(i+1) for i in range(4)],
            plate_n_rows=4,
            plate_n_cols=6)
        platearray.cell_strain_name = 'Test Strain 1'
        platearray.total_media_vol = 16000.*4
        platearray.sample_media_vol = 500.
        platearray.cell_setup_method = 'fixed_volume'
        platearray.cell_predilution = 100
        platearray.cell_predilution_vol = 1000
        platearray.cell_shot_vol = 5
        platearray.apply_inducer(inducer=iptg, apply_to='rows')
        platearray.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(platearray)

        plate5 = platedesign.plate.Plate('P5', n_rows=4, n_cols=6)
        plate5.samples_to_measure = 4
        plate5.cell_strain_name = 'Autofluorescence Strain'
        plate5.total_media_vol = 16000.
        plate5.sample_media_vol = 500.
        plate5.cell_setup_method = 'fixed_volume'
        plate5.cell_predilution = 10
        plate5.cell_predilution_vol = 1000
        plate5.cell_shot_vol = 5
        exp.add_plate(plate5)

        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 50)
        self.assertEqual(iptg.total_vol, 200)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 80)
        self.assertEqual(xyl.total_vol, 300)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate Array PA1",
                                             "Cells for Plate Array PA1",
                                             "Cells for Plate P5",
                                             "Plate Resources"
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = ['PA1', 'PA1', 'PA1', 'PA1', None,]
            summary_table_exp['Plate'] = ['P1', 'P2', 'P3', 'P4', 'P5']
            summary_table_exp['Strain'] = ['Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Autofluorescence Strain',]
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')

            platearray.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate Array PA1'],
                                    ws_2=wb_exp['Inducers for Plate Array PA1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate Array PA1'],
                                    ws_2=wb_exp['Cells for Plate Array PA1'])

            plate5.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P5'],
                                    ws_2=wb_exp['Cells for Plate P5'])
            # Check resources table
            resources_table = get_dataframe_from_worksheet(wb["Plate Resources"])
            resources_table_exp = pandas.DataFrame()
            resources_table_exp['Plate'] = ['P1', 'P2', 'P3', 'P4', 'P5']
            resources_table_exp['Location'] = ['Stack 1-1',
                                               'Stack 1-2',
                                               'Stack 1-3',
                                               'Stack 1-4',
                                               'Stack 1-5',]
            pandas.util.testing.assert_frame_equal(resources_table,
                                                   resources_table_exp)

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             ])
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 11)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1"]*24 + ["P2"]*24 + ["P3"]*24 + ["P4"]*24 + ["P5"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Plate Array'].values,
                ["PA1"]*96 + [None]*4)
            numpy.testing.assert_array_equal(
                samples_table['Location'].values,
                ["Stack 1-1"]*24 + ["Stack 1-2"]*24 + ["Stack 1-3"]*24 + \
                ["Stack 1-4"]*24 + ["Stack 1-5"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1"]*96 + ["Autofluorescence Strain"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100]*96 + [10]*4)
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [((i//6)%4 + 1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(100)])
            numpy.testing.assert_almost_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.concatenate((
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    [numpy.nan]*4)))
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                numpy.concatenate((
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    [numpy.nan]*4)))

    def test_plate_array_two_inducers_with_random_resources(self):
        # Two inducer, three plate experiment
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        exp.plate_resources['Location'] = ['Stack 1-1',
                                           'Stack 1-2',
                                           'Stack 1-3',
                                           'Stack 1-4',
                                           'Stack 1-5',
                                           'Stack 1-6']
        exp.randomize_plate_resources = True
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [ 0,   0.5,  1,   2,   4,   8,
                               16,  32,   64, 128, 256, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0,  0.0005, 0.001, 0.005,
                              0.01, 0.05,   0.1,   0.5]
        exp.add_inducer(xyl)

        # Plates
        platearray = platedesign.plate.PlateArray(
            'PA1',
            array_n_rows=2,
            array_n_cols=2,
            plate_names=['P{}'.format(i+1) for i in range(4)],
            plate_n_rows=4,
            plate_n_cols=6)
        platearray.cell_strain_name = 'Test Strain 1'
        platearray.total_media_vol = 16000.*4
        platearray.sample_media_vol = 500.
        platearray.cell_setup_method = 'fixed_volume'
        platearray.cell_predilution = 100
        platearray.cell_predilution_vol = 1000
        platearray.cell_shot_vol = 5
        platearray.apply_inducer(inducer=iptg, apply_to='rows')
        platearray.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(platearray)

        plate5 = platedesign.plate.Plate('P5', n_rows=4, n_cols=6)
        plate5.samples_to_measure = 4
        plate5.cell_strain_name = 'Autofluorescence Strain'
        plate5.total_media_vol = 16000.
        plate5.sample_media_vol = 500.
        plate5.cell_setup_method = 'fixed_volume'
        plate5.cell_predilution = 10
        plate5.cell_predilution_vol = 1000
        plate5.cell_shot_vol = 5
        exp.add_plate(plate5)

        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 50)
        self.assertEqual(iptg.total_vol, 200)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 80)
        self.assertEqual(xyl.total_vol, 300)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        if six.PY2:
            location_idx = [[1, 2, 5, 3, 4],
                            [1, 5, 0, 4, 3],
                            [1, 4, 0, 3, 2]]
        elif six.PY3:
            location_idx = [[2, 3, 5, 0, 4],
                            [0, 1, 2, 4, 5],
                            [2, 5, 4, 3, 0]]
        locations = ['Stack 1-1',
                     'Stack 1-2',
                     'Stack 1-3',
                     'Stack 1-4',
                     'Stack 1-5',
                     'Stack 1-6']
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate Array PA1",
                                             "Cells for Plate Array PA1",
                                             "Cells for Plate P5",
                                             "Plate Resources"
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = ['PA1', 'PA1', 'PA1', 'PA1', None,]
            summary_table_exp['Plate'] = ['P1', 'P2', 'P3', 'P4', 'P5']
            summary_table_exp['Strain'] = ['Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Autofluorescence Strain',]
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')

            platearray.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate Array PA1'],
                                    ws_2=wb_exp['Inducers for Plate Array PA1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate Array PA1'],
                                    ws_2=wb_exp['Cells for Plate Array PA1'])

            plate5.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P5'],
                                    ws_2=wb_exp['Cells for Plate P5'])
            # Check resources table
            resources_table = get_dataframe_from_worksheet(wb["Plate Resources"])
            resources_table_exp = pandas.DataFrame()
            resources_table_exp['Plate'] = ['P1', 'P2', 'P3', 'P4', 'P5']
            resources_table_exp['Location'] = [locations[i]
                                               for i in location_idx[rep_idx]]
            pandas.util.testing.assert_frame_equal(resources_table,
                                                   resources_table_exp)

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             ])
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 11)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1"]*24 + ["P2"]*24 + ["P3"]*24 + ["P4"]*24 + ["P5"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Plate Array'].values,
                ["PA1"]*96 + [None]*4)
            locations_rep = [locations[i] for i in location_idx[rep_idx]]
            numpy.testing.assert_array_equal(
                samples_table['Location'].values,
                [locations_rep[0]]*24 + [locations_rep[1]]*24 + \
                [locations_rep[2]]*24 + [locations_rep[3]]*24 + \
                [locations_rep[4]]*4)
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1"]*96 + ["Autofluorescence Strain"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100]*96 + [10]*4)
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [((i//6)%4 + 1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(100)])
            numpy.testing.assert_almost_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.concatenate((
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    [numpy.nan]*4)))
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                numpy.concatenate((
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    [numpy.nan]*4)))

    def test_plate_array_two_inducers_with_random_resources_and_sort(self):
        # Two inducer, three plate experiment
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        exp.plate_resources['Location'] = ['Stack 1-1',
                                           'Stack 1-2',
                                           'Stack 1-3',
                                           'Stack 1-4',
                                           'Stack 1-5',
                                           'Stack 1-6']
        exp.randomize_plate_resources = True
        exp.measurement_order = 'Location'
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [ 0,   0.5,  1,   2,   4,   8,
                               16,  32,   64, 128, 256, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0,  0.0005, 0.001, 0.005,
                              0.01, 0.05,   0.1,   0.5]
        exp.add_inducer(xyl)

        # Plates
        platearray = platedesign.plate.PlateArray(
            'PA1',
            array_n_rows=2,
            array_n_cols=2,
            plate_names=['P{}'.format(i+1) for i in range(4)],
            plate_n_rows=4,
            plate_n_cols=6)
        platearray.cell_strain_name = 'Test Strain 1'
        platearray.total_media_vol = 16000.*4
        platearray.sample_media_vol = 500.
        platearray.cell_setup_method = 'fixed_volume'
        platearray.cell_predilution = 100
        platearray.cell_predilution_vol = 1000
        platearray.cell_shot_vol = 5
        platearray.apply_inducer(inducer=iptg, apply_to='rows')
        platearray.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(platearray)

        plate5 = platedesign.plate.Plate('P5', n_rows=4, n_cols=6)
        plate5.samples_to_measure = 4
        plate5.cell_strain_name = 'Autofluorescence Strain'
        plate5.total_media_vol = 16000.
        plate5.sample_media_vol = 500.
        plate5.cell_setup_method = 'fixed_volume'
        plate5.cell_predilution = 10
        plate5.cell_predilution_vol = 1000
        plate5.cell_shot_vol = 5
        exp.add_plate(plate5)

        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 50)
        self.assertEqual(iptg.total_vol, 200)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 80)
        self.assertEqual(xyl.total_vol, 300)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        if six.PY2:
            location_idx = [[1, 2, 5, 3, 4],
                            [1, 5, 0, 4, 3],
                            [1, 4, 0, 3, 2]]
        elif six.PY3:
            location_idx = [[2, 3, 5, 0, 4],
                            [0, 1, 2, 4, 5],
                            [2, 5, 4, 3, 0]]
        locations = ['Stack 1-1',
                     'Stack 1-2',
                     'Stack 1-3',
                     'Stack 1-4',
                     'Stack 1-5',
                     'Stack 1-6']
        plates = ['P1', 'P2', 'P3', 'P4', 'P5']
        samples = [24, 24, 24, 24, 4]
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate Array PA1",
                                             "Cells for Plate Array PA1",
                                             "Cells for Plate P5",
                                             "Plate Resources"
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = ['PA1', 'PA1', 'PA1', 'PA1', None,]
            summary_table_exp['Plate'] = ['P1', 'P2', 'P3', 'P4', 'P5']
            summary_table_exp['Strain'] = ['Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Autofluorescence Strain',]
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')

            platearray.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate Array PA1'],
                                    ws_2=wb_exp['Inducers for Plate Array PA1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate Array PA1'],
                                    ws_2=wb_exp['Cells for Plate Array PA1'])

            plate5.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P5'],
                                    ws_2=wb_exp['Cells for Plate P5'])
            # Check resources table
            resources_table = get_dataframe_from_worksheet(wb["Plate Resources"])
            resources_table_exp = pandas.DataFrame()
            locations_unsorted = [locations[i] for i in location_idx[rep_idx]]
            plates_rep = [plates[i] for i in numpy.argsort(locations_unsorted)]
            locations_rep = [locations_unsorted[i]
                             for i in numpy.argsort(locations_unsorted)]
            resources_table_exp['Plate'] = plates_rep
            resources_table_exp['Location'] = locations_rep
            pandas.util.testing.assert_frame_equal(resources_table,
                                                   resources_table_exp)

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             ])
            # Calculate locations, plate order, and number of samples
            locations_unsorted = [locations[i] for i in location_idx[rep_idx]]
            plates_rep = [plates[i] for i in numpy.argsort(locations_unsorted)]
            locations_rep = [locations_unsorted[i]
                             for i in numpy.argsort(locations_unsorted)]
            samples_rep = [samples[i]
                           for i in numpy.argsort(locations_unsorted)]
            # Calculate expected values in samples table
            plate_exp = []
            plate_array_exp = []
            locations_exp = []
            strain_exp = []
            pre_dilution_exp = []
            cell_vol_exp = []
            rows_exp = []
            cols_exp = []
            iptg_exp = []
            xyl_exp = []
            for plate, location, n_samples in \
                    zip(plates_rep, locations_rep, samples_rep):
                # Define values that depend on plate
                if plate=="P5":
                    plate_array = None
                    strain = "Autofluorescence Strain"
                    pre_dilution = 10
                    cell_vol = 5
                    rows = numpy.array([1,1,1,1])
                    cols = numpy.array([1,2,3,4])
                    iptg_conc = [numpy.nan]*4
                    xyl_conc = [numpy.nan]*4
                else:
                    plate_array = "PA1"
                    strain = "Test Strain 1"
                    pre_dilution = 100
                    cell_vol = 5
                    rows = numpy.repeat(numpy.arange(4) + 1, 6)
                    cols = numpy.tile(numpy.arange(6) + 1, 4)
                    if plate=="P1":
                        iptg_conc = numpy.tile(iptg.concentrations[:6], 4)
                        xyl_conc = numpy.repeat(xyl.concentrations[:4], 6)
                    elif plate=="P2":
                        iptg_conc = numpy.tile(iptg.concentrations[6:], 4)
                        xyl_conc = numpy.repeat(xyl.concentrations[:4], 6)
                    elif plate=="P3":
                        iptg_conc = numpy.tile(iptg.concentrations[:6], 4)
                        xyl_conc = numpy.repeat(xyl.concentrations[4:], 6)
                    elif plate=="P4":
                        iptg_conc = numpy.tile(iptg.concentrations[6:], 4)
                        xyl_conc = numpy.repeat(xyl.concentrations[4:], 6)
                # Add to lists
                plate_exp.extend([plate]*n_samples)
                plate_array_exp.extend([plate_array]*n_samples)
                locations_exp.extend([location]*n_samples)
                strain_exp.extend([strain]*n_samples)
                pre_dilution_exp.extend([pre_dilution]*n_samples)
                cell_vol_exp.extend([cell_vol]*n_samples)
                rows_exp.extend(rows)
                cols_exp.extend(cols)
                iptg_exp.extend(iptg_conc)
                xyl_exp.extend(xyl_conc)
            # Check samples table
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 11)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                plate_exp)
            numpy.testing.assert_array_equal(
                samples_table['Plate Array'].values,
                plate_array_exp)
            numpy.testing.assert_array_equal(
                samples_table['Location'].values,
                locations_exp)
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                strain_exp)
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                pre_dilution_exp)
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                cell_vol_exp)
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                rows_exp)
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                cols_exp)
            numpy.testing.assert_almost_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                iptg_exp)
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                xyl_exp)

    def test_plate_array_two_inducers_plate_measurements(self):
        # Two inducer, three plate experiment
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        exp.plate_measurements = ['Final OD600', 'Incubation time (min)']
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [ 0,   0.5,  1,   2,   4,   8,
                               16,  32,   64, 128, 256, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0,  0.0005, 0.001, 0.005,
                              0.01, 0.05,   0.1,   0.5]
        exp.add_inducer(xyl)

        # Plates
        platearray = platedesign.plate.PlateArray(
            'PA1',
            array_n_rows=2,
            array_n_cols=2,
            plate_names=['P{}'.format(i+1) for i in range(4)],
            plate_n_rows=4,
            plate_n_cols=6)
        platearray.cell_strain_name = 'Test Strain 1'
        platearray.total_media_vol = 16000.*4
        platearray.sample_media_vol = 500.
        platearray.cell_setup_method = 'fixed_volume'
        platearray.cell_predilution = 100
        platearray.cell_predilution_vol = 1000
        platearray.cell_shot_vol = 5
        platearray.apply_inducer(inducer=iptg, apply_to='rows')
        platearray.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(platearray)

        plate5 = platedesign.plate.Plate('P5', n_rows=4, n_cols=6)
        plate5.samples_to_measure = 4
        plate5.cell_strain_name = 'Autofluorescence Strain'
        plate5.total_media_vol = 16000.
        plate5.sample_media_vol = 500.
        plate5.cell_setup_method = 'fixed_volume'
        plate5.cell_predilution = 10
        plate5.cell_predilution_vol = 1000
        plate5.cell_shot_vol = 5
        exp.add_plate(plate5)

        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 50)
        self.assertEqual(iptg.total_vol, 200)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 80)
        self.assertEqual(xyl.total_vol, 300)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate Array PA1",
                                             "Cells for Plate Array PA1",
                                             "Cells for Plate P5",
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = ['PA1', 'PA1', 'PA1', 'PA1', None,]
            summary_table_exp['Plate'] = ['P1', 'P2', 'P3', 'P4', 'P5']
            summary_table_exp['Strain'] = ['Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Autofluorescence Strain',]
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')

            platearray.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate Array PA1'],
                                    ws_2=wb_exp['Inducers for Plate Array PA1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate Array PA1'],
                                    ws_2=wb_exp['Cells for Plate Array PA1'])

            plate5.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P5'],
                                    ws_2=wb_exp['Cells for Plate P5'])

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             "Plate Measurements"])
            # Check samples table
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 10)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1"]*24 + ["P2"]*24 + ["P3"]*24 + ["P4"]*24 + ["P5"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Plate Array'].values,
                ["PA1"]*96 + [None]*4)
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1"]*96 + ["Autofluorescence Strain"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100]*96 + [10]*4)
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [((i//6)%4 + 1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(100)])
            numpy.testing.assert_almost_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.concatenate((
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    [numpy.nan]*4)))
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                numpy.concatenate((
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    [numpy.nan]*4)))
            # Check plate measurements table
            plate_measure_table = \
                get_dataframe_from_worksheet(wb["Plate Measurements"])
            self.assertEqual(len(plate_measure_table.columns), 3)
            numpy.testing.assert_array_equal(
                plate_measure_table['Plate'].values,
                ["P1", "P2", "P3", "P4", "P5"])
            numpy.testing.assert_array_equal(
                plate_measure_table['Final OD600'].values,
                [None]*5)
            numpy.testing.assert_array_equal(
                plate_measure_table['Incubation time (min)'].values,
                [None]*5)

    def test_plate_array_two_inducers_replicate_measurements(self):
        # Two inducer, three plate experiment
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        exp.replicate_measurements = ['Date', 'Run by']
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [ 0,   0.5,  1,   2,   4,   8,
                               16,  32,   64, 128, 256, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0,  0.0005, 0.001, 0.005,
                              0.01, 0.05,   0.1,   0.5]
        exp.add_inducer(xyl)

        # Plates
        platearray = platedesign.plate.PlateArray(
            'PA1',
            array_n_rows=2,
            array_n_cols=2,
            plate_names=['P{}'.format(i+1) for i in range(4)],
            plate_n_rows=4,
            plate_n_cols=6)
        platearray.cell_strain_name = 'Test Strain 1'
        platearray.total_media_vol = 16000.*4
        platearray.sample_media_vol = 500.
        platearray.cell_setup_method = 'fixed_volume'
        platearray.cell_predilution = 100
        platearray.cell_predilution_vol = 1000
        platearray.cell_shot_vol = 5
        platearray.apply_inducer(inducer=iptg, apply_to='rows')
        platearray.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(platearray)

        plate5 = platedesign.plate.Plate('P5', n_rows=4, n_cols=6)
        plate5.samples_to_measure = 4
        plate5.cell_strain_name = 'Autofluorescence Strain'
        plate5.total_media_vol = 16000.
        plate5.sample_media_vol = 500.
        plate5.cell_setup_method = 'fixed_volume'
        plate5.cell_predilution = 10
        plate5.cell_predilution_vol = 1000
        plate5.cell_shot_vol = 5
        exp.add_plate(plate5)

        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 50)
        self.assertEqual(iptg.total_vol, 200)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 80)
        self.assertEqual(xyl.total_vol, 300)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate Array PA1",
                                             "Cells for Plate Array PA1",
                                             "Cells for Plate P5",
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = ['PA1', 'PA1', 'PA1', 'PA1', None,]
            summary_table_exp['Plate'] = ['P1', 'P2', 'P3', 'P4', 'P5']
            summary_table_exp['Strain'] = ['Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Autofluorescence Strain',]
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')

            platearray.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate Array PA1'],
                                    ws_2=wb_exp['Inducers for Plate Array PA1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate Array PA1'],
                                    ws_2=wb_exp['Cells for Plate Array PA1'])

            plate5.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P5'],
                                    ws_2=wb_exp['Cells for Plate P5'])

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Samples",
                                             "Replicate Measurements"])
            # Check samples table
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 10)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1"]*24 + ["P2"]*24 + ["P3"]*24 + ["P4"]*24 + ["P5"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Plate Array'].values,
                ["PA1"]*96 + [None]*4)
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1"]*96 + ["Autofluorescence Strain"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100]*96 + [10]*4)
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [((i//6)%4 + 1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(100)])
            numpy.testing.assert_almost_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.concatenate((
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    [numpy.nan]*4)))
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                numpy.concatenate((
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    [numpy.nan]*4)))
            # Check replicate measurements table
            self.assertEqual(len(list(wb["Replicate Measurements"].rows)), 2)
            self.assertEqual(len(list(wb["Replicate Measurements"].columns)), 2)
            self.assertEqual(wb["Replicate Measurements"]["A1"].value, 'Date')
            self.assertEqual(wb["Replicate Measurements"]["A2"].value, 'Run by')
            self.assertEqual(wb["Replicate Measurements"]["B1"].value, None)
            self.assertEqual(wb["Replicate Measurements"]["B2"].value, None)

    def test_plate_array_two_inducers_measurement_template(self):
        # Two inducer, three plate experiment
        exp = platedesign.experiment.Experiment()
        exp.n_replicates = 3
        exp.measurement_template = 'test/test_experiment_files/template_FlowCal.xlsx'
        # Inducer
        iptg = platedesign.inducer.ChemicalInducer(name='IPTG', units=u'µM')
        iptg.stock_conc = 1e6
        iptg.shot_vol = 5.
        iptg.concentrations = [ 0,   0.5,  1,   2,   4,   8,
                               16,  32,   64, 128, 256, 500]
        exp.add_inducer(iptg)
        xyl = platedesign.inducer.ChemicalInducer(name='Xylose', units='%')
        xyl.stock_conc = 50.
        xyl.shot_vol = 5.
        xyl.concentrations = [0.0,  0.0005, 0.001, 0.005,
                              0.01, 0.05,   0.1,   0.5]
        exp.add_inducer(xyl)

        # Plates
        platearray = platedesign.plate.PlateArray(
            'PA1',
            array_n_rows=2,
            array_n_cols=2,
            plate_names=['P{}'.format(i+1) for i in range(4)],
            plate_n_rows=4,
            plate_n_cols=6)
        platearray.cell_strain_name = 'Test Strain 1'
        platearray.total_media_vol = 16000.*4
        platearray.sample_media_vol = 500.
        platearray.cell_setup_method = 'fixed_volume'
        platearray.cell_predilution = 100
        platearray.cell_predilution_vol = 1000
        platearray.cell_shot_vol = 5
        platearray.apply_inducer(inducer=iptg, apply_to='rows')
        platearray.apply_inducer(inducer=xyl, apply_to='cols')
        exp.add_plate(platearray)

        plate5 = platedesign.plate.Plate('P5', n_rows=4, n_cols=6)
        plate5.samples_to_measure = 4
        plate5.cell_strain_name = 'Autofluorescence Strain'
        plate5.total_media_vol = 16000.
        plate5.sample_media_vol = 500.
        plate5.cell_setup_method = 'fixed_volume'
        plate5.cell_predilution = 10
        plate5.cell_predilution_vol = 1000
        plate5.cell_shot_vol = 5
        exp.add_plate(plate5)

        # Generate experiment files
        exp.generate(path=self.temp_dir)

        # Check that inducers have the right number of replicates for their
        # calculations
        self.assertEqual(iptg.media_vol, 500)
        self.assertEqual(iptg.replicate_vol, 50)
        self.assertEqual(iptg.total_vol, 200)
        self.assertEqual(xyl.media_vol, 500)
        self.assertEqual(xyl.replicate_vol, 80)
        self.assertEqual(xyl.total_vol, 300)

        # Check sheet names of experiment setup file
        wb = openpyxl.load_workbook(filename=os.path.join(
            self.temp_dir,
            'experiment_setup.xlsx'))
        self.assertEqual(wb.sheetnames, ["IPTG", "Xylose"])
        # Check proper inducer preparation instructions
        # Instructions will be compared to the ones generated by the inducer
        # class directly.
        file_name = os.path.join(self.temp_dir, 'IPTG_test.xlsx')
        iptg.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['IPTG'], ws_2=wb_exp['IPTG'])

        file_name = os.path.join(self.temp_dir, 'Xylose_test.xlsx')
        xyl.save_exp_setup_instructions(file_name=file_name)
        wb_exp = openpyxl.load_workbook(file_name)
        assert_worksheets_equal(self, ws_1=wb['Xylose'], ws_2=wb_exp['Xylose'])

        # Check replicate setup files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_setup.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Summary",
                                             "Inducers for Plate Array PA1",
                                             "Cells for Plate Array PA1",
                                             "Cells for Plate P5",
                                             ])
            # Check summary table
            summary_table = get_dataframe_from_worksheet(wb["Summary"])
            summary_table_exp = pandas.DataFrame()
            summary_table_exp['Plate Array'] = ['PA1', 'PA1', 'PA1', 'PA1', None,]
            summary_table_exp['Plate'] = ['P1', 'P2', 'P3', 'P4', 'P5']
            summary_table_exp['Strain'] = ['Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Test Strain 1',
                                           'Autofluorescence Strain',]
            pandas.util.testing.assert_frame_equal(summary_table,
                                                   summary_table_exp)
            # Check proper plate setup instructions
            # Instructions will be compared to the ones generated by the plate
            # class directly
            file_name = os.path.join(self.temp_dir, 'plate_rep.xlsx')

            platearray.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Inducers for Plate Array PA1'],
                                    ws_2=wb_exp['Inducers for Plate Array PA1'])
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate Array PA1'],
                                    ws_2=wb_exp['Cells for Plate Array PA1'])

            plate5.save_rep_setup_instructions(file_name=file_name)
            wb_exp = openpyxl.load_workbook(file_name)
            assert_worksheets_equal(self, 
                                    ws_1=wb['Cells for Plate P5'],
                                    ws_2=wb_exp['Cells for Plate P5'])

        # Check replicate measurement files
        for rep_idx in range(3):
            wb = openpyxl.load_workbook(filename=os.path.join(
                self.temp_dir,
                'replicate_{:03d}'.format(rep_idx+1),
                'replicate_{:03d}_measurement.xlsx'.format(rep_idx+1),))
            self.assertEqual(wb.sheetnames, ["Instruments",
                                             "Beads",
                                             "Samples",
                                             ])
            # Check instruments and beads table
            assert_worksheets_equal(
                self,
                wb_name_1='test/test_experiment_files/template_FlowCal.xlsx',
                ws_name_1='Instruments',
                ws_2=wb['Instruments'])
            assert_worksheets_equal(
                self,
                wb_name_1='test/test_experiment_files/template_FlowCal.xlsx',
                ws_name_1='Beads',
                ws_2=wb['Beads'])
            # Check samples table
            samples_table = get_dataframe_from_worksheet(wb["Samples"])
            self.assertEqual(len(samples_table.columns), 17)
            numpy.testing.assert_array_equal(
                samples_table['ID'].values,
                ["S{:04d}".format(i+1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Plate'].values,
                ["P1"]*24 + ["P2"]*24 + ["P3"]*24 + ["P4"]*24 + ["P5"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Plate Array'].values,
                ["PA1"]*96 + [None]*4)
            numpy.testing.assert_array_equal(
                samples_table['Strain'].values,
                ["Test Strain 1"]*96 + ["Autofluorescence Strain"]*4)
            numpy.testing.assert_array_equal(
                samples_table['Preculture/Aliquot Dilution'].values,
                [100]*96 + [10]*4)
            numpy.testing.assert_array_equal(
                samples_table['Cell Inoculated Vol.'].values,
                [5 for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Row'].values,
                [((i//6)%4 + 1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['Column'].values,
                [(i%6 + 1) for i in range(100)])
            numpy.testing.assert_almost_equal(
                samples_table[u'IPTG Concentration (µM)'].values,
                numpy.concatenate((
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    numpy.tile(iptg.concentrations[:6], 4),
                    numpy.tile(iptg.concentrations[6:], 4),
                    [numpy.nan]*4)))
            numpy.testing.assert_array_equal(
                samples_table['Xylose Concentration (%)'].values,
                numpy.concatenate((
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[:4], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    numpy.repeat(xyl.concentrations[4:], 6),
                    [numpy.nan]*4)))
            numpy.testing.assert_array_equal(
                samples_table['Instrument ID'].values,
                ["FC001"]*100)
            numpy.testing.assert_array_equal(
                samples_table['Beads ID'].values,
                ["B0001"]*100)
            numpy.testing.assert_array_equal(
                samples_table['File Path'].values,
                ["FCFiles/Data{:03d}.fcs".format(i+1) for i in range(100)])
            numpy.testing.assert_array_equal(
                samples_table['FL1 Units'].values,
                ["MEF"]*100)
            numpy.testing.assert_array_equal(
                samples_table['Gate Fraction'].values,
                [0.5]*100)
            numpy.testing.assert_array_equal(
                samples_table[u'Flow Sample Volume (µL)'].values,
                [100]*100)
            numpy.testing.assert_array_equal(
                samples_table[u'Flow PBS Volume (µL)'].values,
                [1000]*100)

