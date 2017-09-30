# -*- coding: UTF-8 -*-
"""
Unit tests for plate classes

"""

import collections
import itertools
import os
import random
import shutil
import unittest

import numpy
import openpyxl
import pandas

import platedesign

class TestPlate(unittest.TestCase):
    """
    Tests for the Plate class

    """
    def setUp(self):
        # Directory where to save temporary files
        self.temp_dir = "test/temp_plate"
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)
        # Set random seed
        random.seed(1)

    def tearDown(self):
        # Delete temporary directory
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def test_create(self):
        p = platedesign.plate.Plate(name='P1')

    def test_default_attributes(self):
        p = platedesign.plate.Plate(name='P1')
        # Check all attributes
        self.assertEqual(p.name, 'P1')
        self.assertEqual(p.n_rows, 4)
        self.assertEqual(p.n_cols, 6)
        self.assertEqual(p.samples_to_measure, 24)
        self.assertIsNone(p.sample_media_vol)
        self.assertIsNone(p.total_media_vol)
        self.assertIsNone(p.cell_strain_name)
        self.assertIsNone(p.cell_setup_method)
        self.assertEqual(p.cell_predilution, 1)
        self.assertIsNone(p.cell_predilution_vol)
        self.assertIsNone(p.cell_initial_od600)
        self.assertIsNone(p.cell_shot_vol)
        self.assertEqual(p.metadata, collections.OrderedDict())
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_non_default_attributes(self):
        p = platedesign.plate.Plate(name='P1', n_rows=8, n_cols=12)
        # Check all attributes
        self.assertEqual(p.name, 'P1')
        self.assertEqual(p.n_rows, 8)
        self.assertEqual(p.n_cols, 12)
        self.assertEqual(p.samples_to_measure, 96)
        self.assertIsNone(p.sample_media_vol)
        self.assertIsNone(p.total_media_vol)
        self.assertIsNone(p.cell_strain_name)
        self.assertIsNone(p.cell_setup_method)
        self.assertEqual(p.cell_predilution, 1)
        self.assertIsNone(p.cell_predilution_vol)
        self.assertIsNone(p.cell_initial_od600)
        self.assertIsNone(p.cell_shot_vol)
        self.assertEqual(p.metadata, collections.OrderedDict())
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_media_vol_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Set media volume
        p.total_media_vol = 13000
        p.sample_media_vol = 500
        # Get media volume for inducer application
        self.assertEqual(p.apply_inducer_media_vol('rows'), 500)
        self.assertEqual(p.apply_inducer_media_vol('cols'), 500)
        self.assertEqual(p.apply_inducer_media_vol('wells'), 500)
        self.assertEqual(p.apply_inducer_media_vol('media'), 13000)

    def test_apply_inducer_media_vol_error_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Set media volume
        p.total_media_vol = 13000
        p.sample_media_vol = 500
        # Limit number of samples to measure
        p.samples_to_measure = 12
        # Get media volume for inducer application
        self.assertEqual(p.apply_inducer_media_vol('wells'), 500)
        self.assertEqual(p.apply_inducer_media_vol('media'), 13000)
        # Verify that exception is raised for 'rows' and 'cols'
        self.assertRaises(ValueError, p.apply_inducer_media_vol, 'rows')
        self.assertRaises(ValueError, p.apply_inducer_media_vol, 'cols')

    def test_apply_inducer_media_vol_error_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Set media volume
        p.total_media_vol = 13000
        p.sample_media_vol = 500
        # Verify that exception is raised for an invalid 'apply_to' argument
        self.assertRaises(ValueError, p.apply_inducer_media_vol, 'all')

    def test_apply_inducer_media_vol_error_3(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Only set total media volume
        p.total_media_vol = 13000
        # Verify that exception is raised
        self.assertRaises(AttributeError, p.apply_inducer_media_vol, 'rows')

    def test_apply_inducer_media_vol_error_4(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Only set sample media volume
        p.sample_media_vol = 500
        # Verify that exception is raised
        self.assertRaises(AttributeError, p.apply_inducer_media_vol, 'rows')

    def test_apply_inducer_media_vol_error_5(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Don't set any media volume
        # Verify that exception is raised
        self.assertRaises(AttributeError, p.apply_inducer_media_vol, 'rows')

    def test_apply_inducer_n_shots(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Get number of shots for inducer application
        self.assertEqual(p.apply_inducer_n_shots('rows'), 4)
        self.assertEqual(p.apply_inducer_n_shots('cols'), 6)
        self.assertEqual(p.apply_inducer_n_shots('wells'), 1)
        self.assertEqual(p.apply_inducer_n_shots('media'), 1)

    def test_apply_inducer_n_shots_error_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Limit number of samples to measure
        p.samples_to_measure = 12
        # Get number of shots for inducer application
        self.assertEqual(p.apply_inducer_n_shots('wells'), 1)
        self.assertEqual(p.apply_inducer_n_shots('media'), 1)
        # Verify that exception is raised for 'rows' and 'cols'
        self.assertRaises(ValueError, p.apply_inducer_n_shots, 'rows')
        self.assertRaises(ValueError, p.apply_inducer_n_shots, 'cols')

    def test_apply_inducer_n_shots_error_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Verify that exception is raised for an invalid 'apply_to' argument
        self.assertRaises(ValueError, p.apply_inducer_n_shots, 'all')

    def test_apply_inducer_rows(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=6, scale='log')
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='rows')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [iptg],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_rows_error_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=4, scale='log')
        # Apply inducer to plate
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='rows')
        # Check that inducers dictionary in plate has not been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_rows_error_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Limit number of samples to measure
        p.samples_to_measure = 12
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=6, scale='log')
        # Apply inducer to plate
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='rows')
        # Check that inducers dictionary in plate has not been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_cols(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=4, scale='log')
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='cols')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [iptg],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_cols_error_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=6, scale='log')
        # Apply inducer to plate
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='cols')
        # Check that inducers dictionary in plate has not been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_cols_error_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Limit number of samples to measure
        p.samples_to_measure = 12
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=4, scale='log')
        # Apply inducer to plate
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='cols')
        # Check that inducers dictionary in plate has not been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_wells(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=24, scale='log')
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='wells')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [iptg],
                                      'media': []})

    def test_apply_inducer_wells_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Limit number of samples to measure
        p.samples_to_measure = 12
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=12, scale='log')
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='wells')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [iptg],
                                      'media': []})

    def test_apply_inducer_wells_error_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=15, scale='log')
        # Apply inducer to plate
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='wells')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_wells_error_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Limit number of samples to measure
        p.samples_to_measure = 12
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=15, scale='log')
        # Apply inducer to plate
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='wells')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_media_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations
        iptg.concentrations = [1]
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='media')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': [iptg]})

    def test_apply_inducer_media_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Limit number of samples to measure
        p.samples_to_measure = 12
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations
        iptg.concentrations = [1]
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='media')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': [iptg]})

    def test_apply_inducer_media_error_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations
        iptg.concentrations = [1, 10]
        # Apply inducer to plate
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='media')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_error_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations
        iptg.concentrations = [1]
        # Apply inducer to plate
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='all')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_save_exp_setup_instructions_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Run save_exp_setup_instructions
        p.save_exp_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_exp.xlsx'))
        # save_exp_setup_instructions does not do anything in Plate. There is
        # no need to check for results.

    def test_save_exp_setup_instructions_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create new spreadsheet
        wb_test = openpyxl.Workbook()
        # Remove sheet created by default
        wb_test.remove_sheet(wb_test.active)
        # Run save_exp_setup_instructions
        p.save_exp_setup_instructions(workbook=wb_test)
        # save_exp_setup_instructions does not do anything in Plate. There is
        # no need to check for results.

    def test_save_exp_setup_instructions_argument_error(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Run save_exp_setup_instructions with no arguments
        self.assertRaises(ValueError, p.save_exp_setup_instructions)

    def test_save_exp_setup_files(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Check that save_exp_setup_files runs successfully
        p.save_exp_setup_files()
        # save_exp_setup_files does not do anything in Plate. There is no need
        # to check for results.

    def test_save_rep_setup_instructions_empty(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should only contain an empty sheet named "Sheet 1"
        self.assertEqual(wb.sheetnames, ["Sheet 1"])

    def test_save_rep_setup_instructions_cell_setup_fixed_volume_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.total_media_vol = 15000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_volume'
        p.cell_shot_vol = 5
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Cells for Plate P1"])
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertEqual(ws.cell(row=2, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=2, column=2).value, 5)
        self.assertEqual(ws.cell(row=2, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=3, column=1).value, "Add into 15.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_cell_setup_fixed_volume_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.total_media_vol = 15000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_volume'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_shot_vol = 5
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Cells for Plate P1"])
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertTrue('A2:C2' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=2, column=1).value, "Predilution")
        self.assertEqual(ws.cell(row=3, column=1).value, "Predilution factor")
        self.assertEqual(ws.cell(row=3, column=2).value, 100)
        self.assertEqual(ws.cell(row=3, column=3).value, "x")
        self.assertEqual(ws.cell(row=4, column=1).value, "Media volume")
        self.assertEqual(ws.cell(row=4, column=2).value, 990)
        self.assertEqual(ws.cell(row=4, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=5, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=5, column=2).value, 10)
        self.assertEqual(ws.cell(row=5, column=3).value, u"µL")
        self.assertTrue('A6:C6' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=6, column=1).value, "Inoculation")
        self.assertEqual(ws.cell(row=7, column=1).value, "Predilution volume")
        self.assertEqual(ws.cell(row=7, column=2).value, 5)
        self.assertEqual(ws.cell(row=7, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=8, column=1).value, "Add into 15.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_cell_setup_fixed_volume_error_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.total_media_vol = 15000.
        # Add some information for cell setup
        # Do not include shot volume
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_volume'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        # Run save_rep_setup_instructions
        self.assertRaises(
            ValueError,
            p.save_rep_setup_instructions,
            file_name=os.path.join(self.temp_dir, 'plate_rep.xlsx'))

    def test_save_rep_setup_instructions_cell_setup_fixed_volume_error_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.total_media_vol = 15000.
        # Add some information for cell setup
        # Do not include predilution volume
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_volume'
        p.cell_predilution = 100
        p.cell_shot_vol = 5
        # Run save_rep_setup_instructions
        self.assertRaises(
            ValueError,
            p.save_rep_setup_instructions,
            file_name=os.path.join(self.temp_dir, 'plate_rep.xlsx'))

    def test_save_rep_setup_instructions_cell_setup_fixed_od600_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.total_media_vol = 15000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_initial_od600 = 1e-5
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Cells for Plate P1"])
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertEqual(ws.cell(row=2, column=1).value, "Preculture OD600")
        self.assertEqual(ws.cell(row=2, column=2).value, None)
        self.assertEqual(ws.cell(row=2, column=3).value, None)
        self.assertEqual(ws.cell(row=3, column=1).value, "Target OD600")
        self.assertEqual(ws.cell(row=3, column=2).value, 1e-5)
        self.assertEqual(ws.cell(row=3, column=3).value, None)
        self.assertEqual(ws.cell(row=4, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=4, column=2).value, "=0.15/B2")
        self.assertEqual(ws.cell(row=4, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=5, column=1).value, "Add into 15.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_cell_setup_fixed_od600_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.total_media_vol = 15000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_initial_od600 = 1e-5
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Cells for Plate P1"])
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertTrue('A2:C2' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=2, column=1).value, "Predilution")
        self.assertEqual(ws.cell(row=3, column=1).value, "Predilution factor")
        self.assertEqual(ws.cell(row=3, column=2).value, 100)
        self.assertEqual(ws.cell(row=3, column=3).value, "x")
        self.assertEqual(ws.cell(row=4, column=1).value, "Media volume")
        self.assertEqual(ws.cell(row=4, column=2).value, 990)
        self.assertEqual(ws.cell(row=4, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=5, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=5, column=2).value, 10)
        self.assertEqual(ws.cell(row=5, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=6, column=1).value, "Predilution OD600")
        self.assertEqual(ws.cell(row=6, column=2).value, None)
        self.assertEqual(ws.cell(row=6, column=3).value, None)
        self.assertTrue('A7:C7' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=7, column=1).value, "Inoculation")
        self.assertEqual(ws.cell(row=8, column=1).value, "Target OD600")
        self.assertEqual(ws.cell(row=8, column=2).value, 1e-5)
        self.assertEqual(ws.cell(row=8, column=3).value, None)
        self.assertEqual(ws.cell(row=9, column=1).value, "Predilution volume")
        self.assertEqual(ws.cell(row=9, column=2).value, "=0.15/B6")
        self.assertEqual(ws.cell(row=9, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=10, column=1).value, "Add into 15.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_cell_setup_fixed_od600_error_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.total_media_vol = 15000.
        # Add some information for cell setup
        # Do not include target od600
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        # Run save_rep_setup_instructions
        self.assertRaises(
            ValueError,
            p.save_rep_setup_instructions,
            file_name=os.path.join(self.temp_dir, 'plate_rep.xlsx'))

    def test_save_rep_setup_instructions_cell_setup_fixed_od600_error_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.total_media_vol = 15000.
        # Add some information for cell setup
        # Do not include predilution volume
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_initial_od600 = 1e-5
        # Run save_rep_setup_instructions
        self.assertRaises(
            ValueError,
            p.save_rep_setup_instructions,
            file_name=os.path.join(self.temp_dir, 'plate_rep.xlsx'))

    def test_save_rep_setup_instructions_inducer_rows_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=6, scale='log')
        p.apply_inducer(iptg, apply_to='rows')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I001")
        self.assertEqual(ws.cell(row=1, column=2).value, "I002")
        self.assertEqual(ws.cell(row=1, column=3).value, "I003")
        self.assertEqual(ws.cell(row=1, column=4).value, "I004")
        self.assertEqual(ws.cell(row=1, column=5).value, "I005")
        self.assertEqual(ws.cell(row=1, column=6).value, "I006")
        self.assertEqual(ws.cell(row=2, column=1).value, "(1, 1)")
        self.assertEqual(ws.cell(row=2, column=2).value, "(1, 2)")
        self.assertEqual(ws.cell(row=2, column=3).value, "(1, 3)")
        self.assertEqual(ws.cell(row=2, column=4).value, "(1, 4)")
        self.assertEqual(ws.cell(row=2, column=5).value, "(1, 5)")
        self.assertEqual(ws.cell(row=2, column=6).value, "(1, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "(2, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "(2, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "(2, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "(2, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "(2, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "(2, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "(3, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "(3, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "(3, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "(3, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "(3, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "(3, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "(4, 1)")
        self.assertEqual(ws.cell(row=5, column=2).value, "(4, 2)")
        self.assertEqual(ws.cell(row=5, column=3).value, "(4, 3)")
        self.assertEqual(ws.cell(row=5, column=4).value, "(4, 4)")
        self.assertEqual(ws.cell(row=5, column=5).value, "(4, 5)")
        self.assertEqual(ws.cell(row=5, column=6).value, "(4, 6)")

    def test_save_rep_setup_instructions_inducer_rows_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=6, scale='log')
        p.apply_inducer(iptg, apply_to='rows')
        # Create second inducer for plate rows
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=6, scale='log')
        p.apply_inducer(atc, apply_to='rows')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I001")
        self.assertEqual(ws.cell(row=1, column=2).value, "I002")
        self.assertEqual(ws.cell(row=1, column=3).value, "I003")
        self.assertEqual(ws.cell(row=1, column=4).value, "I004")
        self.assertEqual(ws.cell(row=1, column=5).value, "I005")
        self.assertEqual(ws.cell(row=1, column=6).value, "I006")
        self.assertEqual(ws.cell(row=2, column=1).value, "a001")
        self.assertEqual(ws.cell(row=2, column=2).value, "a002")
        self.assertEqual(ws.cell(row=2, column=3).value, "a003")
        self.assertEqual(ws.cell(row=2, column=4).value, "a004")
        self.assertEqual(ws.cell(row=2, column=5).value, "a005")
        self.assertEqual(ws.cell(row=2, column=6).value, "a006")
        self.assertEqual(ws.cell(row=3, column=1).value, "(1, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "(1, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "(1, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "(1, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "(1, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "(1, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "(2, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "(2, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "(2, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "(2, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "(2, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "(2, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "(3, 1)")
        self.assertEqual(ws.cell(row=5, column=2).value, "(3, 2)")
        self.assertEqual(ws.cell(row=5, column=3).value, "(3, 3)")
        self.assertEqual(ws.cell(row=5, column=4).value, "(3, 4)")
        self.assertEqual(ws.cell(row=5, column=5).value, "(3, 5)")
        self.assertEqual(ws.cell(row=5, column=6).value, "(3, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "(4, 1)")
        self.assertEqual(ws.cell(row=6, column=2).value, "(4, 2)")
        self.assertEqual(ws.cell(row=6, column=3).value, "(4, 3)")
        self.assertEqual(ws.cell(row=6, column=4).value, "(4, 4)")
        self.assertEqual(ws.cell(row=6, column=5).value, "(4, 5)")
        self.assertEqual(ws.cell(row=6, column=6).value, "(4, 6)")

    def test_save_rep_setup_instructions_inducer_rows_3(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=6, scale='log')
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='rows')
        # Create second inducer for plate rows
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=6, scale='log')
        p.apply_inducer(atc, apply_to='rows')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        # The order of IPTG values after shuffling has been checked manually
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I002")
        self.assertEqual(ws.cell(row=1, column=2).value, "I003")
        self.assertEqual(ws.cell(row=1, column=3).value, "I006")
        self.assertEqual(ws.cell(row=1, column=4).value, "I004")
        self.assertEqual(ws.cell(row=1, column=5).value, "I005")
        self.assertEqual(ws.cell(row=1, column=6).value, "I001")
        self.assertEqual(ws.cell(row=2, column=1).value, "a001")
        self.assertEqual(ws.cell(row=2, column=2).value, "a002")
        self.assertEqual(ws.cell(row=2, column=3).value, "a003")
        self.assertEqual(ws.cell(row=2, column=4).value, "a004")
        self.assertEqual(ws.cell(row=2, column=5).value, "a005")
        self.assertEqual(ws.cell(row=2, column=6).value, "a006")
        self.assertEqual(ws.cell(row=3, column=1).value, "(1, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "(1, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "(1, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "(1, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "(1, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "(1, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "(2, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "(2, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "(2, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "(2, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "(2, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "(2, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "(3, 1)")
        self.assertEqual(ws.cell(row=5, column=2).value, "(3, 2)")
        self.assertEqual(ws.cell(row=5, column=3).value, "(3, 3)")
        self.assertEqual(ws.cell(row=5, column=4).value, "(3, 4)")
        self.assertEqual(ws.cell(row=5, column=5).value, "(3, 5)")
        self.assertEqual(ws.cell(row=5, column=6).value, "(3, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "(4, 1)")
        self.assertEqual(ws.cell(row=6, column=2).value, "(4, 2)")
        self.assertEqual(ws.cell(row=6, column=3).value, "(4, 3)")
        self.assertEqual(ws.cell(row=6, column=4).value, "(4, 4)")
        self.assertEqual(ws.cell(row=6, column=5).value, "(4, 5)")
        self.assertEqual(ws.cell(row=6, column=6).value, "(4, 6)")

    def test_save_rep_setup_instructions_inducer_cols_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate columns
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=4, scale='log')
        p.apply_inducer(iptg, apply_to='cols')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I001")
        self.assertEqual(ws.cell(row=1, column=2).value, "(1, 1)")
        self.assertEqual(ws.cell(row=1, column=3).value, "(1, 2)")
        self.assertEqual(ws.cell(row=1, column=4).value, "(1, 3)")
        self.assertEqual(ws.cell(row=1, column=5).value, "(1, 4)")
        self.assertEqual(ws.cell(row=1, column=6).value, "(1, 5)")
        self.assertEqual(ws.cell(row=1, column=7).value, "(1, 6)")
        self.assertEqual(ws.cell(row=2, column=1).value, "I002")
        self.assertEqual(ws.cell(row=2, column=2).value, "(2, 1)")
        self.assertEqual(ws.cell(row=2, column=3).value, "(2, 2)")
        self.assertEqual(ws.cell(row=2, column=4).value, "(2, 3)")
        self.assertEqual(ws.cell(row=2, column=5).value, "(2, 4)")
        self.assertEqual(ws.cell(row=2, column=6).value, "(2, 5)")
        self.assertEqual(ws.cell(row=2, column=7).value, "(2, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "I003")
        self.assertEqual(ws.cell(row=3, column=2).value, "(3, 1)")
        self.assertEqual(ws.cell(row=3, column=3).value, "(3, 2)")
        self.assertEqual(ws.cell(row=3, column=4).value, "(3, 3)")
        self.assertEqual(ws.cell(row=3, column=5).value, "(3, 4)")
        self.assertEqual(ws.cell(row=3, column=6).value, "(3, 5)")
        self.assertEqual(ws.cell(row=3, column=7).value, "(3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "I004")
        self.assertEqual(ws.cell(row=4, column=2).value, "(4, 1)")
        self.assertEqual(ws.cell(row=4, column=3).value, "(4, 2)")
        self.assertEqual(ws.cell(row=4, column=4).value, "(4, 3)")
        self.assertEqual(ws.cell(row=4, column=5).value, "(4, 4)")
        self.assertEqual(ws.cell(row=4, column=6).value, "(4, 5)")
        self.assertEqual(ws.cell(row=4, column=7).value, "(4, 6)")

    def test_save_rep_setup_instructions_inducer_cols_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate columns
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=4, scale='log')
        p.apply_inducer(iptg, apply_to='cols')
        # Create second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=4, scale='log')
        p.apply_inducer(atc, apply_to='cols')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I001")
        self.assertEqual(ws.cell(row=1, column=2).value, "a001")
        self.assertEqual(ws.cell(row=1, column=3).value, "(1, 1)")
        self.assertEqual(ws.cell(row=1, column=4).value, "(1, 2)")
        self.assertEqual(ws.cell(row=1, column=5).value, "(1, 3)")
        self.assertEqual(ws.cell(row=1, column=6).value, "(1, 4)")
        self.assertEqual(ws.cell(row=1, column=7).value, "(1, 5)")
        self.assertEqual(ws.cell(row=1, column=8).value, "(1, 6)")
        self.assertEqual(ws.cell(row=2, column=1).value, "I002")
        self.assertEqual(ws.cell(row=2, column=2).value, "a002")
        self.assertEqual(ws.cell(row=2, column=3).value, "(2, 1)")
        self.assertEqual(ws.cell(row=2, column=4).value, "(2, 2)")
        self.assertEqual(ws.cell(row=2, column=5).value, "(2, 3)")
        self.assertEqual(ws.cell(row=2, column=6).value, "(2, 4)")
        self.assertEqual(ws.cell(row=2, column=7).value, "(2, 5)")
        self.assertEqual(ws.cell(row=2, column=8).value, "(2, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "I003")
        self.assertEqual(ws.cell(row=3, column=2).value, "a003")
        self.assertEqual(ws.cell(row=3, column=3).value, "(3, 1)")
        self.assertEqual(ws.cell(row=3, column=4).value, "(3, 2)")
        self.assertEqual(ws.cell(row=3, column=5).value, "(3, 3)")
        self.assertEqual(ws.cell(row=3, column=6).value, "(3, 4)")
        self.assertEqual(ws.cell(row=3, column=7).value, "(3, 5)")
        self.assertEqual(ws.cell(row=3, column=8).value, "(3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "I004")
        self.assertEqual(ws.cell(row=4, column=2).value, "a004")
        self.assertEqual(ws.cell(row=4, column=3).value, "(4, 1)")
        self.assertEqual(ws.cell(row=4, column=4).value, "(4, 2)")
        self.assertEqual(ws.cell(row=4, column=5).value, "(4, 3)")
        self.assertEqual(ws.cell(row=4, column=6).value, "(4, 4)")
        self.assertEqual(ws.cell(row=4, column=7).value, "(4, 5)")
        self.assertEqual(ws.cell(row=4, column=8).value, "(4, 6)")

    def test_save_rep_setup_instructions_inducer_cols_3(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate columns
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=4, scale='log')
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='cols')
        # Create second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=4, scale='log')
        p.apply_inducer(atc, apply_to='cols')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        # IPTG concentrations have been checked manually
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I004")
        self.assertEqual(ws.cell(row=1, column=2).value, "a001")
        self.assertEqual(ws.cell(row=1, column=3).value, "(1, 1)")
        self.assertEqual(ws.cell(row=1, column=4).value, "(1, 2)")
        self.assertEqual(ws.cell(row=1, column=5).value, "(1, 3)")
        self.assertEqual(ws.cell(row=1, column=6).value, "(1, 4)")
        self.assertEqual(ws.cell(row=1, column=7).value, "(1, 5)")
        self.assertEqual(ws.cell(row=1, column=8).value, "(1, 6)")
        self.assertEqual(ws.cell(row=2, column=1).value, "I002")
        self.assertEqual(ws.cell(row=2, column=2).value, "a002")
        self.assertEqual(ws.cell(row=2, column=3).value, "(2, 1)")
        self.assertEqual(ws.cell(row=2, column=4).value, "(2, 2)")
        self.assertEqual(ws.cell(row=2, column=5).value, "(2, 3)")
        self.assertEqual(ws.cell(row=2, column=6).value, "(2, 4)")
        self.assertEqual(ws.cell(row=2, column=7).value, "(2, 5)")
        self.assertEqual(ws.cell(row=2, column=8).value, "(2, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "I003")
        self.assertEqual(ws.cell(row=3, column=2).value, "a003")
        self.assertEqual(ws.cell(row=3, column=3).value, "(3, 1)")
        self.assertEqual(ws.cell(row=3, column=4).value, "(3, 2)")
        self.assertEqual(ws.cell(row=3, column=5).value, "(3, 3)")
        self.assertEqual(ws.cell(row=3, column=6).value, "(3, 4)")
        self.assertEqual(ws.cell(row=3, column=7).value, "(3, 5)")
        self.assertEqual(ws.cell(row=3, column=8).value, "(3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "I001")
        self.assertEqual(ws.cell(row=4, column=2).value, "a004")
        self.assertEqual(ws.cell(row=4, column=3).value, "(4, 1)")
        self.assertEqual(ws.cell(row=4, column=4).value, "(4, 2)")
        self.assertEqual(ws.cell(row=4, column=5).value, "(4, 3)")
        self.assertEqual(ws.cell(row=4, column=6).value, "(4, 4)")
        self.assertEqual(ws.cell(row=4, column=7).value, "(4, 5)")
        self.assertEqual(ws.cell(row=4, column=8).value, "(4, 6)")

    def test_save_rep_setup_instructions_inducer_wells_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=24, scale='log')
        p.apply_inducer(iptg, apply_to='wells')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "(1, 1)\nI001")
        self.assertEqual(ws.cell(row=1, column=2).value, "(1, 2)\nI002")
        self.assertEqual(ws.cell(row=1, column=3).value, "(1, 3)\nI003")
        self.assertEqual(ws.cell(row=1, column=4).value, "(1, 4)\nI004")
        self.assertEqual(ws.cell(row=1, column=5).value, "(1, 5)\nI005")
        self.assertEqual(ws.cell(row=1, column=6).value, "(1, 6)\nI006")
        self.assertEqual(ws.cell(row=2, column=1).value, "(2, 1)\nI007")
        self.assertEqual(ws.cell(row=2, column=2).value, "(2, 2)\nI008")
        self.assertEqual(ws.cell(row=2, column=3).value, "(2, 3)\nI009")
        self.assertEqual(ws.cell(row=2, column=4).value, "(2, 4)\nI010")
        self.assertEqual(ws.cell(row=2, column=5).value, "(2, 5)\nI011")
        self.assertEqual(ws.cell(row=2, column=6).value, "(2, 6)\nI012")
        self.assertEqual(ws.cell(row=3, column=1).value, "(3, 1)\nI013")
        self.assertEqual(ws.cell(row=3, column=2).value, "(3, 2)\nI014")
        self.assertEqual(ws.cell(row=3, column=3).value, "(3, 3)\nI015")
        self.assertEqual(ws.cell(row=3, column=4).value, "(3, 4)\nI016")
        self.assertEqual(ws.cell(row=3, column=5).value, "(3, 5)\nI017")
        self.assertEqual(ws.cell(row=3, column=6).value, "(3, 6)\nI018")
        self.assertEqual(ws.cell(row=4, column=1).value, "(4, 1)\nI019")
        self.assertEqual(ws.cell(row=4, column=2).value, "(4, 2)\nI020")
        self.assertEqual(ws.cell(row=4, column=3).value, "(4, 3)\nI021")
        self.assertEqual(ws.cell(row=4, column=4).value, "(4, 4)\nI022")
        self.assertEqual(ws.cell(row=4, column=5).value, "(4, 5)\nI023")
        self.assertEqual(ws.cell(row=4, column=6).value, "(4, 6)\nI024")

    def test_save_rep_setup_instructions_inducer_wells_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=24, scale='log')
        p.apply_inducer(iptg, apply_to='wells')
        # Create second inducer for plate
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=24, scale='log')
        p.apply_inducer(atc, apply_to='wells')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "(1, 1)\nI001\na001")
        self.assertEqual(ws.cell(row=1, column=2).value, "(1, 2)\nI002\na002")
        self.assertEqual(ws.cell(row=1, column=3).value, "(1, 3)\nI003\na003")
        self.assertEqual(ws.cell(row=1, column=4).value, "(1, 4)\nI004\na004")
        self.assertEqual(ws.cell(row=1, column=5).value, "(1, 5)\nI005\na005")
        self.assertEqual(ws.cell(row=1, column=6).value, "(1, 6)\nI006\na006")
        self.assertEqual(ws.cell(row=2, column=1).value, "(2, 1)\nI007\na007")
        self.assertEqual(ws.cell(row=2, column=2).value, "(2, 2)\nI008\na008")
        self.assertEqual(ws.cell(row=2, column=3).value, "(2, 3)\nI009\na009")
        self.assertEqual(ws.cell(row=2, column=4).value, "(2, 4)\nI010\na010")
        self.assertEqual(ws.cell(row=2, column=5).value, "(2, 5)\nI011\na011")
        self.assertEqual(ws.cell(row=2, column=6).value, "(2, 6)\nI012\na012")
        self.assertEqual(ws.cell(row=3, column=1).value, "(3, 1)\nI013\na013")
        self.assertEqual(ws.cell(row=3, column=2).value, "(3, 2)\nI014\na014")
        self.assertEqual(ws.cell(row=3, column=3).value, "(3, 3)\nI015\na015")
        self.assertEqual(ws.cell(row=3, column=4).value, "(3, 4)\nI016\na016")
        self.assertEqual(ws.cell(row=3, column=5).value, "(3, 5)\nI017\na017")
        self.assertEqual(ws.cell(row=3, column=6).value, "(3, 6)\nI018\na018")
        self.assertEqual(ws.cell(row=4, column=1).value, "(4, 1)\nI019\na019")
        self.assertEqual(ws.cell(row=4, column=2).value, "(4, 2)\nI020\na020")
        self.assertEqual(ws.cell(row=4, column=3).value, "(4, 3)\nI021\na021")
        self.assertEqual(ws.cell(row=4, column=4).value, "(4, 4)\nI022\na022")
        self.assertEqual(ws.cell(row=4, column=5).value, "(4, 5)\nI023\na023")
        self.assertEqual(ws.cell(row=4, column=6).value, "(4, 6)\nI024\na024")

    def test_save_rep_setup_instructions_inducer_wells_3(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.samples_to_measure = 12
        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=12, scale='log')
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='wells')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "(1, 1)\nI001")
        self.assertEqual(ws.cell(row=1, column=2).value, "(1, 2)\nI002")
        self.assertEqual(ws.cell(row=1, column=3).value, "(1, 3)\nI003")
        self.assertEqual(ws.cell(row=1, column=4).value, "(1, 4)\nI004")
        self.assertEqual(ws.cell(row=1, column=5).value, "(1, 5)\nI005")
        self.assertEqual(ws.cell(row=1, column=6).value, "(1, 6)\nI006")
        self.assertEqual(ws.cell(row=2, column=1).value, "(2, 1)\nI007")
        self.assertEqual(ws.cell(row=2, column=2).value, "(2, 2)\nI008")
        self.assertEqual(ws.cell(row=2, column=3).value, "(2, 3)\nI009")
        self.assertEqual(ws.cell(row=2, column=4).value, "(2, 4)\nI010")
        self.assertEqual(ws.cell(row=2, column=5).value, "(2, 5)\nI011")
        self.assertEqual(ws.cell(row=2, column=6).value, "(2, 6)\nI012")
        self.assertEqual(ws.cell(row=3, column=1).value, "(3, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "(3, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "(3, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "(3, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "(3, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "(3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "(4, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "(4, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "(4, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "(4, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "(4, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "(4, 6)")

    def test_save_rep_setup_instructions_inducer_wells_4(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=24, scale='log')
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='wells')
        # Create second inducer for plate
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=24, scale='log')
        p.apply_inducer(atc, apply_to='wells')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "(1, 1)\nI024\na001")
        self.assertEqual(ws.cell(row=1, column=2).value, "(1, 2)\nI003\na002")
        self.assertEqual(ws.cell(row=1, column=3).value, "(1, 3)\nI008\na003")
        self.assertEqual(ws.cell(row=1, column=4).value, "(1, 4)\nI022\na004")
        self.assertEqual(ws.cell(row=1, column=5).value, "(1, 5)\nI011\na005")
        self.assertEqual(ws.cell(row=1, column=6).value, "(1, 6)\nI013\na006")
        self.assertEqual(ws.cell(row=2, column=1).value, "(2, 1)\nI019\na007")
        self.assertEqual(ws.cell(row=2, column=2).value, "(2, 2)\nI016\na008")
        self.assertEqual(ws.cell(row=2, column=3).value, "(2, 3)\nI007\na009")
        self.assertEqual(ws.cell(row=2, column=4).value, "(2, 4)\nI005\na010")
        self.assertEqual(ws.cell(row=2, column=5).value, "(2, 5)\nI015\na011")
        self.assertEqual(ws.cell(row=2, column=6).value, "(2, 6)\nI023\na012")
        self.assertEqual(ws.cell(row=3, column=1).value, "(3, 1)\nI021\na013")
        self.assertEqual(ws.cell(row=3, column=2).value, "(3, 2)\nI018\na014")
        self.assertEqual(ws.cell(row=3, column=3).value, "(3, 3)\nI001\na015")
        self.assertEqual(ws.cell(row=3, column=4).value, "(3, 4)\nI002\na016")
        self.assertEqual(ws.cell(row=3, column=5).value, "(3, 5)\nI014\na017")
        self.assertEqual(ws.cell(row=3, column=6).value, "(3, 6)\nI012\na018")
        self.assertEqual(ws.cell(row=4, column=1).value, "(4, 1)\nI009\na019")
        self.assertEqual(ws.cell(row=4, column=2).value, "(4, 2)\nI010\na020")
        self.assertEqual(ws.cell(row=4, column=3).value, "(4, 3)\nI006\na021")
        self.assertEqual(ws.cell(row=4, column=4).value, "(4, 4)\nI017\na022")
        self.assertEqual(ws.cell(row=4, column=5).value, "(4, 5)\nI020\na023")
        self.assertEqual(ws.cell(row=4, column=6).value, "(4, 6)\nI004\na024")

    def test_save_rep_setup_instructions_inducer_media_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.shot_vol = 5.
        # Set single concentration
        iptg.concentrations = [12]
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='media')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "(1, 1)")
        self.assertEqual(ws.cell(row=1, column=2).value, "(1, 2)")
        self.assertEqual(ws.cell(row=1, column=3).value, "(1, 3)")
        self.assertEqual(ws.cell(row=1, column=4).value, "(1, 4)")
        self.assertEqual(ws.cell(row=1, column=5).value, "(1, 5)")
        self.assertEqual(ws.cell(row=1, column=6).value, "(1, 6)")
        self.assertEqual(ws.cell(row=2, column=1).value, "(2, 1)")
        self.assertEqual(ws.cell(row=2, column=2).value, "(2, 2)")
        self.assertEqual(ws.cell(row=2, column=3).value, "(2, 3)")
        self.assertEqual(ws.cell(row=2, column=4).value, "(2, 4)")
        self.assertEqual(ws.cell(row=2, column=5).value, "(2, 5)")
        self.assertEqual(ws.cell(row=2, column=6).value, "(2, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "(3, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "(3, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "(3, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "(3, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "(3, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "(3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "(4, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "(4, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "(4, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "(4, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "(4, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "(4, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value,
                         u"Add 5.00µL of IPTG to media")

    def test_save_rep_setup_instructions_inducer_media_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.shot_vol = 5.
        # Set single concentration
        iptg.concentrations = [12]
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='media')
        # Create inducer for plate
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.shot_vol = 10.
        # Set single concentration
        atc.concentrations = [2]
        # Apply inducer to plate
        p.apply_inducer(atc, apply_to='media')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "(1, 1)")
        self.assertEqual(ws.cell(row=1, column=2).value, "(1, 2)")
        self.assertEqual(ws.cell(row=1, column=3).value, "(1, 3)")
        self.assertEqual(ws.cell(row=1, column=4).value, "(1, 4)")
        self.assertEqual(ws.cell(row=1, column=5).value, "(1, 5)")
        self.assertEqual(ws.cell(row=1, column=6).value, "(1, 6)")
        self.assertEqual(ws.cell(row=2, column=1).value, "(2, 1)")
        self.assertEqual(ws.cell(row=2, column=2).value, "(2, 2)")
        self.assertEqual(ws.cell(row=2, column=3).value, "(2, 3)")
        self.assertEqual(ws.cell(row=2, column=4).value, "(2, 4)")
        self.assertEqual(ws.cell(row=2, column=5).value, "(2, 5)")
        self.assertEqual(ws.cell(row=2, column=6).value, "(2, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "(3, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "(3, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "(3, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "(3, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "(3, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "(3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "(4, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "(4, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "(4, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "(4, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "(4, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "(4, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value,
                         u"Add 5.00µL of IPTG to media")
        self.assertEqual(ws.cell(row=7, column=1).value,
                         u"Add 10.00µL of aTc to media")

    def test_save_rep_setup_instructions_inducer_mixed(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=6, scale='log')
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='rows')

        # Create second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        # Set concentrations from gradient
        atc.set_gradient(min=0.5, max=50, n=4, scale='log')
        # Apply inducer to plate
        p.apply_inducer(atc, apply_to='cols')

        # Create inducer for plate wells
        xyl = platedesign.inducer.ChemicalInducer(
            name='Xylose',
            units=u'%')
        # Set concentrations from gradient
        xyl.set_gradient(min=1e-6, max=1e-3, n=24, scale='log')
        # Apply inducer to plate
        p.apply_inducer(xyl, apply_to='wells')

        # Create inducer for plate
        sugar = platedesign.inducer.ChemicalInducer(
            name='Sugar',
            units=u'ng/µL')
        sugar.shot_vol = 10.
        # Set single concentration
        sugar.concentrations = [2]
        # Apply inducer to plate
        p.apply_inducer(sugar, apply_to='media')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate P1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, None)
        self.assertEqual(ws.cell(row=1, column=2).value, "I001")
        self.assertEqual(ws.cell(row=1, column=3).value, "I002")
        self.assertEqual(ws.cell(row=1, column=4).value, "I003")
        self.assertEqual(ws.cell(row=1, column=5).value, "I004")
        self.assertEqual(ws.cell(row=1, column=6).value, "I005")
        self.assertEqual(ws.cell(row=1, column=7).value, "I006")
        self.assertEqual(ws.cell(row=2, column=1).value, "a001")
        self.assertEqual(ws.cell(row=2, column=2).value, "(1, 1)\nX001")
        self.assertEqual(ws.cell(row=2, column=3).value, "(1, 2)\nX002")
        self.assertEqual(ws.cell(row=2, column=4).value, "(1, 3)\nX003")
        self.assertEqual(ws.cell(row=2, column=5).value, "(1, 4)\nX004")
        self.assertEqual(ws.cell(row=2, column=6).value, "(1, 5)\nX005")
        self.assertEqual(ws.cell(row=2, column=7).value, "(1, 6)\nX006")
        self.assertEqual(ws.cell(row=3, column=1).value, "a002")
        self.assertEqual(ws.cell(row=3, column=2).value, "(2, 1)\nX007")
        self.assertEqual(ws.cell(row=3, column=3).value, "(2, 2)\nX008")
        self.assertEqual(ws.cell(row=3, column=4).value, "(2, 3)\nX009")
        self.assertEqual(ws.cell(row=3, column=5).value, "(2, 4)\nX010")
        self.assertEqual(ws.cell(row=3, column=6).value, "(2, 5)\nX011")
        self.assertEqual(ws.cell(row=3, column=7).value, "(2, 6)\nX012")
        self.assertEqual(ws.cell(row=4, column=1).value, "a003")
        self.assertEqual(ws.cell(row=4, column=2).value, "(3, 1)\nX013")
        self.assertEqual(ws.cell(row=4, column=3).value, "(3, 2)\nX014")
        self.assertEqual(ws.cell(row=4, column=4).value, "(3, 3)\nX015")
        self.assertEqual(ws.cell(row=4, column=5).value, "(3, 4)\nX016")
        self.assertEqual(ws.cell(row=4, column=6).value, "(3, 5)\nX017")
        self.assertEqual(ws.cell(row=4, column=7).value, "(3, 6)\nX018")
        self.assertEqual(ws.cell(row=5, column=1).value, "a004")
        self.assertEqual(ws.cell(row=5, column=2).value, "(4, 1)\nX019")
        self.assertEqual(ws.cell(row=5, column=3).value, "(4, 2)\nX020")
        self.assertEqual(ws.cell(row=5, column=4).value, "(4, 3)\nX021")
        self.assertEqual(ws.cell(row=5, column=5).value, "(4, 4)\nX022")
        self.assertEqual(ws.cell(row=5, column=6).value, "(4, 5)\nX023")
        self.assertEqual(ws.cell(row=5, column=7).value, "(4, 6)\nX024")
        self.assertEqual(ws.cell(row=7, column=1).value,
                         u"Add 10.00µL of Sugar to media")

    def test_save_rep_setup_instructions_cells_and_inducer(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.total_media_vol = 15000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_initial_od600 = 1e-5

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=6, scale='log')
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='rows')

        # Create second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        # Set concentrations from gradient
        atc.set_gradient(min=0.5, max=50, n=4, scale='log')
        # Apply inducer to plate
        p.apply_inducer(atc, apply_to='cols')

        # Create inducer for plate wells
        xyl = platedesign.inducer.ChemicalInducer(
            name='Xylose',
            units=u'%')
        # Set concentrations from gradient
        xyl.set_gradient(min=1e-6, max=1e-3, n=24, scale='log')
        # Apply inducer to plate
        p.apply_inducer(xyl, apply_to='wells')

        # Create inducer for plate
        sugar = platedesign.inducer.ChemicalInducer(
            name='Sugar',
            units=u'ng/µL')
        sugar.shot_vol = 10.
        # Set single concentration
        sugar.concentrations = [2]
        # Apply inducer to plate
        p.apply_inducer(sugar, apply_to='media')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Check that sheet exists in spreadsheet
        self.assertTrue("Inducers for Plate P1" in wb.sheetnames)
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, None)
        self.assertEqual(ws.cell(row=1, column=2).value, "I001")
        self.assertEqual(ws.cell(row=1, column=3).value, "I002")
        self.assertEqual(ws.cell(row=1, column=4).value, "I003")
        self.assertEqual(ws.cell(row=1, column=5).value, "I004")
        self.assertEqual(ws.cell(row=1, column=6).value, "I005")
        self.assertEqual(ws.cell(row=1, column=7).value, "I006")
        self.assertEqual(ws.cell(row=2, column=1).value, "a001")
        self.assertEqual(ws.cell(row=2, column=2).value, "(1, 1)\nX001")
        self.assertEqual(ws.cell(row=2, column=3).value, "(1, 2)\nX002")
        self.assertEqual(ws.cell(row=2, column=4).value, "(1, 3)\nX003")
        self.assertEqual(ws.cell(row=2, column=5).value, "(1, 4)\nX004")
        self.assertEqual(ws.cell(row=2, column=6).value, "(1, 5)\nX005")
        self.assertEqual(ws.cell(row=2, column=7).value, "(1, 6)\nX006")
        self.assertEqual(ws.cell(row=3, column=1).value, "a002")
        self.assertEqual(ws.cell(row=3, column=2).value, "(2, 1)\nX007")
        self.assertEqual(ws.cell(row=3, column=3).value, "(2, 2)\nX008")
        self.assertEqual(ws.cell(row=3, column=4).value, "(2, 3)\nX009")
        self.assertEqual(ws.cell(row=3, column=5).value, "(2, 4)\nX010")
        self.assertEqual(ws.cell(row=3, column=6).value, "(2, 5)\nX011")
        self.assertEqual(ws.cell(row=3, column=7).value, "(2, 6)\nX012")
        self.assertEqual(ws.cell(row=4, column=1).value, "a003")
        self.assertEqual(ws.cell(row=4, column=2).value, "(3, 1)\nX013")
        self.assertEqual(ws.cell(row=4, column=3).value, "(3, 2)\nX014")
        self.assertEqual(ws.cell(row=4, column=4).value, "(3, 3)\nX015")
        self.assertEqual(ws.cell(row=4, column=5).value, "(3, 4)\nX016")
        self.assertEqual(ws.cell(row=4, column=6).value, "(3, 5)\nX017")
        self.assertEqual(ws.cell(row=4, column=7).value, "(3, 6)\nX018")
        self.assertEqual(ws.cell(row=5, column=1).value, "a004")
        self.assertEqual(ws.cell(row=5, column=2).value, "(4, 1)\nX019")
        self.assertEqual(ws.cell(row=5, column=3).value, "(4, 2)\nX020")
        self.assertEqual(ws.cell(row=5, column=4).value, "(4, 3)\nX021")
        self.assertEqual(ws.cell(row=5, column=5).value, "(4, 4)\nX022")
        self.assertEqual(ws.cell(row=5, column=6).value, "(4, 5)\nX023")
        self.assertEqual(ws.cell(row=5, column=7).value, "(4, 6)\nX024")
        self.assertEqual(ws.cell(row=7, column=1).value,
                         u"Add 10.00µL of Sugar to media")

        # Check that sheet exists in spreadsheet
        self.assertTrue("Cells for Plate P1" in wb.sheetnames)
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertTrue('A2:C2' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=2, column=1).value, "Predilution")
        self.assertEqual(ws.cell(row=3, column=1).value, "Predilution factor")
        self.assertEqual(ws.cell(row=3, column=2).value, 100)
        self.assertEqual(ws.cell(row=3, column=3).value, "x")
        self.assertEqual(ws.cell(row=4, column=1).value, "Media volume")
        self.assertEqual(ws.cell(row=4, column=2).value, 990)
        self.assertEqual(ws.cell(row=4, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=5, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=5, column=2).value, 10)
        self.assertEqual(ws.cell(row=5, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=6, column=1).value, "Predilution OD600")
        self.assertEqual(ws.cell(row=6, column=2).value, None)
        self.assertEqual(ws.cell(row=6, column=3).value, None)
        self.assertTrue('A7:C7' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=7, column=1).value, "Inoculation")
        self.assertEqual(ws.cell(row=8, column=1).value, "Target OD600")
        self.assertEqual(ws.cell(row=8, column=2).value, 1e-5)
        self.assertEqual(ws.cell(row=8, column=3).value, None)
        self.assertEqual(ws.cell(row=9, column=1).value, "Predilution volume")
        self.assertEqual(ws.cell(row=9, column=2).value, "=0.15/B6")
        self.assertEqual(ws.cell(row=9, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=10, column=1).value, "Add into 15.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_cells_and_inducer_workbook(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.total_media_vol = 15000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_initial_od600 = 1e-5

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations from gradient
        iptg.set_gradient(min=1e-6, max=1e-3, n=6, scale='log')
        # Apply inducer to plate
        p.apply_inducer(iptg, apply_to='rows')

        # Create second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        # Set concentrations from gradient
        atc.set_gradient(min=0.5, max=50, n=4, scale='log')
        # Apply inducer to plate
        p.apply_inducer(atc, apply_to='cols')

        # Create inducer for plate wells
        xyl = platedesign.inducer.ChemicalInducer(
            name='Xylose',
            units=u'%')
        # Set concentrations from gradient
        xyl.set_gradient(min=1e-6, max=1e-3, n=24, scale='log')
        # Apply inducer to plate
        p.apply_inducer(xyl, apply_to='wells')

        # Create inducer for plate
        sugar = platedesign.inducer.ChemicalInducer(
            name='Sugar',
            units=u'ng/µL')
        sugar.shot_vol = 10.
        # Set single concentration
        sugar.concentrations = [2]
        # Apply inducer to plate
        p.apply_inducer(sugar, apply_to='media')

        # Create new spreadsheet
        wb = openpyxl.Workbook()
        # Save instructions on existing spreadsheet
        p.save_rep_setup_instructions(workbook=wb)

        # Check that sheet exists in spreadsheet
        self.assertTrue("Inducers for Plate P1" in wb.sheetnames)
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "")
        self.assertEqual(ws.cell(row=1, column=2).value, "I001")
        self.assertEqual(ws.cell(row=1, column=3).value, "I002")
        self.assertEqual(ws.cell(row=1, column=4).value, "I003")
        self.assertEqual(ws.cell(row=1, column=5).value, "I004")
        self.assertEqual(ws.cell(row=1, column=6).value, "I005")
        self.assertEqual(ws.cell(row=1, column=7).value, "I006")
        self.assertEqual(ws.cell(row=2, column=1).value, "a001")
        self.assertEqual(ws.cell(row=2, column=2).value, "(1, 1)\nX001")
        self.assertEqual(ws.cell(row=2, column=3).value, "(1, 2)\nX002")
        self.assertEqual(ws.cell(row=2, column=4).value, "(1, 3)\nX003")
        self.assertEqual(ws.cell(row=2, column=5).value, "(1, 4)\nX004")
        self.assertEqual(ws.cell(row=2, column=6).value, "(1, 5)\nX005")
        self.assertEqual(ws.cell(row=2, column=7).value, "(1, 6)\nX006")
        self.assertEqual(ws.cell(row=3, column=1).value, "a002")
        self.assertEqual(ws.cell(row=3, column=2).value, "(2, 1)\nX007")
        self.assertEqual(ws.cell(row=3, column=3).value, "(2, 2)\nX008")
        self.assertEqual(ws.cell(row=3, column=4).value, "(2, 3)\nX009")
        self.assertEqual(ws.cell(row=3, column=5).value, "(2, 4)\nX010")
        self.assertEqual(ws.cell(row=3, column=6).value, "(2, 5)\nX011")
        self.assertEqual(ws.cell(row=3, column=7).value, "(2, 6)\nX012")
        self.assertEqual(ws.cell(row=4, column=1).value, "a003")
        self.assertEqual(ws.cell(row=4, column=2).value, "(3, 1)\nX013")
        self.assertEqual(ws.cell(row=4, column=3).value, "(3, 2)\nX014")
        self.assertEqual(ws.cell(row=4, column=4).value, "(3, 3)\nX015")
        self.assertEqual(ws.cell(row=4, column=5).value, "(3, 4)\nX016")
        self.assertEqual(ws.cell(row=4, column=6).value, "(3, 5)\nX017")
        self.assertEqual(ws.cell(row=4, column=7).value, "(3, 6)\nX018")
        self.assertEqual(ws.cell(row=5, column=1).value, "a004")
        self.assertEqual(ws.cell(row=5, column=2).value, "(4, 1)\nX019")
        self.assertEqual(ws.cell(row=5, column=3).value, "(4, 2)\nX020")
        self.assertEqual(ws.cell(row=5, column=4).value, "(4, 3)\nX021")
        self.assertEqual(ws.cell(row=5, column=5).value, "(4, 4)\nX022")
        self.assertEqual(ws.cell(row=5, column=6).value, "(4, 5)\nX023")
        self.assertEqual(ws.cell(row=5, column=7).value, "(4, 6)\nX024")
        self.assertEqual(ws.cell(row=7, column=1).value,
                         u"Add 10.00µL of Sugar to media")

        # Check that sheet exists in spreadsheet
        self.assertTrue("Cells for Plate P1" in wb.sheetnames)
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate P1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertTrue('A2:C2' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=2, column=1).value, "Predilution")
        self.assertEqual(ws.cell(row=3, column=1).value, "Predilution factor")
        self.assertEqual(ws.cell(row=3, column=2).value, 100)
        self.assertEqual(ws.cell(row=3, column=3).value, "x")
        self.assertEqual(ws.cell(row=4, column=1).value, "Media volume")
        self.assertEqual(ws.cell(row=4, column=2).value, 990)
        self.assertEqual(ws.cell(row=4, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=5, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=5, column=2).value, 10)
        self.assertEqual(ws.cell(row=5, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=6, column=1).value, "Predilution OD600")
        self.assertEqual(ws.cell(row=6, column=2).value, None)
        self.assertEqual(ws.cell(row=6, column=3).value, None)
        self.assertTrue('A7:C7' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=7, column=1).value, "Inoculation")
        self.assertEqual(ws.cell(row=8, column=1).value, "Target OD600")
        self.assertEqual(ws.cell(row=8, column=2).value, 1e-5)
        self.assertEqual(ws.cell(row=8, column=3).value, None)
        self.assertEqual(ws.cell(row=9, column=1).value, "Predilution volume")
        self.assertEqual(ws.cell(row=9, column=2).value, "=0.15/B6")
        self.assertEqual(ws.cell(row=9, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=10, column=1).value, "Add into 15.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_argument_error(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Run save_rep_setup_instructions with no arguments
        self.assertRaises(ValueError, p.save_rep_setup_instructions)

    def test_save_rep_setup_files(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Check that save_rep_setup_files runs successfully
        p.save_rep_setup_files()
        # save_exp_setup_files does not do anything in Plate. There is no need
        # to check for results.

    def test_add_inducer_setup_instructions_sheet_error(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create workbook
        wb = openpyxl.Workbook()
        # Add sheet
        wb.create_sheet("Test sheet")
        # Attempt to add cell setup instructions
        self.assertRaises(ValueError,
                          p.add_inducer_setup_instructions,
                          wb,
                          "Test sheet")

    def test_add_cell_setup_instructions_sheet_error(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        # Create workbook
        wb = openpyxl.Workbook()
        # Add sheet
        wb.create_sheet("Test sheet")
        # Attempt to add cell setup instructions
        self.assertRaises(ValueError,
                          p.add_cell_setup_instructions,
                          wb,
                          "Test sheet")

    def test_close_plates_no_inducer_no_cell_inoculation(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_metadata(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'
        # Add some metadata
        p.metadata['Meta 1'] = 'Value 1'
        p.metadata['Meta 2'] = 'Value 2'
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 3)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        self.assertTrue('Meta 1' in cp.plate_info)
        self.assertEqual(cp.plate_info['Meta 1'], 'Value 1')
        self.assertTrue('Meta 2' in cp.plate_info)
        self.assertEqual(cp.plate_info['Meta 2'], 'Value 2')
        # Check well info
        well_info = pandas.DataFrame()
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_cell_inoculation_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'
        # Add cell inoculation info
        p.cell_setup_method = 'fixed_volume'
        p.cell_shot_vol = 5
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 3)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        self.assertTrue('Preculture Dilution' in cp.plate_info)
        self.assertEqual(cp.plate_info['Preculture Dilution'], 1)
        self.assertTrue('Cell Inoculated Vol.' in cp.plate_info)
        self.assertEqual(cp.plate_info['Cell Inoculated Vol.'], 5)
        # Check well info
        well_info = pandas.DataFrame()
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_cell_inoculation_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'
        # Add cell inoculation info
        p.cell_setup_method = 'fixed_volume'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_shot_vol = 5
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 3)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        self.assertTrue('Preculture Dilution' in cp.plate_info)
        self.assertEqual(cp.plate_info['Preculture Dilution'], 100)
        self.assertTrue('Cell Inoculated Vol.' in cp.plate_info)
        self.assertEqual(cp.plate_info['Cell Inoculated Vol.'], 5)
        # Check well info
        well_info = pandas.DataFrame()
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_cell_inoculation_3(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'
        # Add cell inoculation info
        p.cell_setup_method = 'fixed_od600'
        p.cell_initial_od600 = 1e-5
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 3)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        self.assertTrue('Preculture Dilution' in cp.plate_info)
        self.assertEqual(cp.plate_info['Preculture Dilution'], 1)
        self.assertTrue('Initial OD600' in cp.plate_info)
        self.assertEqual(cp.plate_info['Initial OD600'], 1e-5)
        # Check well info
        well_info = pandas.DataFrame()
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_cell_inoculation_4(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'
        # Add cell inoculation info
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_initial_od600 = 1e-5
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 3)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        self.assertTrue('Preculture Dilution' in cp.plate_info)
        self.assertEqual(cp.plate_info['Preculture Dilution'], 100)
        self.assertTrue('Initial OD600' in cp.plate_info)
        self.assertEqual(cp.plate_info['Initial OD600'], 1e-5)
        # Check well info
        well_info = pandas.DataFrame()
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_row_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3, 4, 5, 6, 7, 8]
        p.apply_inducer(iptg, apply_to='rows')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [3., 4., 5., 6., 7., 8.,
                                                 3., 4., 5., 6., 7., 8.,
                                                 3., 4., 5., 6., 7., 8.,
                                                 3., 4., 5., 6., 7., 8.]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_row_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3, 4, 5, 6, 7, 8]
        p.apply_inducer(iptg, apply_to='rows')

        # Second inducer for plate rows
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6]
        p.apply_inducer(atc, apply_to='rows')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [3., 4., 5., 6., 7., 8.,
                                                 3., 4., 5., 6., 7., 8.,
                                                 3., 4., 5., 6., 7., 8.,
                                                 3., 4., 5., 6., 7., 8.]
        well_info[u'aTc Concentration (ng/µL)'] = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
                                                   0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
                                                   0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
                                                   0.1, 0.2, 0.3, 0.4, 0.5, 0.6,]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_row_3(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3, 4, 5, 6, 7, 8]
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='rows')

        # Second inducer for plate rows
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6]
        p.apply_inducer(atc, apply_to='rows')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [4., 5., 8., 6., 7., 3.,
                                                 4., 5., 8., 6., 7., 3.,
                                                 4., 5., 8., 6., 7., 3.,
                                                 4., 5., 8., 6., 7., 3.,]
        well_info[u'aTc Concentration (ng/µL)'] = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
                                                   0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
                                                   0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
                                                   0.1, 0.2, 0.3, 0.4, 0.5, 0.6,]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_col_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate columns
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3, 4, 5, 6]
        p.apply_inducer(iptg, apply_to='cols')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [3., 3., 3., 3., 3., 3.,
                                                 4., 4., 4., 4., 4., 4.,
                                                 5., 5., 5., 5., 5., 5.,
                                                 6., 6., 6., 6., 6., 6.,]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_col_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate columns
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3, 4, 5, 6]
        p.apply_inducer(iptg, apply_to='cols')

        # Second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = [0.1, 0.2, 0.3, 0.4]
        p.apply_inducer(atc, apply_to='cols')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [3., 3., 3., 3., 3., 3.,
                                                 4., 4., 4., 4., 4., 4.,
                                                 5., 5., 5., 5., 5., 5.,
                                                 6., 6., 6., 6., 6., 6.,]
        well_info[u'aTc Concentration (ng/µL)'] = [0.1, 0.1, 0.1, 0.1, 0.1, 0.1,
                                                   0.2, 0.2, 0.2, 0.2, 0.2, 0.2,
                                                   0.3, 0.3, 0.3, 0.3, 0.3, 0.3,
                                                   0.4, 0.4, 0.4, 0.4, 0.4, 0.4,]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_col_3(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate columns
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3, 4, 5, 6]
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='cols')

        # Second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = [0.1, 0.2, 0.3, 0.4]
        p.apply_inducer(atc, apply_to='cols')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [6., 6., 6., 6., 6., 6.,
                                                 4., 4., 4., 4., 4., 4.,
                                                 5., 5., 5., 5., 5., 5.,
                                                 3., 3., 3., 3., 3., 3.,]
        well_info[u'aTc Concentration (ng/µL)'] = [0.1, 0.1, 0.1, 0.1, 0.1, 0.1,
                                                   0.2, 0.2, 0.2, 0.2, 0.2, 0.2,
                                                   0.3, 0.3, 0.3, 0.3, 0.3, 0.3,
                                                   0.4, 0.4, 0.4, 0.4, 0.4, 0.4,]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_wells_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(24) + 1
        p.apply_inducer(iptg, apply_to='wells')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [1., 2., 3., 4., 5., 6.,
                                                 7., 8., 9., 10., 11., 12.,
                                                 13., 14., 15., 16., 17., 18.,
                                                 19., 20., 21., 22., 23., 24.]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_wells_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducers for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(24) + 1
        p.apply_inducer(iptg, apply_to='wells')

        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = (numpy.arange(24) + 1)/10.
        p.apply_inducer(atc, apply_to='wells')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [1., 2., 3., 4., 5., 6.,
                                                 7., 8., 9., 10., 11., 12.,
                                                 13., 14., 15., 16., 17., 18.,
                                                 19., 20., 21., 22., 23., 24.]
        well_info[u'aTc Concentration (ng/µL)'] = [0.1, 0.2, 0.3, 0.4, 0.5 ,0.6,
                                                   0.7, 0.8, 0.9, 1.0, 1.1, 1.2,
                                                   1.3, 1.4, 1.5, 1.6, 1.7, 1.8,
                                                   1.9, 2.0, 2.1, 2.2, 2.3, 2.4]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_wells_3(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducers for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(24) + 1
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='wells')

        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = (numpy.arange(24) + 1)/10.
        p.apply_inducer(atc, apply_to='wells')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [24., 3., 8., 22., 11., 13.,
                                                 19., 16., 7., 5., 15., 23.,
                                                 21., 18., 1., 2., 14., 12.,
                                                 9., 10., 6., 17., 20., 4.,]
        well_info[u'aTc Concentration (ng/µL)'] = [0.1, 0.2, 0.3, 0.4, 0.5 ,0.6,
                                                   0.7, 0.8, 0.9, 1.0, 1.1, 1.2,
                                                   1.3, 1.4, 1.5, 1.6, 1.7, 1.8,
                                                   1.9, 2.0, 2.1, 2.2, 2.3, 2.4]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_wells_4(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.samples_to_measure = 12
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(12) + 1
        p.apply_inducer(iptg, apply_to='wells')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [1., 2., 3., 4., 5., 6.,
                                                 7., 8., 9., 10., 11., 12.,
                                                 None, None, None, None, None, None,
                                                 None, None, None, None, None, None]
        well_info['Measure'] = [True]*12 + [False]*12
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_media_1(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3]
        p.apply_inducer(iptg, apply_to='media')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [3., 3., 3., 3., 3., 3.,
                                                 3., 3., 3., 3., 3., 3.,
                                                 3., 3., 3., 3., 3., 3.,
                                                 3., 3., 3., 3., 3., 3.,]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_media_2(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3]
        p.apply_inducer(iptg, apply_to='media')

        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = [5]
        p.apply_inducer(atc, apply_to='media')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [3., 3., 3., 3., 3., 3.,
                                                 3., 3., 3., 3., 3., 3.,
                                                 3., 3., 3., 3., 3., 3.,
                                                 3., 3., 3., 3., 3., 3.,]
        well_info[u'aTc Concentration (ng/µL)'] = [5., 5., 5., 5., 5., 5.,
                                                   5., 5., 5., 5., 5., 5.,
                                                   5., 5., 5., 5., 5., 5.,
                                                   5., 5., 5., 5., 5., 5.,]
        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_media_3(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.samples_to_measure = 12
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3]
        p.apply_inducer(iptg, apply_to='media')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)
        # Check plate info
        self.assertEqual(len(cp.plate_info), 1)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [3., 3., 3., 3., 3., 3.,
                                                 3., 3., 3., 3., 3., 3.,
                                                 None, None, None, None, None, None,
                                                 None, None, None, None, None, None]
        well_info['Measure'] = [True]*12 + [False]*12
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_combined(self):
        # Create plate
        p = platedesign.plate.Plate(name='P1')
        p.cell_strain_name = 'Test strain 1'
        # Add some metadata
        p.metadata['Meta 1'] = 'Value 1'
        p.metadata['Meta 2'] = 'Value 2'
        # Add cell inoculation info
        p.cell_setup_method = 'fixed_volume'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_shot_vol = 5

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3, 4, 5, 6, 7, 8]
        p.apply_inducer(iptg, apply_to='rows')

        # Second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = [0.1, 0.2, 0.3, 0.4]
        p.apply_inducer(atc, apply_to='cols')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 1)
        # Get the only closed place
        cp = cps[0]
        # Check basic properties
        self.assertEqual(cp.name, 'P1')
        self.assertEqual(cp.n_rows, 4)
        self.assertEqual(cp.n_cols, 6)

        # Check plate info
        self.assertEqual(len(cp.plate_info), 5)
        self.assertTrue('Strain' in cp.plate_info)
        self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
        self.assertTrue('Meta 1' in cp.plate_info)
        self.assertEqual(cp.plate_info['Meta 1'], 'Value 1')
        self.assertTrue('Meta 2' in cp.plate_info)
        self.assertEqual(cp.plate_info['Meta 2'], 'Value 2')
        self.assertTrue('Preculture Dilution' in cp.plate_info)
        self.assertEqual(cp.plate_info['Preculture Dilution'], 100)
        self.assertTrue('Cell Inoculated Vol.' in cp.plate_info)
        self.assertEqual(cp.plate_info['Cell Inoculated Vol.'], 5)

        # Check well info
        well_info = pandas.DataFrame()
        well_info[u'IPTG Concentration (µM)'] = [3., 4., 5., 6., 7., 8.,
                                                 3., 4., 5., 6., 7., 8.,
                                                 3., 4., 5., 6., 7., 8.,
                                                 3., 4., 5., 6., 7., 8.]
        well_info[u'aTc Concentration (ng/µL)'] = [0.1, 0.1, 0.1, 0.1, 0.1, 0.1,
                                                   0.2, 0.2, 0.2, 0.2, 0.2, 0.2,
                                                   0.3, 0.3, 0.3, 0.3, 0.3, 0.3,
                                                   0.4, 0.4, 0.4, 0.4, 0.4, 0.4,]

        well_info['Measure'] = [True]*24
        pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

class TestPlateArray(unittest.TestCase):
    """
    Tests for the PlateArray class

    """
    def setUp(self):
        # Directory where to save temporary files
        self.temp_dir = "test/temp_plate_array"
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)
        # Set random seed
        random.seed(1)

    def tearDown(self):
        # Delete temporary directory
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def test_create(self):
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])

    def test_create_dim_mismatch_error(self):
        self.assertRaises(ValueError,
                          platedesign.plate.PlateArray,
                          name='A1',
                          array_n_rows=2,
                          array_n_cols=3,
                          plate_names=['P{}'.format(i+1)
                                       for i in range(4)])

    def test_default_attributes(self):
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Check all attributes
        self.assertEqual(p.name, 'A1')
        self.assertEqual(p.plate_names, ['P{}'.format(i+1)
                                         for i in range(6)])
        self.assertEqual(p.array_n_rows, 2)
        self.assertEqual(p.array_n_cols, 3)
        self.assertEqual(p.plate_n_rows, 4)
        self.assertEqual(p.plate_n_cols, 6)
        self.assertEqual(p.n_rows, 8)
        self.assertEqual(p.n_cols, 18)
        self.assertEqual(p.samples_to_measure, 144)
        self.assertIsNone(p.sample_media_vol)
        self.assertIsNone(p.total_media_vol)
        self.assertIsNone(p.cell_strain_name)
        self.assertIsNone(p.cell_setup_method)
        self.assertEqual(p.cell_predilution, 1)
        self.assertIsNone(p.cell_predilution_vol)
        self.assertIsNone(p.cell_initial_od600)
        self.assertIsNone(p.cell_shot_vol)
        self.assertEqual(p.metadata, collections.OrderedDict())
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_non_default_attributes(self):
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)],
                                         plate_n_rows=8,
                                         plate_n_cols=12)
        # Check all attributes
        self.assertEqual(p.name, 'A1')
        self.assertEqual(p.plate_names, ['P{}'.format(i+1)
                                         for i in range(6)])
        self.assertEqual(p.array_n_rows, 2)
        self.assertEqual(p.array_n_cols, 3)
        self.assertEqual(p.plate_n_rows, 8)
        self.assertEqual(p.plate_n_cols, 12)
        self.assertEqual(p.n_rows, 16)
        self.assertEqual(p.n_cols, 36)
        self.assertEqual(p.samples_to_measure, 576)
        self.assertIsNone(p.sample_media_vol)
        self.assertIsNone(p.total_media_vol)
        self.assertIsNone(p.cell_strain_name)
        self.assertIsNone(p.cell_setup_method)
        self.assertEqual(p.cell_predilution, 1)
        self.assertIsNone(p.cell_predilution_vol)
        self.assertIsNone(p.cell_initial_od600)
        self.assertIsNone(p.cell_shot_vol)
        self.assertEqual(p.metadata, collections.OrderedDict())
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_media_vol_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Set media volume
        p.total_media_vol = 80000
        p.sample_media_vol = 500
        # Get media volume for inducer application
        self.assertEqual(p.apply_inducer_media_vol('rows'), 500)
        self.assertEqual(p.apply_inducer_media_vol('cols'), 500)
        self.assertEqual(p.apply_inducer_media_vol('wells'), 500)
        self.assertEqual(p.apply_inducer_media_vol('media'), 80000)

    def test_apply_inducer_media_vol_error_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Set media volume
        p.total_media_vol = 80000
        p.sample_media_vol = 500
        # Limit number of samples to measure
        p.samples_to_measure = 80
        # Get media volume for inducer application
        self.assertEqual(p.apply_inducer_media_vol('wells'), 500)
        self.assertEqual(p.apply_inducer_media_vol('media'), 80000)
        # Verify that exception is raised for 'rows' and 'cols'
        self.assertRaises(ValueError, p.apply_inducer_media_vol, 'rows')
        self.assertRaises(ValueError, p.apply_inducer_media_vol, 'cols')

    def test_apply_inducer_media_vol_error_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Set media volume
        p.total_media_vol = 80000
        p.sample_media_vol = 500
        # Verify that exception is raised for an invalid 'apply_to' argument
        self.assertRaises(ValueError, p.apply_inducer_media_vol, 'all')

    def test_apply_inducer_media_vol_error_3(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Only set total media volume
        p.total_media_vol = 80000
        # Verify that exception is raised
        self.assertRaises(AttributeError, p.apply_inducer_media_vol, 'rows')

    def test_apply_inducer_media_vol_error_4(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Only set sample media volume
        p.sample_media_vol = 500
        # Verify that exception is raised
        self.assertRaises(AttributeError, p.apply_inducer_media_vol, 'rows')

    def test_apply_inducer_media_vol_error_5(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Don't set any media volume
        # Verify that exception is raised
        self.assertRaises(AttributeError, p.apply_inducer_media_vol, 'rows')

    def test_apply_inducer_n_shots(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Get number of shots for inducer application
        self.assertEqual(p.apply_inducer_n_shots('rows'), 8)
        self.assertEqual(p.apply_inducer_n_shots('cols'), 18)
        self.assertEqual(p.apply_inducer_n_shots('wells'), 1)
        self.assertEqual(p.apply_inducer_n_shots('media'), 1)

    def test_apply_inducer_n_shots_error_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Limit number of samples to measure
        p.samples_to_measure = 80
        # Get number of shots for inducer application
        self.assertEqual(p.apply_inducer_n_shots('wells'), 1)
        self.assertEqual(p.apply_inducer_n_shots('media'), 1)
        # Verify that exception is raised for 'rows' and 'cols'
        self.assertRaises(ValueError, p.apply_inducer_n_shots, 'rows')
        self.assertRaises(ValueError, p.apply_inducer_n_shots, 'cols')

    def test_apply_inducer_n_shots_error_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Verify that exception is raised for an invalid 'apply_to' argument
        self.assertRaises(ValueError, p.apply_inducer_n_shots, 'all')

    def test_apply_inducer_rows(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=18, scale='log')
        p.apply_inducer(iptg, apply_to='rows')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [iptg],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_rows_error_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=6, scale='log')
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='rows')
        # Check that inducers dictionary in plate has not been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_rows_error_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Limit number of samples to measure
        p.samples_to_measure = 80
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=18, scale='log')
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='rows')
        # Check that inducers dictionary in plate has not been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_cols(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=8, scale='log')
        p.apply_inducer(iptg, apply_to='cols')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [iptg],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_cols_error_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=4, scale='log')
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='cols')
        # Check that inducers dictionary in plate has not been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_cols_error_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Limit number of samples to measure
        p.samples_to_measure = 80
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=8, scale='log')
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='cols')
        # Check that inducers dictionary in plate has not been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_wells_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=144, scale='log')
        p.apply_inducer(iptg, apply_to='wells')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [iptg],
                                      'media': []})

    def test_apply_inducer_wells_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Limit number of samples to measure
        p.samples_to_measure = 80
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=80, scale='log')
        p.apply_inducer(iptg, apply_to='wells')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [iptg],
                                      'media': []})

    def test_apply_inducer_wells_error_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=24, scale='log')
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='wells')
        # Check that inducers dictionary in plate has not been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_wells_error_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Limit number of samples to measure
        p.samples_to_measure = 80
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=24, scale='log')
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='wells')
        # Check that inducers dictionary in plate has not been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_media_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [1]
        p.apply_inducer(iptg, apply_to='media')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': [iptg]})

    def test_apply_inducer_media_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Limit number of samples to measure
        p.samples_to_measure = 80
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [1]
        p.apply_inducer(iptg, apply_to='media')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': [iptg]})

    def test_apply_inducer_media_error_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [1, 10]
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='media')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_apply_inducer_error_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        # Set concentrations
        iptg.concentrations = [1]
        # Apply inducer to plate
        self.assertRaises(ValueError, p.apply_inducer, iptg, apply_to='all')
        # Check that inducers dictionary in plate has been updated
        self.assertEqual(p.inducers, {'rows': [],
                                      'cols': [],
                                      'wells': [],
                                      'media': []})

    def test_save_exp_setup_instructions_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Run save_exp_setup_instructions
        p.save_exp_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_exp.xlsx'))
        # save_exp_setup_instructions does not do anything in Plate. There is
        # no need to check for results.

    def test_save_exp_setup_instructions_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create new spreadsheet
        wb_test = openpyxl.Workbook()
        # Remove sheet created by default
        wb_test.remove_sheet(wb_test.active)
        # Run save_exp_setup_instructions
        p.save_exp_setup_instructions(workbook=wb_test)
        # save_exp_setup_instructions does not do anything in Plate. There is
        # no need to check for results.

    def test_save_exp_setup_instructions_argument_error(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Run save_exp_setup_instructions with no arguments
        self.assertRaises(ValueError, p.save_exp_setup_instructions)

    def test_save_exp_setup_files(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Check that save_exp_setup_files runs successfully
        p.save_exp_setup_files()
        # save_exp_setup_files does not do anything in Plate. There is no need
        # to check for results.

    def test_save_rep_setup_instructions_empty(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should only contain an empty sheet named "Sheet 1"
        self.assertEqual(wb.sheetnames, ["Sheet 1"])

    def test_save_rep_setup_instructions_cell_setup_fixed_volume_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.total_media_vol = 80000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_volume'
        p.cell_shot_vol = 5
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Cells for Plate Array A1"])
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertEqual(ws.cell(row=2, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=2, column=2).value, 5)
        self.assertEqual(ws.cell(row=2, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=3, column=1).value, "Add into 80.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_cell_setup_fixed_volume_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.total_media_vol = 80000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_volume'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_shot_vol = 5
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Cells for Plate Array A1"])
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertTrue('A2:C2' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=2, column=1).value, "Predilution")
        self.assertEqual(ws.cell(row=3, column=1).value, "Predilution factor")
        self.assertEqual(ws.cell(row=3, column=2).value, 100)
        self.assertEqual(ws.cell(row=3, column=3).value, "x")
        self.assertEqual(ws.cell(row=4, column=1).value, "Media volume")
        self.assertEqual(ws.cell(row=4, column=2).value, 990)
        self.assertEqual(ws.cell(row=4, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=5, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=5, column=2).value, 10)
        self.assertEqual(ws.cell(row=5, column=3).value, u"µL")
        self.assertTrue('A6:C6' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=6, column=1).value, "Inoculation")
        self.assertEqual(ws.cell(row=7, column=1).value, "Predilution volume")
        self.assertEqual(ws.cell(row=7, column=2).value, 5)
        self.assertEqual(ws.cell(row=7, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=8, column=1).value, "Add into 80.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_cell_setup_fixed_volume_error_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.total_media_vol = 80000.
        # Add some information for cell setup
        # Do not include shot volume
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_volume'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        # Run save_rep_setup_instructions
        self.assertRaises(
            ValueError,
            p.save_rep_setup_instructions,
            file_name=os.path.join(self.temp_dir, 'plate_rep.xlsx'))

    def test_save_rep_setup_instructions_cell_setup_fixed_volume_error_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.total_media_vol = 80000.
        # Add some information for cell setup
        # Do not include predilution volume
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_volume'
        p.cell_predilution = 100
        p.cell_shot_vol = 5
        # Run save_rep_setup_instructions
        self.assertRaises(
            ValueError,
            p.save_rep_setup_instructions,
            file_name=os.path.join(self.temp_dir, 'plate_rep.xlsx'))

    def test_save_rep_setup_instructions_cell_setup_fixed_od600_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.total_media_vol = 80000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_initial_od600 = 1e-5
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Cells for Plate Array A1"])
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertEqual(ws.cell(row=2, column=1).value, "Preculture OD600")
        self.assertEqual(ws.cell(row=2, column=2).value, None)
        self.assertEqual(ws.cell(row=2, column=3).value, None)
        self.assertEqual(ws.cell(row=3, column=1).value, "Target OD600")
        self.assertEqual(ws.cell(row=3, column=2).value, 1e-5)
        self.assertEqual(ws.cell(row=3, column=3).value, None)
        self.assertEqual(ws.cell(row=4, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=4, column=2).value, "=0.8/B2")
        self.assertEqual(ws.cell(row=4, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=5, column=1).value, "Add into 80.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_cell_setup_fixed_od600_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.total_media_vol = 80000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_initial_od600 = 1e-5
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Cells for Plate Array A1"])
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertTrue('A2:C2' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=2, column=1).value, "Predilution")
        self.assertEqual(ws.cell(row=3, column=1).value, "Predilution factor")
        self.assertEqual(ws.cell(row=3, column=2).value, 100)
        self.assertEqual(ws.cell(row=3, column=3).value, "x")
        self.assertEqual(ws.cell(row=4, column=1).value, "Media volume")
        self.assertEqual(ws.cell(row=4, column=2).value, 990)
        self.assertEqual(ws.cell(row=4, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=5, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=5, column=2).value, 10)
        self.assertEqual(ws.cell(row=5, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=6, column=1).value, "Predilution OD600")
        self.assertEqual(ws.cell(row=6, column=2).value, None)
        self.assertEqual(ws.cell(row=6, column=3).value, None)
        self.assertTrue('A7:C7' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=7, column=1).value, "Inoculation")
        self.assertEqual(ws.cell(row=8, column=1).value, "Target OD600")
        self.assertEqual(ws.cell(row=8, column=2).value, 1e-5)
        self.assertEqual(ws.cell(row=8, column=3).value, None)
        self.assertEqual(ws.cell(row=9, column=1).value, "Predilution volume")
        self.assertEqual(ws.cell(row=9, column=2).value, "=0.8/B6")
        self.assertEqual(ws.cell(row=9, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=10, column=1).value, "Add into 80.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_cell_setup_fixed_od600_error_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.total_media_vol = 80000.
        # Add some information for cell setup
        # Do not include target od600
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        # Run save_rep_setup_instructions
        self.assertRaises(
            ValueError,
            p.save_rep_setup_instructions,
            file_name=os.path.join(self.temp_dir, 'plate_rep.xlsx'))

    def test_save_rep_setup_instructions_cell_setup_fixed_od600_error_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.total_media_vol = 80000.
        # Add some information for cell setup
        # Do not include predilution volume
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_initial_od600 = 1e-5
        # Run save_rep_setup_instructions
        self.assertRaises(
            ValueError,
            p.save_rep_setup_instructions,
            file_name=os.path.join(self.temp_dir, 'plate_rep.xlsx'))

    def test_save_rep_setup_instructions_inducer_rows_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=18, scale='log')
        p.apply_inducer(iptg, apply_to='rows')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I001")
        self.assertEqual(ws.cell(row=1, column=2).value, "I002")
        self.assertEqual(ws.cell(row=1, column=3).value, "I003")
        self.assertEqual(ws.cell(row=1, column=4).value, "I004")
        self.assertEqual(ws.cell(row=1, column=5).value, "I005")
        self.assertEqual(ws.cell(row=1, column=6).value, "I006")
        self.assertEqual(ws.cell(row=1, column=7).value, "I007")
        self.assertEqual(ws.cell(row=1, column=8).value, "I008")
        self.assertEqual(ws.cell(row=1, column=9).value, "I009")
        self.assertEqual(ws.cell(row=1, column=10).value, "I010")
        self.assertEqual(ws.cell(row=1, column=11).value, "I011")
        self.assertEqual(ws.cell(row=1, column=12).value, "I012")
        self.assertEqual(ws.cell(row=1, column=13).value, "I013")
        self.assertEqual(ws.cell(row=1, column=14).value, "I014")
        self.assertEqual(ws.cell(row=1, column=15).value, "I015")
        self.assertEqual(ws.cell(row=1, column=16).value, "I016")
        self.assertEqual(ws.cell(row=1, column=17).value, "I017")
        self.assertEqual(ws.cell(row=1, column=18).value, "I018")
        self.assertEqual(ws.cell(row=2, column=1).value, "P1 (1, 1)")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (1, 2)")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (1, 3)")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (1, 4)")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (1, 5)")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (1, 6)")
        self.assertEqual(ws.cell(row=2, column=7).value, "P2 (1, 1)")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (1, 2)")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (1, 3)")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (1, 4)")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (1, 5)")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (1, 6)")
        self.assertEqual(ws.cell(row=2, column=13).value, "P3 (1, 1)")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (1, 2)")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (1, 3)")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (1, 4)")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (1, 5)")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (1, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "P1 (2, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (2, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (2, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (2, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (2, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (2, 6)")
        self.assertEqual(ws.cell(row=3, column=7).value, "P2 (2, 1)")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (2, 2)")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (2, 3)")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (2, 4)")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (2, 5)")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (2, 6)")
        self.assertEqual(ws.cell(row=3, column=13).value, "P3 (2, 1)")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (2, 2)")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (2, 3)")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (2, 4)")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (2, 5)")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (2, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "P1 (3, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (3, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (3, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (3, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (3, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (3, 6)")
        self.assertEqual(ws.cell(row=4, column=7).value, "P2 (3, 1)")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (3, 2)")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (3, 3)")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (3, 4)")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (3, 5)")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (3, 6)")
        self.assertEqual(ws.cell(row=4, column=13).value, "P3 (3, 1)")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (3, 2)")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (3, 3)")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (3, 4)")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (3, 5)")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (3, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "P1 (4, 1)")
        self.assertEqual(ws.cell(row=5, column=2).value, "P1 (4, 2)")
        self.assertEqual(ws.cell(row=5, column=3).value, "P1 (4, 3)")
        self.assertEqual(ws.cell(row=5, column=4).value, "P1 (4, 4)")
        self.assertEqual(ws.cell(row=5, column=5).value, "P1 (4, 5)")
        self.assertEqual(ws.cell(row=5, column=6).value, "P1 (4, 6)")
        self.assertEqual(ws.cell(row=5, column=7).value, "P2 (4, 1)")
        self.assertEqual(ws.cell(row=5, column=8).value, "P2 (4, 2)")
        self.assertEqual(ws.cell(row=5, column=9).value, "P2 (4, 3)")
        self.assertEqual(ws.cell(row=5, column=10).value, "P2 (4, 4)")
        self.assertEqual(ws.cell(row=5, column=11).value, "P2 (4, 5)")
        self.assertEqual(ws.cell(row=5, column=12).value, "P2 (4, 6)")
        self.assertEqual(ws.cell(row=5, column=13).value, "P3 (4, 1)")
        self.assertEqual(ws.cell(row=5, column=14).value, "P3 (4, 2)")
        self.assertEqual(ws.cell(row=5, column=15).value, "P3 (4, 3)")
        self.assertEqual(ws.cell(row=5, column=16).value, "P3 (4, 4)")
        self.assertEqual(ws.cell(row=5, column=17).value, "P3 (4, 5)")
        self.assertEqual(ws.cell(row=5, column=18).value, "P3 (4, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "P4 (1, 1)")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (1, 2)")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (1, 3)")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (1, 4)")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (1, 5)")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (1, 6)")
        self.assertEqual(ws.cell(row=6, column=7).value, "P5 (1, 1)")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (1, 2)")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (1, 3)")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (1, 4)")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (1, 5)")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (1, 6)")
        self.assertEqual(ws.cell(row=6, column=13).value, "P6 (1, 1)")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (1, 2)")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (1, 3)")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (1, 4)")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (1, 5)")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (1, 6)")
        self.assertEqual(ws.cell(row=7, column=1).value, "P4 (2, 1)")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (2, 2)")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (2, 3)")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (2, 4)")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (2, 5)")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (2, 6)")
        self.assertEqual(ws.cell(row=7, column=7).value, "P5 (2, 1)")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (2, 2)")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (2, 3)")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (2, 4)")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (2, 5)")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (2, 6)")
        self.assertEqual(ws.cell(row=7, column=13).value, "P6 (2, 1)")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (2, 2)")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (2, 3)")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (2, 4)")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (2, 5)")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (2, 6)")
        self.assertEqual(ws.cell(row=8, column=1).value, "P4 (3, 1)")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (3, 2)")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (3, 3)")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (3, 4)")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (3, 5)")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (3, 6)")
        self.assertEqual(ws.cell(row=8, column=7).value, "P5 (3, 1)")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (3, 2)")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (3, 3)")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (3, 4)")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (3, 5)")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (3, 6)")
        self.assertEqual(ws.cell(row=8, column=13).value, "P6 (3, 1)")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (3, 2)")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (3, 3)")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (3, 4)")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (3, 5)")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (3, 6)")
        self.assertEqual(ws.cell(row=9, column=1).value, "P4 (4, 1)")
        self.assertEqual(ws.cell(row=9, column=2).value, "P4 (4, 2)")
        self.assertEqual(ws.cell(row=9, column=3).value, "P4 (4, 3)")
        self.assertEqual(ws.cell(row=9, column=4).value, "P4 (4, 4)")
        self.assertEqual(ws.cell(row=9, column=5).value, "P4 (4, 5)")
        self.assertEqual(ws.cell(row=9, column=6).value, "P4 (4, 6)")
        self.assertEqual(ws.cell(row=9, column=7).value, "P5 (4, 1)")
        self.assertEqual(ws.cell(row=9, column=8).value, "P5 (4, 2)")
        self.assertEqual(ws.cell(row=9, column=9).value, "P5 (4, 3)")
        self.assertEqual(ws.cell(row=9, column=10).value, "P5 (4, 4)")
        self.assertEqual(ws.cell(row=9, column=11).value, "P5 (4, 5)")
        self.assertEqual(ws.cell(row=9, column=12).value, "P5 (4, 6)")
        self.assertEqual(ws.cell(row=9, column=13).value, "P6 (4, 1)")
        self.assertEqual(ws.cell(row=9, column=14).value, "P6 (4, 2)")
        self.assertEqual(ws.cell(row=9, column=15).value, "P6 (4, 3)")
        self.assertEqual(ws.cell(row=9, column=16).value, "P6 (4, 4)")
        self.assertEqual(ws.cell(row=9, column=17).value, "P6 (4, 5)")
        self.assertEqual(ws.cell(row=9, column=18).value, "P6 (4, 6)")

    def test_save_rep_setup_instructions_inducer_rows_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=18, scale='log')
        p.apply_inducer(iptg, apply_to='rows')
        # Create second inducer for plate rows
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=18, scale='log')
        p.apply_inducer(atc, apply_to='rows')

        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I001")
        self.assertEqual(ws.cell(row=1, column=2).value, "I002")
        self.assertEqual(ws.cell(row=1, column=3).value, "I003")
        self.assertEqual(ws.cell(row=1, column=4).value, "I004")
        self.assertEqual(ws.cell(row=1, column=5).value, "I005")
        self.assertEqual(ws.cell(row=1, column=6).value, "I006")
        self.assertEqual(ws.cell(row=1, column=7).value, "I007")
        self.assertEqual(ws.cell(row=1, column=8).value, "I008")
        self.assertEqual(ws.cell(row=1, column=9).value, "I009")
        self.assertEqual(ws.cell(row=1, column=10).value, "I010")
        self.assertEqual(ws.cell(row=1, column=11).value, "I011")
        self.assertEqual(ws.cell(row=1, column=12).value, "I012")
        self.assertEqual(ws.cell(row=1, column=13).value, "I013")
        self.assertEqual(ws.cell(row=1, column=14).value, "I014")
        self.assertEqual(ws.cell(row=1, column=15).value, "I015")
        self.assertEqual(ws.cell(row=1, column=16).value, "I016")
        self.assertEqual(ws.cell(row=1, column=17).value, "I017")
        self.assertEqual(ws.cell(row=1, column=18).value, "I018")
        self.assertEqual(ws.cell(row=2, column=1).value, "a001")
        self.assertEqual(ws.cell(row=2, column=2).value, "a002")
        self.assertEqual(ws.cell(row=2, column=3).value, "a003")
        self.assertEqual(ws.cell(row=2, column=4).value, "a004")
        self.assertEqual(ws.cell(row=2, column=5).value, "a005")
        self.assertEqual(ws.cell(row=2, column=6).value, "a006")
        self.assertEqual(ws.cell(row=2, column=7).value, "a007")
        self.assertEqual(ws.cell(row=2, column=8).value, "a008")
        self.assertEqual(ws.cell(row=2, column=9).value, "a009")
        self.assertEqual(ws.cell(row=2, column=10).value, "a010")
        self.assertEqual(ws.cell(row=2, column=11).value, "a011")
        self.assertEqual(ws.cell(row=2, column=12).value, "a012")
        self.assertEqual(ws.cell(row=2, column=13).value, "a013")
        self.assertEqual(ws.cell(row=2, column=14).value, "a014")
        self.assertEqual(ws.cell(row=2, column=15).value, "a015")
        self.assertEqual(ws.cell(row=2, column=16).value, "a016")
        self.assertEqual(ws.cell(row=2, column=17).value, "a017")
        self.assertEqual(ws.cell(row=2, column=18).value, "a018")
        self.assertEqual(ws.cell(row=3, column=1).value, "P1 (1, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (1, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (1, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (1, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (1, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (1, 6)")
        self.assertEqual(ws.cell(row=3, column=7).value, "P2 (1, 1)")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (1, 2)")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (1, 3)")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (1, 4)")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (1, 5)")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (1, 6)")
        self.assertEqual(ws.cell(row=3, column=13).value, "P3 (1, 1)")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (1, 2)")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (1, 3)")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (1, 4)")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (1, 5)")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (1, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "P1 (2, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (2, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (2, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (2, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (2, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (2, 6)")
        self.assertEqual(ws.cell(row=4, column=7).value, "P2 (2, 1)")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (2, 2)")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (2, 3)")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (2, 4)")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (2, 5)")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (2, 6)")
        self.assertEqual(ws.cell(row=4, column=13).value, "P3 (2, 1)")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (2, 2)")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (2, 3)")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (2, 4)")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (2, 5)")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (2, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "P1 (3, 1)")
        self.assertEqual(ws.cell(row=5, column=2).value, "P1 (3, 2)")
        self.assertEqual(ws.cell(row=5, column=3).value, "P1 (3, 3)")
        self.assertEqual(ws.cell(row=5, column=4).value, "P1 (3, 4)")
        self.assertEqual(ws.cell(row=5, column=5).value, "P1 (3, 5)")
        self.assertEqual(ws.cell(row=5, column=6).value, "P1 (3, 6)")
        self.assertEqual(ws.cell(row=5, column=7).value, "P2 (3, 1)")
        self.assertEqual(ws.cell(row=5, column=8).value, "P2 (3, 2)")
        self.assertEqual(ws.cell(row=5, column=9).value, "P2 (3, 3)")
        self.assertEqual(ws.cell(row=5, column=10).value, "P2 (3, 4)")
        self.assertEqual(ws.cell(row=5, column=11).value, "P2 (3, 5)")
        self.assertEqual(ws.cell(row=5, column=12).value, "P2 (3, 6)")
        self.assertEqual(ws.cell(row=5, column=13).value, "P3 (3, 1)")
        self.assertEqual(ws.cell(row=5, column=14).value, "P3 (3, 2)")
        self.assertEqual(ws.cell(row=5, column=15).value, "P3 (3, 3)")
        self.assertEqual(ws.cell(row=5, column=16).value, "P3 (3, 4)")
        self.assertEqual(ws.cell(row=5, column=17).value, "P3 (3, 5)")
        self.assertEqual(ws.cell(row=5, column=18).value, "P3 (3, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "P1 (4, 1)")
        self.assertEqual(ws.cell(row=6, column=2).value, "P1 (4, 2)")
        self.assertEqual(ws.cell(row=6, column=3).value, "P1 (4, 3)")
        self.assertEqual(ws.cell(row=6, column=4).value, "P1 (4, 4)")
        self.assertEqual(ws.cell(row=6, column=5).value, "P1 (4, 5)")
        self.assertEqual(ws.cell(row=6, column=6).value, "P1 (4, 6)")
        self.assertEqual(ws.cell(row=6, column=7).value, "P2 (4, 1)")
        self.assertEqual(ws.cell(row=6, column=8).value, "P2 (4, 2)")
        self.assertEqual(ws.cell(row=6, column=9).value, "P2 (4, 3)")
        self.assertEqual(ws.cell(row=6, column=10).value, "P2 (4, 4)")
        self.assertEqual(ws.cell(row=6, column=11).value, "P2 (4, 5)")
        self.assertEqual(ws.cell(row=6, column=12).value, "P2 (4, 6)")
        self.assertEqual(ws.cell(row=6, column=13).value, "P3 (4, 1)")
        self.assertEqual(ws.cell(row=6, column=14).value, "P3 (4, 2)")
        self.assertEqual(ws.cell(row=6, column=15).value, "P3 (4, 3)")
        self.assertEqual(ws.cell(row=6, column=16).value, "P3 (4, 4)")
        self.assertEqual(ws.cell(row=6, column=17).value, "P3 (4, 5)")
        self.assertEqual(ws.cell(row=6, column=18).value, "P3 (4, 6)")
        self.assertEqual(ws.cell(row=7, column=1).value, "P4 (1, 1)")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (1, 2)")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (1, 3)")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (1, 4)")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (1, 5)")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (1, 6)")
        self.assertEqual(ws.cell(row=7, column=7).value, "P5 (1, 1)")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (1, 2)")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (1, 3)")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (1, 4)")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (1, 5)")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (1, 6)")
        self.assertEqual(ws.cell(row=7, column=13).value, "P6 (1, 1)")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (1, 2)")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (1, 3)")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (1, 4)")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (1, 5)")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (1, 6)")
        self.assertEqual(ws.cell(row=8, column=1).value, "P4 (2, 1)")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (2, 2)")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (2, 3)")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (2, 4)")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (2, 5)")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (2, 6)")
        self.assertEqual(ws.cell(row=8, column=7).value, "P5 (2, 1)")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (2, 2)")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (2, 3)")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (2, 4)")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (2, 5)")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (2, 6)")
        self.assertEqual(ws.cell(row=8, column=13).value, "P6 (2, 1)")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (2, 2)")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (2, 3)")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (2, 4)")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (2, 5)")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (2, 6)")
        self.assertEqual(ws.cell(row=9, column=1).value, "P4 (3, 1)")
        self.assertEqual(ws.cell(row=9, column=2).value, "P4 (3, 2)")
        self.assertEqual(ws.cell(row=9, column=3).value, "P4 (3, 3)")
        self.assertEqual(ws.cell(row=9, column=4).value, "P4 (3, 4)")
        self.assertEqual(ws.cell(row=9, column=5).value, "P4 (3, 5)")
        self.assertEqual(ws.cell(row=9, column=6).value, "P4 (3, 6)")
        self.assertEqual(ws.cell(row=9, column=7).value, "P5 (3, 1)")
        self.assertEqual(ws.cell(row=9, column=8).value, "P5 (3, 2)")
        self.assertEqual(ws.cell(row=9, column=9).value, "P5 (3, 3)")
        self.assertEqual(ws.cell(row=9, column=10).value, "P5 (3, 4)")
        self.assertEqual(ws.cell(row=9, column=11).value, "P5 (3, 5)")
        self.assertEqual(ws.cell(row=9, column=12).value, "P5 (3, 6)")
        self.assertEqual(ws.cell(row=9, column=13).value, "P6 (3, 1)")
        self.assertEqual(ws.cell(row=9, column=14).value, "P6 (3, 2)")
        self.assertEqual(ws.cell(row=9, column=15).value, "P6 (3, 3)")
        self.assertEqual(ws.cell(row=9, column=16).value, "P6 (3, 4)")
        self.assertEqual(ws.cell(row=9, column=17).value, "P6 (3, 5)")
        self.assertEqual(ws.cell(row=9, column=18).value, "P6 (3, 6)")
        self.assertEqual(ws.cell(row=10, column=1).value, "P4 (4, 1)")
        self.assertEqual(ws.cell(row=10, column=2).value, "P4 (4, 2)")
        self.assertEqual(ws.cell(row=10, column=3).value, "P4 (4, 3)")
        self.assertEqual(ws.cell(row=10, column=4).value, "P4 (4, 4)")
        self.assertEqual(ws.cell(row=10, column=5).value, "P4 (4, 5)")
        self.assertEqual(ws.cell(row=10, column=6).value, "P4 (4, 6)")
        self.assertEqual(ws.cell(row=10, column=7).value, "P5 (4, 1)")
        self.assertEqual(ws.cell(row=10, column=8).value, "P5 (4, 2)")
        self.assertEqual(ws.cell(row=10, column=9).value, "P5 (4, 3)")
        self.assertEqual(ws.cell(row=10, column=10).value, "P5 (4, 4)")
        self.assertEqual(ws.cell(row=10, column=11).value, "P5 (4, 5)")
        self.assertEqual(ws.cell(row=10, column=12).value, "P5 (4, 6)")
        self.assertEqual(ws.cell(row=10, column=13).value, "P6 (4, 1)")
        self.assertEqual(ws.cell(row=10, column=14).value, "P6 (4, 2)")
        self.assertEqual(ws.cell(row=10, column=15).value, "P6 (4, 3)")
        self.assertEqual(ws.cell(row=10, column=16).value, "P6 (4, 4)")
        self.assertEqual(ws.cell(row=10, column=17).value, "P6 (4, 5)")
        self.assertEqual(ws.cell(row=10, column=18).value, "P6 (4, 6)")

    def test_save_rep_setup_instructions_inducer_rows_3(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=18, scale='log')
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='rows')
        # Create second inducer for plate rows
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=18, scale='log')
        p.apply_inducer(atc, apply_to='rows')

        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I012")
        self.assertEqual(ws.cell(row=1, column=2).value, "I016")
        self.assertEqual(ws.cell(row=1, column=3).value, "I018")
        self.assertEqual(ws.cell(row=1, column=4).value, "I002")
        self.assertEqual(ws.cell(row=1, column=5).value, "I011")
        self.assertEqual(ws.cell(row=1, column=6).value, "I005")
        self.assertEqual(ws.cell(row=1, column=7).value, "I017")
        self.assertEqual(ws.cell(row=1, column=8).value, "I014")
        self.assertEqual(ws.cell(row=1, column=9).value, "I010")
        self.assertEqual(ws.cell(row=1, column=10).value, "I001")
        self.assertEqual(ws.cell(row=1, column=11).value, "I009")
        self.assertEqual(ws.cell(row=1, column=12).value, "I008")
        self.assertEqual(ws.cell(row=1, column=13).value, "I006")
        self.assertEqual(ws.cell(row=1, column=14).value, "I007")
        self.assertEqual(ws.cell(row=1, column=15).value, "I004")
        self.assertEqual(ws.cell(row=1, column=16).value, "I013")
        self.assertEqual(ws.cell(row=1, column=17).value, "I015")
        self.assertEqual(ws.cell(row=1, column=18).value, "I003")
        self.assertEqual(ws.cell(row=2, column=1).value, "a001")
        self.assertEqual(ws.cell(row=2, column=2).value, "a002")
        self.assertEqual(ws.cell(row=2, column=3).value, "a003")
        self.assertEqual(ws.cell(row=2, column=4).value, "a004")
        self.assertEqual(ws.cell(row=2, column=5).value, "a005")
        self.assertEqual(ws.cell(row=2, column=6).value, "a006")
        self.assertEqual(ws.cell(row=2, column=7).value, "a007")
        self.assertEqual(ws.cell(row=2, column=8).value, "a008")
        self.assertEqual(ws.cell(row=2, column=9).value, "a009")
        self.assertEqual(ws.cell(row=2, column=10).value, "a010")
        self.assertEqual(ws.cell(row=2, column=11).value, "a011")
        self.assertEqual(ws.cell(row=2, column=12).value, "a012")
        self.assertEqual(ws.cell(row=2, column=13).value, "a013")
        self.assertEqual(ws.cell(row=2, column=14).value, "a014")
        self.assertEqual(ws.cell(row=2, column=15).value, "a015")
        self.assertEqual(ws.cell(row=2, column=16).value, "a016")
        self.assertEqual(ws.cell(row=2, column=17).value, "a017")
        self.assertEqual(ws.cell(row=2, column=18).value, "a018")
        self.assertEqual(ws.cell(row=3, column=1).value, "P1 (1, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (1, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (1, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (1, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (1, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (1, 6)")
        self.assertEqual(ws.cell(row=3, column=7).value, "P2 (1, 1)")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (1, 2)")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (1, 3)")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (1, 4)")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (1, 5)")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (1, 6)")
        self.assertEqual(ws.cell(row=3, column=13).value, "P3 (1, 1)")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (1, 2)")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (1, 3)")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (1, 4)")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (1, 5)")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (1, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "P1 (2, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (2, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (2, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (2, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (2, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (2, 6)")
        self.assertEqual(ws.cell(row=4, column=7).value, "P2 (2, 1)")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (2, 2)")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (2, 3)")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (2, 4)")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (2, 5)")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (2, 6)")
        self.assertEqual(ws.cell(row=4, column=13).value, "P3 (2, 1)")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (2, 2)")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (2, 3)")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (2, 4)")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (2, 5)")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (2, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "P1 (3, 1)")
        self.assertEqual(ws.cell(row=5, column=2).value, "P1 (3, 2)")
        self.assertEqual(ws.cell(row=5, column=3).value, "P1 (3, 3)")
        self.assertEqual(ws.cell(row=5, column=4).value, "P1 (3, 4)")
        self.assertEqual(ws.cell(row=5, column=5).value, "P1 (3, 5)")
        self.assertEqual(ws.cell(row=5, column=6).value, "P1 (3, 6)")
        self.assertEqual(ws.cell(row=5, column=7).value, "P2 (3, 1)")
        self.assertEqual(ws.cell(row=5, column=8).value, "P2 (3, 2)")
        self.assertEqual(ws.cell(row=5, column=9).value, "P2 (3, 3)")
        self.assertEqual(ws.cell(row=5, column=10).value, "P2 (3, 4)")
        self.assertEqual(ws.cell(row=5, column=11).value, "P2 (3, 5)")
        self.assertEqual(ws.cell(row=5, column=12).value, "P2 (3, 6)")
        self.assertEqual(ws.cell(row=5, column=13).value, "P3 (3, 1)")
        self.assertEqual(ws.cell(row=5, column=14).value, "P3 (3, 2)")
        self.assertEqual(ws.cell(row=5, column=15).value, "P3 (3, 3)")
        self.assertEqual(ws.cell(row=5, column=16).value, "P3 (3, 4)")
        self.assertEqual(ws.cell(row=5, column=17).value, "P3 (3, 5)")
        self.assertEqual(ws.cell(row=5, column=18).value, "P3 (3, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "P1 (4, 1)")
        self.assertEqual(ws.cell(row=6, column=2).value, "P1 (4, 2)")
        self.assertEqual(ws.cell(row=6, column=3).value, "P1 (4, 3)")
        self.assertEqual(ws.cell(row=6, column=4).value, "P1 (4, 4)")
        self.assertEqual(ws.cell(row=6, column=5).value, "P1 (4, 5)")
        self.assertEqual(ws.cell(row=6, column=6).value, "P1 (4, 6)")
        self.assertEqual(ws.cell(row=6, column=7).value, "P2 (4, 1)")
        self.assertEqual(ws.cell(row=6, column=8).value, "P2 (4, 2)")
        self.assertEqual(ws.cell(row=6, column=9).value, "P2 (4, 3)")
        self.assertEqual(ws.cell(row=6, column=10).value, "P2 (4, 4)")
        self.assertEqual(ws.cell(row=6, column=11).value, "P2 (4, 5)")
        self.assertEqual(ws.cell(row=6, column=12).value, "P2 (4, 6)")
        self.assertEqual(ws.cell(row=6, column=13).value, "P3 (4, 1)")
        self.assertEqual(ws.cell(row=6, column=14).value, "P3 (4, 2)")
        self.assertEqual(ws.cell(row=6, column=15).value, "P3 (4, 3)")
        self.assertEqual(ws.cell(row=6, column=16).value, "P3 (4, 4)")
        self.assertEqual(ws.cell(row=6, column=17).value, "P3 (4, 5)")
        self.assertEqual(ws.cell(row=6, column=18).value, "P3 (4, 6)")
        self.assertEqual(ws.cell(row=7, column=1).value, "P4 (1, 1)")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (1, 2)")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (1, 3)")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (1, 4)")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (1, 5)")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (1, 6)")
        self.assertEqual(ws.cell(row=7, column=7).value, "P5 (1, 1)")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (1, 2)")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (1, 3)")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (1, 4)")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (1, 5)")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (1, 6)")
        self.assertEqual(ws.cell(row=7, column=13).value, "P6 (1, 1)")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (1, 2)")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (1, 3)")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (1, 4)")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (1, 5)")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (1, 6)")
        self.assertEqual(ws.cell(row=8, column=1).value, "P4 (2, 1)")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (2, 2)")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (2, 3)")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (2, 4)")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (2, 5)")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (2, 6)")
        self.assertEqual(ws.cell(row=8, column=7).value, "P5 (2, 1)")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (2, 2)")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (2, 3)")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (2, 4)")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (2, 5)")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (2, 6)")
        self.assertEqual(ws.cell(row=8, column=13).value, "P6 (2, 1)")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (2, 2)")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (2, 3)")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (2, 4)")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (2, 5)")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (2, 6)")
        self.assertEqual(ws.cell(row=9, column=1).value, "P4 (3, 1)")
        self.assertEqual(ws.cell(row=9, column=2).value, "P4 (3, 2)")
        self.assertEqual(ws.cell(row=9, column=3).value, "P4 (3, 3)")
        self.assertEqual(ws.cell(row=9, column=4).value, "P4 (3, 4)")
        self.assertEqual(ws.cell(row=9, column=5).value, "P4 (3, 5)")
        self.assertEqual(ws.cell(row=9, column=6).value, "P4 (3, 6)")
        self.assertEqual(ws.cell(row=9, column=7).value, "P5 (3, 1)")
        self.assertEqual(ws.cell(row=9, column=8).value, "P5 (3, 2)")
        self.assertEqual(ws.cell(row=9, column=9).value, "P5 (3, 3)")
        self.assertEqual(ws.cell(row=9, column=10).value, "P5 (3, 4)")
        self.assertEqual(ws.cell(row=9, column=11).value, "P5 (3, 5)")
        self.assertEqual(ws.cell(row=9, column=12).value, "P5 (3, 6)")
        self.assertEqual(ws.cell(row=9, column=13).value, "P6 (3, 1)")
        self.assertEqual(ws.cell(row=9, column=14).value, "P6 (3, 2)")
        self.assertEqual(ws.cell(row=9, column=15).value, "P6 (3, 3)")
        self.assertEqual(ws.cell(row=9, column=16).value, "P6 (3, 4)")
        self.assertEqual(ws.cell(row=9, column=17).value, "P6 (3, 5)")
        self.assertEqual(ws.cell(row=9, column=18).value, "P6 (3, 6)")
        self.assertEqual(ws.cell(row=10, column=1).value, "P4 (4, 1)")
        self.assertEqual(ws.cell(row=10, column=2).value, "P4 (4, 2)")
        self.assertEqual(ws.cell(row=10, column=3).value, "P4 (4, 3)")
        self.assertEqual(ws.cell(row=10, column=4).value, "P4 (4, 4)")
        self.assertEqual(ws.cell(row=10, column=5).value, "P4 (4, 5)")
        self.assertEqual(ws.cell(row=10, column=6).value, "P4 (4, 6)")
        self.assertEqual(ws.cell(row=10, column=7).value, "P5 (4, 1)")
        self.assertEqual(ws.cell(row=10, column=8).value, "P5 (4, 2)")
        self.assertEqual(ws.cell(row=10, column=9).value, "P5 (4, 3)")
        self.assertEqual(ws.cell(row=10, column=10).value, "P5 (4, 4)")
        self.assertEqual(ws.cell(row=10, column=11).value, "P5 (4, 5)")
        self.assertEqual(ws.cell(row=10, column=12).value, "P5 (4, 6)")
        self.assertEqual(ws.cell(row=10, column=13).value, "P6 (4, 1)")
        self.assertEqual(ws.cell(row=10, column=14).value, "P6 (4, 2)")
        self.assertEqual(ws.cell(row=10, column=15).value, "P6 (4, 3)")
        self.assertEqual(ws.cell(row=10, column=16).value, "P6 (4, 4)")
        self.assertEqual(ws.cell(row=10, column=17).value, "P6 (4, 5)")
        self.assertEqual(ws.cell(row=10, column=18).value, "P6 (4, 6)")

    def test_save_rep_setup_instructions_inducer_cols_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer for plate columns
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=8, scale='log')
        p.apply_inducer(iptg, apply_to='cols')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I001")
        self.assertEqual(ws.cell(row=1, column=2).value, "P1 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=3).value, "P1 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=4).value, "P1 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=5).value, "P1 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=6).value, "P1 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=7).value, "P1 (1, 6)")
        self.assertEqual(ws.cell(row=1, column=8).value, "P2 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=9).value, "P2 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=10).value, "P2 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=11).value, "P2 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=12).value, "P2 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=13).value, "P2 (1, 6)")
        self.assertEqual(ws.cell(row=1, column=14).value, "P3 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=15).value, "P3 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=16).value, "P3 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=17).value, "P3 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=18).value, "P3 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=19).value, "P3 (1, 6)")
        self.assertEqual(ws.cell(row=2, column=1).value, "I002")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=7).value, "P1 (2, 6)")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=13).value, "P2 (2, 6)")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=19).value, "P3 (2, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "I003")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=7).value, "P1 (3, 6)")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=13).value, "P2 (3, 6)")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=19).value, "P3 (3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "I004")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=7).value, "P1 (4, 6)")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=13).value, "P2 (4, 6)")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=19).value, "P3 (4, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "I005")
        self.assertEqual(ws.cell(row=5, column=2).value, "P4 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=3).value, "P4 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=4).value, "P4 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=5).value, "P4 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=6).value, "P4 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=7).value, "P4 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=8).value, "P5 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=9).value, "P5 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=10).value, "P5 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=11).value, "P5 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=12).value, "P5 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=13).value, "P5 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=14).value, "P6 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=15).value, "P6 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=16).value, "P6 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=17).value, "P6 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=18).value, "P6 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=19).value, "P6 (1, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "I006")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=7).value, "P4 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=13).value, "P5 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=19).value, "P6 (2, 6)")
        self.assertEqual(ws.cell(row=7, column=1).value, "I007")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=7).value, "P4 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=13).value, "P5 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=19).value, "P6 (3, 6)")
        self.assertEqual(ws.cell(row=8, column=1).value, "I008")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=7).value, "P4 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=13).value, "P5 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=19).value, "P6 (4, 6)")

    def test_save_rep_setup_instructions_inducer_cols_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer for plate columns
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=8, scale='log')
        p.apply_inducer(iptg, apply_to='cols')
        # Create second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=8, scale='log')
        p.apply_inducer(atc, apply_to='cols')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I001")
        self.assertEqual(ws.cell(row=1, column=2).value, "a001")
        self.assertEqual(ws.cell(row=1, column=3).value, "P1 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=4).value, "P1 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=5).value, "P1 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=6).value, "P1 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=7).value, "P1 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=8).value, "P1 (1, 6)")
        self.assertEqual(ws.cell(row=1, column=9).value, "P2 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=10).value, "P2 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=11).value, "P2 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=12).value, "P2 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=13).value, "P2 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=14).value, "P2 (1, 6)")
        self.assertEqual(ws.cell(row=1, column=15).value, "P3 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=16).value, "P3 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=17).value, "P3 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=18).value, "P3 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=19).value, "P3 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=20).value, "P3 (1, 6)")
        self.assertEqual(ws.cell(row=2, column=1).value, "I002")
        self.assertEqual(ws.cell(row=2, column=2).value, "a002")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=7).value, "P1 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=8).value, "P1 (2, 6)")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=13).value, "P2 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=14).value, "P2 (2, 6)")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=19).value, "P3 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=20).value, "P3 (2, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "I003")
        self.assertEqual(ws.cell(row=3, column=2).value, "a003")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=7).value, "P1 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=8).value, "P1 (3, 6)")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=13).value, "P2 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=14).value, "P2 (3, 6)")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=19).value, "P3 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=20).value, "P3 (3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "I004")
        self.assertEqual(ws.cell(row=4, column=2).value, "a004")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=7).value, "P1 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=8).value, "P1 (4, 6)")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=13).value, "P2 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=14).value, "P2 (4, 6)")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=19).value, "P3 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=20).value, "P3 (4, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "I005")
        self.assertEqual(ws.cell(row=5, column=2).value, "a005")
        self.assertEqual(ws.cell(row=5, column=3).value, "P4 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=4).value, "P4 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=5).value, "P4 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=6).value, "P4 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=7).value, "P4 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=8).value, "P4 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=9).value, "P5 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=10).value, "P5 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=11).value, "P5 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=12).value, "P5 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=13).value, "P5 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=14).value, "P5 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=15).value, "P6 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=16).value, "P6 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=17).value, "P6 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=18).value, "P6 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=19).value, "P6 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=20).value, "P6 (1, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "I006")
        self.assertEqual(ws.cell(row=6, column=2).value, "a006")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=7).value, "P4 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=8).value, "P4 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=13).value, "P5 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=14).value, "P5 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=19).value, "P6 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=20).value, "P6 (2, 6)")
        self.assertEqual(ws.cell(row=7, column=1).value, "I007")
        self.assertEqual(ws.cell(row=7, column=2).value, "a007")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=7).value, "P4 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=8).value, "P4 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=13).value, "P5 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=14).value, "P5 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=19).value, "P6 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=20).value, "P6 (3, 6)")
        self.assertEqual(ws.cell(row=8, column=1).value, "I008")
        self.assertEqual(ws.cell(row=8, column=2).value, "a008")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=7).value, "P4 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=8).value, "P4 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=13).value, "P5 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=14).value, "P5 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=19).value, "P6 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=20).value, "P6 (4, 6)")

    def test_save_rep_setup_instructions_inducer_cols_3(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer for plate columns
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=8, scale='log')
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='cols')
        # Create second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=8, scale='log')
        p.apply_inducer(atc, apply_to='cols')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "I001")
        self.assertEqual(ws.cell(row=1, column=2).value, "a001")
        self.assertEqual(ws.cell(row=1, column=3).value, "P1 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=4).value, "P1 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=5).value, "P1 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=6).value, "P1 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=7).value, "P1 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=8).value, "P1 (1, 6)")
        self.assertEqual(ws.cell(row=1, column=9).value, "P2 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=10).value, "P2 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=11).value, "P2 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=12).value, "P2 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=13).value, "P2 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=14).value, "P2 (1, 6)")
        self.assertEqual(ws.cell(row=1, column=15).value, "P3 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=16).value, "P3 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=17).value, "P3 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=18).value, "P3 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=19).value, "P3 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=20).value, "P3 (1, 6)")
        self.assertEqual(ws.cell(row=2, column=1).value, "I003")
        self.assertEqual(ws.cell(row=2, column=2).value, "a002")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=7).value, "P1 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=8).value, "P1 (2, 6)")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=13).value, "P2 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=14).value, "P2 (2, 6)")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=19).value, "P3 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=20).value, "P3 (2, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "I004")
        self.assertEqual(ws.cell(row=3, column=2).value, "a003")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=7).value, "P1 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=8).value, "P1 (3, 6)")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=13).value, "P2 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=14).value, "P2 (3, 6)")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=19).value, "P3 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=20).value, "P3 (3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "I007")
        self.assertEqual(ws.cell(row=4, column=2).value, "a004")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=7).value, "P1 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=8).value, "P1 (4, 6)")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=13).value, "P2 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=14).value, "P2 (4, 6)")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=19).value, "P3 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=20).value, "P3 (4, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "I008")
        self.assertEqual(ws.cell(row=5, column=2).value, "a005")
        self.assertEqual(ws.cell(row=5, column=3).value, "P4 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=4).value, "P4 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=5).value, "P4 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=6).value, "P4 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=7).value, "P4 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=8).value, "P4 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=9).value, "P5 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=10).value, "P5 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=11).value, "P5 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=12).value, "P5 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=13).value, "P5 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=14).value, "P5 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=15).value, "P6 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=16).value, "P6 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=17).value, "P6 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=18).value, "P6 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=19).value, "P6 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=20).value, "P6 (1, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "I005")
        self.assertEqual(ws.cell(row=6, column=2).value, "a006")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=7).value, "P4 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=8).value, "P4 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=13).value, "P5 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=14).value, "P5 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=19).value, "P6 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=20).value, "P6 (2, 6)")
        self.assertEqual(ws.cell(row=7, column=1).value, "I006")
        self.assertEqual(ws.cell(row=7, column=2).value, "a007")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=7).value, "P4 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=8).value, "P4 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=13).value, "P5 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=14).value, "P5 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=19).value, "P6 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=20).value, "P6 (3, 6)")
        self.assertEqual(ws.cell(row=8, column=1).value, "I002")
        self.assertEqual(ws.cell(row=8, column=2).value, "a008")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=7).value, "P4 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=8).value, "P4 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=13).value, "P5 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=14).value, "P5 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=19).value, "P6 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=20).value, "P6 (4, 6)")

    def test_save_rep_setup_instructions_inducer_wells_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=144, scale='log')
        p.apply_inducer(iptg, apply_to='wells')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "P1 (1, 1)\nI001")
        self.assertEqual(ws.cell(row=1, column=2).value, "P1 (1, 2)\nI002")
        self.assertEqual(ws.cell(row=1, column=3).value, "P1 (1, 3)\nI003")
        self.assertEqual(ws.cell(row=1, column=4).value, "P1 (1, 4)\nI004")
        self.assertEqual(ws.cell(row=1, column=5).value, "P1 (1, 5)\nI005")
        self.assertEqual(ws.cell(row=1, column=6).value, "P1 (1, 6)\nI006")
        self.assertEqual(ws.cell(row=1, column=7).value, "P2 (1, 1)\nI007")
        self.assertEqual(ws.cell(row=1, column=8).value, "P2 (1, 2)\nI008")
        self.assertEqual(ws.cell(row=1, column=9).value, "P2 (1, 3)\nI009")
        self.assertEqual(ws.cell(row=1, column=10).value, "P2 (1, 4)\nI010")
        self.assertEqual(ws.cell(row=1, column=11).value, "P2 (1, 5)\nI011")
        self.assertEqual(ws.cell(row=1, column=12).value, "P2 (1, 6)\nI012")
        self.assertEqual(ws.cell(row=1, column=13).value, "P3 (1, 1)\nI013")
        self.assertEqual(ws.cell(row=1, column=14).value, "P3 (1, 2)\nI014")
        self.assertEqual(ws.cell(row=1, column=15).value, "P3 (1, 3)\nI015")
        self.assertEqual(ws.cell(row=1, column=16).value, "P3 (1, 4)\nI016")
        self.assertEqual(ws.cell(row=1, column=17).value, "P3 (1, 5)\nI017")
        self.assertEqual(ws.cell(row=1, column=18).value, "P3 (1, 6)\nI018")
        self.assertEqual(ws.cell(row=2, column=1).value, "P1 (2, 1)\nI019")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (2, 2)\nI020")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (2, 3)\nI021")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (2, 4)\nI022")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (2, 5)\nI023")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (2, 6)\nI024")
        self.assertEqual(ws.cell(row=2, column=7).value, "P2 (2, 1)\nI025")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (2, 2)\nI026")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (2, 3)\nI027")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (2, 4)\nI028")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (2, 5)\nI029")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (2, 6)\nI030")
        self.assertEqual(ws.cell(row=2, column=13).value, "P3 (2, 1)\nI031")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (2, 2)\nI032")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (2, 3)\nI033")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (2, 4)\nI034")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (2, 5)\nI035")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (2, 6)\nI036")
        self.assertEqual(ws.cell(row=3, column=1).value, "P1 (3, 1)\nI037")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (3, 2)\nI038")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (3, 3)\nI039")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (3, 4)\nI040")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (3, 5)\nI041")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (3, 6)\nI042")
        self.assertEqual(ws.cell(row=3, column=7).value, "P2 (3, 1)\nI043")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (3, 2)\nI044")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (3, 3)\nI045")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (3, 4)\nI046")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (3, 5)\nI047")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (3, 6)\nI048")
        self.assertEqual(ws.cell(row=3, column=13).value, "P3 (3, 1)\nI049")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (3, 2)\nI050")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (3, 3)\nI051")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (3, 4)\nI052")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (3, 5)\nI053")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (3, 6)\nI054")
        self.assertEqual(ws.cell(row=4, column=1).value, "P1 (4, 1)\nI055")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (4, 2)\nI056")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (4, 3)\nI057")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (4, 4)\nI058")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (4, 5)\nI059")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (4, 6)\nI060")
        self.assertEqual(ws.cell(row=4, column=7).value, "P2 (4, 1)\nI061")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (4, 2)\nI062")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (4, 3)\nI063")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (4, 4)\nI064")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (4, 5)\nI065")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (4, 6)\nI066")
        self.assertEqual(ws.cell(row=4, column=13).value, "P3 (4, 1)\nI067")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (4, 2)\nI068")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (4, 3)\nI069")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (4, 4)\nI070")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (4, 5)\nI071")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (4, 6)\nI072")
        self.assertEqual(ws.cell(row=5, column=1).value, "P4 (1, 1)\nI073")
        self.assertEqual(ws.cell(row=5, column=2).value, "P4 (1, 2)\nI074")
        self.assertEqual(ws.cell(row=5, column=3).value, "P4 (1, 3)\nI075")
        self.assertEqual(ws.cell(row=5, column=4).value, "P4 (1, 4)\nI076")
        self.assertEqual(ws.cell(row=5, column=5).value, "P4 (1, 5)\nI077")
        self.assertEqual(ws.cell(row=5, column=6).value, "P4 (1, 6)\nI078")
        self.assertEqual(ws.cell(row=5, column=7).value, "P5 (1, 1)\nI079")
        self.assertEqual(ws.cell(row=5, column=8).value, "P5 (1, 2)\nI080")
        self.assertEqual(ws.cell(row=5, column=9).value, "P5 (1, 3)\nI081")
        self.assertEqual(ws.cell(row=5, column=10).value, "P5 (1, 4)\nI082")
        self.assertEqual(ws.cell(row=5, column=11).value, "P5 (1, 5)\nI083")
        self.assertEqual(ws.cell(row=5, column=12).value, "P5 (1, 6)\nI084")
        self.assertEqual(ws.cell(row=5, column=13).value, "P6 (1, 1)\nI085")
        self.assertEqual(ws.cell(row=5, column=14).value, "P6 (1, 2)\nI086")
        self.assertEqual(ws.cell(row=5, column=15).value, "P6 (1, 3)\nI087")
        self.assertEqual(ws.cell(row=5, column=16).value, "P6 (1, 4)\nI088")
        self.assertEqual(ws.cell(row=5, column=17).value, "P6 (1, 5)\nI089")
        self.assertEqual(ws.cell(row=5, column=18).value, "P6 (1, 6)\nI090")
        self.assertEqual(ws.cell(row=6, column=1).value, "P4 (2, 1)\nI091")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (2, 2)\nI092")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (2, 3)\nI093")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (2, 4)\nI094")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (2, 5)\nI095")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (2, 6)\nI096")
        self.assertEqual(ws.cell(row=6, column=7).value, "P5 (2, 1)\nI097")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (2, 2)\nI098")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (2, 3)\nI099")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (2, 4)\nI100")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (2, 5)\nI101")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (2, 6)\nI102")
        self.assertEqual(ws.cell(row=6, column=13).value, "P6 (2, 1)\nI103")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (2, 2)\nI104")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (2, 3)\nI105")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (2, 4)\nI106")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (2, 5)\nI107")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (2, 6)\nI108")
        self.assertEqual(ws.cell(row=7, column=1).value, "P4 (3, 1)\nI109")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (3, 2)\nI110")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (3, 3)\nI111")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (3, 4)\nI112")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (3, 5)\nI113")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (3, 6)\nI114")
        self.assertEqual(ws.cell(row=7, column=7).value, "P5 (3, 1)\nI115")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (3, 2)\nI116")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (3, 3)\nI117")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (3, 4)\nI118")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (3, 5)\nI119")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (3, 6)\nI120")
        self.assertEqual(ws.cell(row=7, column=13).value, "P6 (3, 1)\nI121")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (3, 2)\nI122")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (3, 3)\nI123")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (3, 4)\nI124")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (3, 5)\nI125")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (3, 6)\nI126")
        self.assertEqual(ws.cell(row=8, column=1).value, "P4 (4, 1)\nI127")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (4, 2)\nI128")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (4, 3)\nI129")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (4, 4)\nI130")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (4, 5)\nI131")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (4, 6)\nI132")
        self.assertEqual(ws.cell(row=8, column=7).value, "P5 (4, 1)\nI133")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (4, 2)\nI134")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (4, 3)\nI135")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (4, 4)\nI136")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (4, 5)\nI137")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (4, 6)\nI138")
        self.assertEqual(ws.cell(row=8, column=13).value, "P6 (4, 1)\nI139")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (4, 2)\nI140")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (4, 3)\nI141")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (4, 4)\nI142")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (4, 5)\nI143")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (4, 6)\nI144")

    def test_save_rep_setup_instructions_inducer_wells_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=144, scale='log')
        p.apply_inducer(iptg, apply_to='wells')
        # Create second inducer for plate
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=144, scale='log')
        p.apply_inducer(atc, apply_to='wells')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "P1 (1, 1)\nI001\na001")
        self.assertEqual(ws.cell(row=1, column=2).value, "P1 (1, 2)\nI002\na002")
        self.assertEqual(ws.cell(row=1, column=3).value, "P1 (1, 3)\nI003\na003")
        self.assertEqual(ws.cell(row=1, column=4).value, "P1 (1, 4)\nI004\na004")
        self.assertEqual(ws.cell(row=1, column=5).value, "P1 (1, 5)\nI005\na005")
        self.assertEqual(ws.cell(row=1, column=6).value, "P1 (1, 6)\nI006\na006")
        self.assertEqual(ws.cell(row=1, column=7).value, "P2 (1, 1)\nI007\na007")
        self.assertEqual(ws.cell(row=1, column=8).value, "P2 (1, 2)\nI008\na008")
        self.assertEqual(ws.cell(row=1, column=9).value, "P2 (1, 3)\nI009\na009")
        self.assertEqual(ws.cell(row=1, column=10).value, "P2 (1, 4)\nI010\na010")
        self.assertEqual(ws.cell(row=1, column=11).value, "P2 (1, 5)\nI011\na011")
        self.assertEqual(ws.cell(row=1, column=12).value, "P2 (1, 6)\nI012\na012")
        self.assertEqual(ws.cell(row=1, column=13).value, "P3 (1, 1)\nI013\na013")
        self.assertEqual(ws.cell(row=1, column=14).value, "P3 (1, 2)\nI014\na014")
        self.assertEqual(ws.cell(row=1, column=15).value, "P3 (1, 3)\nI015\na015")
        self.assertEqual(ws.cell(row=1, column=16).value, "P3 (1, 4)\nI016\na016")
        self.assertEqual(ws.cell(row=1, column=17).value, "P3 (1, 5)\nI017\na017")
        self.assertEqual(ws.cell(row=1, column=18).value, "P3 (1, 6)\nI018\na018")
        self.assertEqual(ws.cell(row=2, column=1).value, "P1 (2, 1)\nI019\na019")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (2, 2)\nI020\na020")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (2, 3)\nI021\na021")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (2, 4)\nI022\na022")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (2, 5)\nI023\na023")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (2, 6)\nI024\na024")
        self.assertEqual(ws.cell(row=2, column=7).value, "P2 (2, 1)\nI025\na025")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (2, 2)\nI026\na026")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (2, 3)\nI027\na027")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (2, 4)\nI028\na028")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (2, 5)\nI029\na029")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (2, 6)\nI030\na030")
        self.assertEqual(ws.cell(row=2, column=13).value, "P3 (2, 1)\nI031\na031")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (2, 2)\nI032\na032")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (2, 3)\nI033\na033")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (2, 4)\nI034\na034")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (2, 5)\nI035\na035")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (2, 6)\nI036\na036")
        self.assertEqual(ws.cell(row=3, column=1).value, "P1 (3, 1)\nI037\na037")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (3, 2)\nI038\na038")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (3, 3)\nI039\na039")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (3, 4)\nI040\na040")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (3, 5)\nI041\na041")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (3, 6)\nI042\na042")
        self.assertEqual(ws.cell(row=3, column=7).value, "P2 (3, 1)\nI043\na043")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (3, 2)\nI044\na044")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (3, 3)\nI045\na045")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (3, 4)\nI046\na046")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (3, 5)\nI047\na047")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (3, 6)\nI048\na048")
        self.assertEqual(ws.cell(row=3, column=13).value, "P3 (3, 1)\nI049\na049")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (3, 2)\nI050\na050")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (3, 3)\nI051\na051")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (3, 4)\nI052\na052")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (3, 5)\nI053\na053")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (3, 6)\nI054\na054")
        self.assertEqual(ws.cell(row=4, column=1).value, "P1 (4, 1)\nI055\na055")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (4, 2)\nI056\na056")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (4, 3)\nI057\na057")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (4, 4)\nI058\na058")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (4, 5)\nI059\na059")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (4, 6)\nI060\na060")
        self.assertEqual(ws.cell(row=4, column=7).value, "P2 (4, 1)\nI061\na061")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (4, 2)\nI062\na062")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (4, 3)\nI063\na063")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (4, 4)\nI064\na064")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (4, 5)\nI065\na065")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (4, 6)\nI066\na066")
        self.assertEqual(ws.cell(row=4, column=13).value, "P3 (4, 1)\nI067\na067")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (4, 2)\nI068\na068")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (4, 3)\nI069\na069")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (4, 4)\nI070\na070")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (4, 5)\nI071\na071")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (4, 6)\nI072\na072")
        self.assertEqual(ws.cell(row=5, column=1).value, "P4 (1, 1)\nI073\na073")
        self.assertEqual(ws.cell(row=5, column=2).value, "P4 (1, 2)\nI074\na074")
        self.assertEqual(ws.cell(row=5, column=3).value, "P4 (1, 3)\nI075\na075")
        self.assertEqual(ws.cell(row=5, column=4).value, "P4 (1, 4)\nI076\na076")
        self.assertEqual(ws.cell(row=5, column=5).value, "P4 (1, 5)\nI077\na077")
        self.assertEqual(ws.cell(row=5, column=6).value, "P4 (1, 6)\nI078\na078")
        self.assertEqual(ws.cell(row=5, column=7).value, "P5 (1, 1)\nI079\na079")
        self.assertEqual(ws.cell(row=5, column=8).value, "P5 (1, 2)\nI080\na080")
        self.assertEqual(ws.cell(row=5, column=9).value, "P5 (1, 3)\nI081\na081")
        self.assertEqual(ws.cell(row=5, column=10).value, "P5 (1, 4)\nI082\na082")
        self.assertEqual(ws.cell(row=5, column=11).value, "P5 (1, 5)\nI083\na083")
        self.assertEqual(ws.cell(row=5, column=12).value, "P5 (1, 6)\nI084\na084")
        self.assertEqual(ws.cell(row=5, column=13).value, "P6 (1, 1)\nI085\na085")
        self.assertEqual(ws.cell(row=5, column=14).value, "P6 (1, 2)\nI086\na086")
        self.assertEqual(ws.cell(row=5, column=15).value, "P6 (1, 3)\nI087\na087")
        self.assertEqual(ws.cell(row=5, column=16).value, "P6 (1, 4)\nI088\na088")
        self.assertEqual(ws.cell(row=5, column=17).value, "P6 (1, 5)\nI089\na089")
        self.assertEqual(ws.cell(row=5, column=18).value, "P6 (1, 6)\nI090\na090")
        self.assertEqual(ws.cell(row=6, column=1).value, "P4 (2, 1)\nI091\na091")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (2, 2)\nI092\na092")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (2, 3)\nI093\na093")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (2, 4)\nI094\na094")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (2, 5)\nI095\na095")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (2, 6)\nI096\na096")
        self.assertEqual(ws.cell(row=6, column=7).value, "P5 (2, 1)\nI097\na097")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (2, 2)\nI098\na098")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (2, 3)\nI099\na099")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (2, 4)\nI100\na100")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (2, 5)\nI101\na101")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (2, 6)\nI102\na102")
        self.assertEqual(ws.cell(row=6, column=13).value, "P6 (2, 1)\nI103\na103")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (2, 2)\nI104\na104")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (2, 3)\nI105\na105")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (2, 4)\nI106\na106")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (2, 5)\nI107\na107")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (2, 6)\nI108\na108")
        self.assertEqual(ws.cell(row=7, column=1).value, "P4 (3, 1)\nI109\na109")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (3, 2)\nI110\na110")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (3, 3)\nI111\na111")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (3, 4)\nI112\na112")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (3, 5)\nI113\na113")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (3, 6)\nI114\na114")
        self.assertEqual(ws.cell(row=7, column=7).value, "P5 (3, 1)\nI115\na115")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (3, 2)\nI116\na116")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (3, 3)\nI117\na117")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (3, 4)\nI118\na118")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (3, 5)\nI119\na119")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (3, 6)\nI120\na120")
        self.assertEqual(ws.cell(row=7, column=13).value, "P6 (3, 1)\nI121\na121")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (3, 2)\nI122\na122")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (3, 3)\nI123\na123")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (3, 4)\nI124\na124")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (3, 5)\nI125\na125")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (3, 6)\nI126\na126")
        self.assertEqual(ws.cell(row=8, column=1).value, "P4 (4, 1)\nI127\na127")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (4, 2)\nI128\na128")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (4, 3)\nI129\na129")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (4, 4)\nI130\na130")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (4, 5)\nI131\na131")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (4, 6)\nI132\na132")
        self.assertEqual(ws.cell(row=8, column=7).value, "P5 (4, 1)\nI133\na133")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (4, 2)\nI134\na134")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (4, 3)\nI135\na135")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (4, 4)\nI136\na136")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (4, 5)\nI137\na137")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (4, 6)\nI138\na138")
        self.assertEqual(ws.cell(row=8, column=13).value, "P6 (4, 1)\nI139\na139")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (4, 2)\nI140\na140")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (4, 3)\nI141\na141")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (4, 4)\nI142\na142")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (4, 5)\nI143\na143")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (4, 6)\nI144\na144")

    def test_save_rep_setup_instructions_inducer_wells_3(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Limit number of samples to measure
        p.samples_to_measure = 80
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=80, scale='log')
        p.apply_inducer(iptg, apply_to='wells')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "P1 (1, 1)\nI001")
        self.assertEqual(ws.cell(row=1, column=2).value, "P1 (1, 2)\nI002")
        self.assertEqual(ws.cell(row=1, column=3).value, "P1 (1, 3)\nI003")
        self.assertEqual(ws.cell(row=1, column=4).value, "P1 (1, 4)\nI004")
        self.assertEqual(ws.cell(row=1, column=5).value, "P1 (1, 5)\nI005")
        self.assertEqual(ws.cell(row=1, column=6).value, "P1 (1, 6)\nI006")
        self.assertEqual(ws.cell(row=1, column=7).value, "P2 (1, 1)\nI007")
        self.assertEqual(ws.cell(row=1, column=8).value, "P2 (1, 2)\nI008")
        self.assertEqual(ws.cell(row=1, column=9).value, "P2 (1, 3)\nI009")
        self.assertEqual(ws.cell(row=1, column=10).value, "P2 (1, 4)\nI010")
        self.assertEqual(ws.cell(row=1, column=11).value, "P2 (1, 5)\nI011")
        self.assertEqual(ws.cell(row=1, column=12).value, "P2 (1, 6)\nI012")
        self.assertEqual(ws.cell(row=1, column=13).value, "P3 (1, 1)\nI013")
        self.assertEqual(ws.cell(row=1, column=14).value, "P3 (1, 2)\nI014")
        self.assertEqual(ws.cell(row=1, column=15).value, "P3 (1, 3)\nI015")
        self.assertEqual(ws.cell(row=1, column=16).value, "P3 (1, 4)\nI016")
        self.assertEqual(ws.cell(row=1, column=17).value, "P3 (1, 5)\nI017")
        self.assertEqual(ws.cell(row=1, column=18).value, "P3 (1, 6)\nI018")
        self.assertEqual(ws.cell(row=2, column=1).value, "P1 (2, 1)\nI019")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (2, 2)\nI020")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (2, 3)\nI021")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (2, 4)\nI022")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (2, 5)\nI023")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (2, 6)\nI024")
        self.assertEqual(ws.cell(row=2, column=7).value, "P2 (2, 1)\nI025")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (2, 2)\nI026")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (2, 3)\nI027")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (2, 4)\nI028")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (2, 5)\nI029")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (2, 6)\nI030")
        self.assertEqual(ws.cell(row=2, column=13).value, "P3 (2, 1)\nI031")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (2, 2)\nI032")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (2, 3)\nI033")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (2, 4)\nI034")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (2, 5)\nI035")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (2, 6)\nI036")
        self.assertEqual(ws.cell(row=3, column=1).value, "P1 (3, 1)\nI037")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (3, 2)\nI038")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (3, 3)\nI039")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (3, 4)\nI040")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (3, 5)\nI041")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (3, 6)\nI042")
        self.assertEqual(ws.cell(row=3, column=7).value, "P2 (3, 1)\nI043")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (3, 2)\nI044")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (3, 3)\nI045")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (3, 4)\nI046")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (3, 5)\nI047")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (3, 6)\nI048")
        self.assertEqual(ws.cell(row=3, column=13).value, "P3 (3, 1)\nI049")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (3, 2)\nI050")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (3, 3)\nI051")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (3, 4)\nI052")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (3, 5)\nI053")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (3, 6)\nI054")
        self.assertEqual(ws.cell(row=4, column=1).value, "P1 (4, 1)\nI055")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (4, 2)\nI056")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (4, 3)\nI057")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (4, 4)\nI058")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (4, 5)\nI059")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (4, 6)\nI060")
        self.assertEqual(ws.cell(row=4, column=7).value, "P2 (4, 1)\nI061")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (4, 2)\nI062")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (4, 3)\nI063")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (4, 4)\nI064")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (4, 5)\nI065")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (4, 6)\nI066")
        self.assertEqual(ws.cell(row=4, column=13).value, "P3 (4, 1)\nI067")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (4, 2)\nI068")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (4, 3)\nI069")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (4, 4)\nI070")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (4, 5)\nI071")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (4, 6)\nI072")
        self.assertEqual(ws.cell(row=5, column=1).value, "P4 (1, 1)\nI073")
        self.assertEqual(ws.cell(row=5, column=2).value, "P4 (1, 2)\nI074")
        self.assertEqual(ws.cell(row=5, column=3).value, "P4 (1, 3)\nI075")
        self.assertEqual(ws.cell(row=5, column=4).value, "P4 (1, 4)\nI076")
        self.assertEqual(ws.cell(row=5, column=5).value, "P4 (1, 5)\nI077")
        self.assertEqual(ws.cell(row=5, column=6).value, "P4 (1, 6)\nI078")
        self.assertEqual(ws.cell(row=5, column=7).value, "P5 (1, 1)\nI079")
        self.assertEqual(ws.cell(row=5, column=8).value, "P5 (1, 2)\nI080")
        self.assertEqual(ws.cell(row=5, column=9).value, "P5 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=10).value, "P5 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=11).value, "P5 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=12).value, "P5 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=13).value, "P6 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=14).value, "P6 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=15).value, "P6 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=16).value, "P6 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=17).value, "P6 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=18).value, "P6 (1, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "P4 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=7).value, "P5 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=13).value, "P6 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (2, 6)")
        self.assertEqual(ws.cell(row=7, column=1).value, "P4 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=7).value, "P5 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=13).value, "P6 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (3, 6)")
        self.assertEqual(ws.cell(row=8, column=1).value, "P4 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=7).value, "P5 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=13).value, "P6 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (4, 6)")

    def test_save_rep_setup_instructions_inducer_wells_4(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.set_gradient(min=1e-6, max=1e-3, n=144, scale='log')
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='wells')
        # Create second inducer for plate
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.set_gradient(min=0.5, max=50, n=144, scale='log')
        p.apply_inducer(atc, apply_to='wells')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "P1 (1, 1)\nI018\na001")
        self.assertEqual(ws.cell(row=1, column=2).value, "P1 (1, 2)\nI116\na002")
        self.assertEqual(ws.cell(row=1, column=3).value, "P1 (1, 3)\nI081\na003")
        self.assertEqual(ws.cell(row=1, column=4).value, "P1 (1, 4)\nI049\na004")
        self.assertEqual(ws.cell(row=1, column=5).value, "P1 (1, 5)\nI062\na005")
        self.assertEqual(ws.cell(row=1, column=6).value, "P1 (1, 6)\nI059\na006")
        self.assertEqual(ws.cell(row=1, column=7).value, "P2 (1, 1)\nI072\na007")
        self.assertEqual(ws.cell(row=1, column=8).value, "P2 (1, 2)\nI011\na008")
        self.assertEqual(ws.cell(row=1, column=9).value, "P2 (1, 3)\nI044\na009")
        self.assertEqual(ws.cell(row=1, column=10).value, "P2 (1, 4)\nI123\na010")
        self.assertEqual(ws.cell(row=1, column=11).value, "P2 (1, 5)\nI075\na011")
        self.assertEqual(ws.cell(row=1, column=12).value, "P2 (1, 6)\nI033\na012")
        self.assertEqual(ws.cell(row=1, column=13).value, "P3 (1, 1)\nI048\na013")
        self.assertEqual(ws.cell(row=1, column=14).value, "P3 (1, 2)\nI129\na014")
        self.assertEqual(ws.cell(row=1, column=15).value, "P3 (1, 3)\nI010\na015")
        self.assertEqual(ws.cell(row=1, column=16).value, "P3 (1, 4)\nI028\na016")
        self.assertEqual(ws.cell(row=1, column=17).value, "P3 (1, 5)\nI005\na017")
        self.assertEqual(ws.cell(row=1, column=18).value, "P3 (1, 6)\nI043\na018")
        self.assertEqual(ws.cell(row=2, column=1).value, "P1 (2, 1)\nI023\na019")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (2, 2)\nI110\na020")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (2, 3)\nI002\na021")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (2, 4)\nI061\na022")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (2, 5)\nI085\na023")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (2, 6)\nI007\na024")
        self.assertEqual(ws.cell(row=2, column=7).value, "P2 (2, 1)\nI083\na025")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (2, 2)\nI098\na026")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (2, 3)\nI079\na027")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (2, 4)\nI127\na028")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (2, 5)\nI017\na029")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (2, 6)\nI006\na030")
        self.assertEqual(ws.cell(row=2, column=13).value, "P3 (2, 1)\nI008\na031")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (2, 2)\nI103\na032")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (2, 3)\nI056\na033")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (2, 4)\nI076\na034")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (2, 5)\nI087\na035")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (2, 6)\nI065\na036")
        self.assertEqual(ws.cell(row=3, column=1).value, "P1 (3, 1)\nI102\na037")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (3, 2)\nI014\na038")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (3, 3)\nI019\na039")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (3, 4)\nI021\na040")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (3, 5)\nI009\na041")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (3, 6)\nI024\na042")
        self.assertEqual(ws.cell(row=3, column=7).value, "P2 (3, 1)\nI038\na043")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (3, 2)\nI100\na044")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (3, 3)\nI105\na045")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (3, 4)\nI117\na046")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (3, 5)\nI134\na047")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (3, 6)\nI039\na048")
        self.assertEqual(ws.cell(row=3, column=13).value, "P3 (3, 1)\nI074\na049")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (3, 2)\nI045\na050")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (3, 3)\nI042\na051")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (3, 4)\nI140\na052")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (3, 5)\nI131\na053")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (3, 6)\nI111\na054")
        self.assertEqual(ws.cell(row=4, column=1).value, "P1 (4, 1)\nI077\na055")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (4, 2)\nI016\na056")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (4, 3)\nI068\na057")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (4, 4)\nI139\na058")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (4, 5)\nI057\na059")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (4, 6)\nI031\na060")
        self.assertEqual(ws.cell(row=4, column=7).value, "P2 (4, 1)\nI084\na061")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (4, 2)\nI054\na062")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (4, 3)\nI066\na063")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (4, 4)\nI099\na064")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (4, 5)\nI095\na065")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (4, 6)\nI132\na066")
        self.assertEqual(ws.cell(row=4, column=13).value, "P3 (4, 1)\nI012\na067")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (4, 2)\nI113\na068")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (4, 3)\nI097\na069")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (4, 4)\nI086\na070")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (4, 5)\nI050\na071")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (4, 6)\nI088\na072")
        self.assertEqual(ws.cell(row=5, column=1).value, "P4 (1, 1)\nI137\na073")
        self.assertEqual(ws.cell(row=5, column=2).value, "P4 (1, 2)\nI037\na074")
        self.assertEqual(ws.cell(row=5, column=3).value, "P4 (1, 3)\nI096\na075")
        self.assertEqual(ws.cell(row=5, column=4).value, "P4 (1, 4)\nI108\na076")
        self.assertEqual(ws.cell(row=5, column=5).value, "P4 (1, 5)\nI107\na077")
        self.assertEqual(ws.cell(row=5, column=6).value, "P4 (1, 6)\nI040\na078")
        self.assertEqual(ws.cell(row=5, column=7).value, "P5 (1, 1)\nI035\na079")
        self.assertEqual(ws.cell(row=5, column=8).value, "P5 (1, 2)\nI128\na080")
        self.assertEqual(ws.cell(row=5, column=9).value, "P5 (1, 3)\nI093\na081")
        self.assertEqual(ws.cell(row=5, column=10).value, "P5 (1, 4)\nI143\na082")
        self.assertEqual(ws.cell(row=5, column=11).value, "P5 (1, 5)\nI138\na083")
        self.assertEqual(ws.cell(row=5, column=12).value, "P5 (1, 6)\nI015\na084")
        self.assertEqual(ws.cell(row=5, column=13).value, "P6 (1, 1)\nI141\na085")
        self.assertEqual(ws.cell(row=5, column=14).value, "P6 (1, 2)\nI106\na086")
        self.assertEqual(ws.cell(row=5, column=15).value, "P6 (1, 3)\nI022\na087")
        self.assertEqual(ws.cell(row=5, column=16).value, "P6 (1, 4)\nI118\na088")
        self.assertEqual(ws.cell(row=5, column=17).value, "P6 (1, 5)\nI053\na089")
        self.assertEqual(ws.cell(row=5, column=18).value, "P6 (1, 6)\nI046\na090")
        self.assertEqual(ws.cell(row=6, column=1).value, "P4 (2, 1)\nI078\na091")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (2, 2)\nI082\na092")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (2, 3)\nI055\na093")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (2, 4)\nI029\na094")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (2, 5)\nI064\na095")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (2, 6)\nI080\na096")
        self.assertEqual(ws.cell(row=6, column=7).value, "P5 (2, 1)\nI041\na097")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (2, 2)\nI092\na098")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (2, 3)\nI071\na099")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (2, 4)\nI073\na100")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (2, 5)\nI034\na101")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (2, 6)\nI136\na102")
        self.assertEqual(ws.cell(row=6, column=13).value, "P6 (2, 1)\nI089\na103")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (2, 2)\nI104\na104")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (2, 3)\nI144\na105")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (2, 4)\nI069\na106")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (2, 5)\nI060\na107")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (2, 6)\nI091\na108")
        self.assertEqual(ws.cell(row=7, column=1).value, "P4 (3, 1)\nI003\na109")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (3, 2)\nI032\na110")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (3, 3)\nI052\na111")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (3, 4)\nI025\na112")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (3, 5)\nI126\na113")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (3, 6)\nI027\na114")
        self.assertEqual(ws.cell(row=7, column=7).value, "P5 (3, 1)\nI130\na115")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (3, 2)\nI119\na116")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (3, 3)\nI120\na117")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (3, 4)\nI124\na118")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (3, 5)\nI051\na119")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (3, 6)\nI026\na120")
        self.assertEqual(ws.cell(row=7, column=13).value, "P6 (3, 1)\nI047\na121")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (3, 2)\nI115\na122")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (3, 3)\nI067\na123")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (3, 4)\nI125\na124")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (3, 5)\nI135\na125")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (3, 6)\nI114\na126")
        self.assertEqual(ws.cell(row=8, column=1).value, "P4 (4, 1)\nI121\na127")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (4, 2)\nI030\na128")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (4, 3)\nI094\na129")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (4, 4)\nI133\na130")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (4, 5)\nI001\na131")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (4, 6)\nI101\na132")
        self.assertEqual(ws.cell(row=8, column=7).value, "P5 (4, 1)\nI058\na133")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (4, 2)\nI112\na134")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (4, 3)\nI004\na135")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (4, 4)\nI013\na136")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (4, 5)\nI142\na137")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (4, 6)\nI090\na138")
        self.assertEqual(ws.cell(row=8, column=13).value, "P6 (4, 1)\nI063\na139")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (4, 2)\nI070\na140")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (4, 3)\nI036\na141")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (4, 4)\nI109\na142")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (4, 5)\nI122\na143")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (4, 6)\nI020\na144")

    def test_save_rep_setup_instructions_inducer_media_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Limit number of samples to measure
        p.samples_to_measure = 80
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.shot_vol = 5.
        iptg.concentrations = [10]
        p.apply_inducer(iptg, apply_to='media')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "P1 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=2).value, "P1 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=3).value, "P1 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=4).value, "P1 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=5).value, "P1 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=6).value, "P1 (1, 6)")
        self.assertEqual(ws.cell(row=1, column=7).value, "P2 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=8).value, "P2 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=9).value, "P2 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=10).value, "P2 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=11).value, "P2 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=12).value, "P2 (1, 6)")
        self.assertEqual(ws.cell(row=1, column=13).value, "P3 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=14).value, "P3 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=15).value, "P3 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=16).value, "P3 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=17).value, "P3 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=18).value, "P3 (1, 6)")
        self.assertEqual(ws.cell(row=2, column=1).value, "P1 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (2, 6)")
        self.assertEqual(ws.cell(row=2, column=7).value, "P2 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (2, 6)")
        self.assertEqual(ws.cell(row=2, column=13).value, "P3 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (2, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "P1 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (3, 6)")
        self.assertEqual(ws.cell(row=3, column=7).value, "P2 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (3, 6)")
        self.assertEqual(ws.cell(row=3, column=13).value, "P3 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "P1 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (4, 6)")
        self.assertEqual(ws.cell(row=4, column=7).value, "P2 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (4, 6)")
        self.assertEqual(ws.cell(row=4, column=13).value, "P3 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (4, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "P4 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=2).value, "P4 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=3).value, "P4 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=4).value, "P4 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=5).value, "P4 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=6).value, "P4 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=7).value, "P5 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=8).value, "P5 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=9).value, "P5 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=10).value, "P5 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=11).value, "P5 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=12).value, "P5 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=13).value, "P6 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=14).value, "P6 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=15).value, "P6 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=16).value, "P6 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=17).value, "P6 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=18).value, "P6 (1, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "P4 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=7).value, "P5 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=13).value, "P6 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (2, 6)")
        self.assertEqual(ws.cell(row=7, column=1).value, "P4 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=7).value, "P5 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=13).value, "P6 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (3, 6)")
        self.assertEqual(ws.cell(row=8, column=1).value, "P4 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=7).value, "P5 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=13).value, "P6 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (4, 6)")
        self.assertEqual(ws.cell(row=10, column=1).value,
                         u"Add 5.00µL of IPTG to media")

    def test_save_rep_setup_instructions_inducer_media_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Limit number of samples to measure
        p.samples_to_measure = 80
        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.shot_vol = 5.
        iptg.concentrations = [10]
        p.apply_inducer(iptg, apply_to='media')
        # Create second inducer for plate
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.shot_vol = 10.
        atc.concentrations = [12]
        p.apply_inducer(atc, apply_to='media')
        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "P1 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=2).value, "P1 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=3).value, "P1 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=4).value, "P1 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=5).value, "P1 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=6).value, "P1 (1, 6)")
        self.assertEqual(ws.cell(row=1, column=7).value, "P2 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=8).value, "P2 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=9).value, "P2 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=10).value, "P2 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=11).value, "P2 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=12).value, "P2 (1, 6)")
        self.assertEqual(ws.cell(row=1, column=13).value, "P3 (1, 1)")
        self.assertEqual(ws.cell(row=1, column=14).value, "P3 (1, 2)")
        self.assertEqual(ws.cell(row=1, column=15).value, "P3 (1, 3)")
        self.assertEqual(ws.cell(row=1, column=16).value, "P3 (1, 4)")
        self.assertEqual(ws.cell(row=1, column=17).value, "P3 (1, 5)")
        self.assertEqual(ws.cell(row=1, column=18).value, "P3 (1, 6)")
        self.assertEqual(ws.cell(row=2, column=1).value, "P1 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (2, 6)")
        self.assertEqual(ws.cell(row=2, column=7).value, "P2 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (2, 6)")
        self.assertEqual(ws.cell(row=2, column=13).value, "P3 (2, 1)")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (2, 2)")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (2, 3)")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (2, 4)")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (2, 5)")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (2, 6)")
        self.assertEqual(ws.cell(row=3, column=1).value, "P1 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (3, 6)")
        self.assertEqual(ws.cell(row=3, column=7).value, "P2 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (3, 6)")
        self.assertEqual(ws.cell(row=3, column=13).value, "P3 (3, 1)")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (3, 2)")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (3, 3)")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (3, 4)")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (3, 5)")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (3, 6)")
        self.assertEqual(ws.cell(row=4, column=1).value, "P1 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (4, 6)")
        self.assertEqual(ws.cell(row=4, column=7).value, "P2 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (4, 6)")
        self.assertEqual(ws.cell(row=4, column=13).value, "P3 (4, 1)")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (4, 2)")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (4, 3)")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (4, 4)")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (4, 5)")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (4, 6)")
        self.assertEqual(ws.cell(row=5, column=1).value, "P4 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=2).value, "P4 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=3).value, "P4 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=4).value, "P4 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=5).value, "P4 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=6).value, "P4 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=7).value, "P5 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=8).value, "P5 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=9).value, "P5 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=10).value, "P5 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=11).value, "P5 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=12).value, "P5 (1, 6)")
        self.assertEqual(ws.cell(row=5, column=13).value, "P6 (1, 1)")
        self.assertEqual(ws.cell(row=5, column=14).value, "P6 (1, 2)")
        self.assertEqual(ws.cell(row=5, column=15).value, "P6 (1, 3)")
        self.assertEqual(ws.cell(row=5, column=16).value, "P6 (1, 4)")
        self.assertEqual(ws.cell(row=5, column=17).value, "P6 (1, 5)")
        self.assertEqual(ws.cell(row=5, column=18).value, "P6 (1, 6)")
        self.assertEqual(ws.cell(row=6, column=1).value, "P4 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=7).value, "P5 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (2, 6)")
        self.assertEqual(ws.cell(row=6, column=13).value, "P6 (2, 1)")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (2, 2)")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (2, 3)")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (2, 4)")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (2, 5)")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (2, 6)")
        self.assertEqual(ws.cell(row=7, column=1).value, "P4 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=7).value, "P5 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (3, 6)")
        self.assertEqual(ws.cell(row=7, column=13).value, "P6 (3, 1)")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (3, 2)")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (3, 3)")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (3, 4)")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (3, 5)")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (3, 6)")
        self.assertEqual(ws.cell(row=8, column=1).value, "P4 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=7).value, "P5 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (4, 6)")
        self.assertEqual(ws.cell(row=8, column=13).value, "P6 (4, 1)")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (4, 2)")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (4, 3)")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (4, 4)")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (4, 5)")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (4, 6)")
        self.assertEqual(ws.cell(row=10, column=1).value,
                         u"Add 5.00µL of IPTG to media")
        self.assertEqual(ws.cell(row=11, column=1).value,
                         u"Add 10.00µL of aTc to media")

    def test_save_rep_setup_instructions_inducer_mixed(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.shot_vol = 5.
        iptg.set_gradient(min=1e-6, max=1e-3, n=18, scale='log')
        p.apply_inducer(iptg, apply_to='rows')

        # Create second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.shot_vol = 5.
        atc.set_gradient(min=0.5, max=50, n=8, scale='log')
        p.apply_inducer(atc, apply_to='cols')

        # Create inducer for plate wells
        xyl = platedesign.inducer.ChemicalInducer(
            name='Xylose',
            units=u'%')
        xyl.set_gradient(min=1e-6, max=1e-3, n=144, scale='log')
        p.apply_inducer(xyl, apply_to='wells')

        # Create inducer for plate
        sugar = platedesign.inducer.ChemicalInducer(
            name='Sugar',
            units=u'ng/µL')
        sugar.shot_vol = 10.
        sugar.concentrations = [2]
        p.apply_inducer(sugar, apply_to='media')

        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Spreadsheet should contain one sheet
        self.assertEqual(wb.sheetnames, ["Inducers for Plate Array A1"])
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, None)
        self.assertEqual(ws.cell(row=1, column=2).value, "I001")
        self.assertEqual(ws.cell(row=1, column=3).value, "I002")
        self.assertEqual(ws.cell(row=1, column=4).value, "I003")
        self.assertEqual(ws.cell(row=1, column=5).value, "I004")
        self.assertEqual(ws.cell(row=1, column=6).value, "I005")
        self.assertEqual(ws.cell(row=1, column=7).value, "I006")
        self.assertEqual(ws.cell(row=1, column=8).value, "I007")
        self.assertEqual(ws.cell(row=1, column=9).value, "I008")
        self.assertEqual(ws.cell(row=1, column=10).value, "I009")
        self.assertEqual(ws.cell(row=1, column=11).value, "I010")
        self.assertEqual(ws.cell(row=1, column=12).value, "I011")
        self.assertEqual(ws.cell(row=1, column=13).value, "I012")
        self.assertEqual(ws.cell(row=1, column=14).value, "I013")
        self.assertEqual(ws.cell(row=1, column=15).value, "I014")
        self.assertEqual(ws.cell(row=1, column=16).value, "I015")
        self.assertEqual(ws.cell(row=1, column=17).value, "I016")
        self.assertEqual(ws.cell(row=1, column=18).value, "I017")
        self.assertEqual(ws.cell(row=1, column=19).value, "I018")
        self.assertEqual(ws.cell(row=2, column=1).value, "a001")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (1, 1)\nX001")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (1, 2)\nX002")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (1, 3)\nX003")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (1, 4)\nX004")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (1, 5)\nX005")
        self.assertEqual(ws.cell(row=2, column=7).value, "P1 (1, 6)\nX006")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (1, 1)\nX007")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (1, 2)\nX008")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (1, 3)\nX009")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (1, 4)\nX010")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (1, 5)\nX011")
        self.assertEqual(ws.cell(row=2, column=13).value, "P2 (1, 6)\nX012")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (1, 1)\nX013")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (1, 2)\nX014")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (1, 3)\nX015")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (1, 4)\nX016")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (1, 5)\nX017")
        self.assertEqual(ws.cell(row=2, column=19).value, "P3 (1, 6)\nX018")
        self.assertEqual(ws.cell(row=3, column=1).value, "a002")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (2, 1)\nX019")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (2, 2)\nX020")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (2, 3)\nX021")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (2, 4)\nX022")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (2, 5)\nX023")
        self.assertEqual(ws.cell(row=3, column=7).value, "P1 (2, 6)\nX024")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (2, 1)\nX025")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (2, 2)\nX026")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (2, 3)\nX027")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (2, 4)\nX028")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (2, 5)\nX029")
        self.assertEqual(ws.cell(row=3, column=13).value, "P2 (2, 6)\nX030")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (2, 1)\nX031")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (2, 2)\nX032")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (2, 3)\nX033")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (2, 4)\nX034")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (2, 5)\nX035")
        self.assertEqual(ws.cell(row=3, column=19).value, "P3 (2, 6)\nX036")
        self.assertEqual(ws.cell(row=4, column=1).value, "a003")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (3, 1)\nX037")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (3, 2)\nX038")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (3, 3)\nX039")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (3, 4)\nX040")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (3, 5)\nX041")
        self.assertEqual(ws.cell(row=4, column=7).value, "P1 (3, 6)\nX042")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (3, 1)\nX043")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (3, 2)\nX044")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (3, 3)\nX045")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (3, 4)\nX046")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (3, 5)\nX047")
        self.assertEqual(ws.cell(row=4, column=13).value, "P2 (3, 6)\nX048")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (3, 1)\nX049")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (3, 2)\nX050")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (3, 3)\nX051")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (3, 4)\nX052")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (3, 5)\nX053")
        self.assertEqual(ws.cell(row=4, column=19).value, "P3 (3, 6)\nX054")
        self.assertEqual(ws.cell(row=5, column=1).value, "a004")
        self.assertEqual(ws.cell(row=5, column=2).value, "P1 (4, 1)\nX055")
        self.assertEqual(ws.cell(row=5, column=3).value, "P1 (4, 2)\nX056")
        self.assertEqual(ws.cell(row=5, column=4).value, "P1 (4, 3)\nX057")
        self.assertEqual(ws.cell(row=5, column=5).value, "P1 (4, 4)\nX058")
        self.assertEqual(ws.cell(row=5, column=6).value, "P1 (4, 5)\nX059")
        self.assertEqual(ws.cell(row=5, column=7).value, "P1 (4, 6)\nX060")
        self.assertEqual(ws.cell(row=5, column=8).value, "P2 (4, 1)\nX061")
        self.assertEqual(ws.cell(row=5, column=9).value, "P2 (4, 2)\nX062")
        self.assertEqual(ws.cell(row=5, column=10).value, "P2 (4, 3)\nX063")
        self.assertEqual(ws.cell(row=5, column=11).value, "P2 (4, 4)\nX064")
        self.assertEqual(ws.cell(row=5, column=12).value, "P2 (4, 5)\nX065")
        self.assertEqual(ws.cell(row=5, column=13).value, "P2 (4, 6)\nX066")
        self.assertEqual(ws.cell(row=5, column=14).value, "P3 (4, 1)\nX067")
        self.assertEqual(ws.cell(row=5, column=15).value, "P3 (4, 2)\nX068")
        self.assertEqual(ws.cell(row=5, column=16).value, "P3 (4, 3)\nX069")
        self.assertEqual(ws.cell(row=5, column=17).value, "P3 (4, 4)\nX070")
        self.assertEqual(ws.cell(row=5, column=18).value, "P3 (4, 5)\nX071")
        self.assertEqual(ws.cell(row=5, column=19).value, "P3 (4, 6)\nX072")
        self.assertEqual(ws.cell(row=6, column=1).value, "a005")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (1, 1)\nX073")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (1, 2)\nX074")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (1, 3)\nX075")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (1, 4)\nX076")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (1, 5)\nX077")
        self.assertEqual(ws.cell(row=6, column=7).value, "P4 (1, 6)\nX078")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (1, 1)\nX079")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (1, 2)\nX080")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (1, 3)\nX081")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (1, 4)\nX082")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (1, 5)\nX083")
        self.assertEqual(ws.cell(row=6, column=13).value, "P5 (1, 6)\nX084")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (1, 1)\nX085")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (1, 2)\nX086")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (1, 3)\nX087")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (1, 4)\nX088")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (1, 5)\nX089")
        self.assertEqual(ws.cell(row=6, column=19).value, "P6 (1, 6)\nX090")
        self.assertEqual(ws.cell(row=7, column=1).value, "a006")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (2, 1)\nX091")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (2, 2)\nX092")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (2, 3)\nX093")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (2, 4)\nX094")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (2, 5)\nX095")
        self.assertEqual(ws.cell(row=7, column=7).value, "P4 (2, 6)\nX096")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (2, 1)\nX097")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (2, 2)\nX098")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (2, 3)\nX099")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (2, 4)\nX100")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (2, 5)\nX101")
        self.assertEqual(ws.cell(row=7, column=13).value, "P5 (2, 6)\nX102")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (2, 1)\nX103")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (2, 2)\nX104")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (2, 3)\nX105")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (2, 4)\nX106")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (2, 5)\nX107")
        self.assertEqual(ws.cell(row=7, column=19).value, "P6 (2, 6)\nX108")
        self.assertEqual(ws.cell(row=8, column=1).value, "a007")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (3, 1)\nX109")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (3, 2)\nX110")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (3, 3)\nX111")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (3, 4)\nX112")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (3, 5)\nX113")
        self.assertEqual(ws.cell(row=8, column=7).value, "P4 (3, 6)\nX114")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (3, 1)\nX115")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (3, 2)\nX116")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (3, 3)\nX117")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (3, 4)\nX118")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (3, 5)\nX119")
        self.assertEqual(ws.cell(row=8, column=13).value, "P5 (3, 6)\nX120")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (3, 1)\nX121")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (3, 2)\nX122")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (3, 3)\nX123")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (3, 4)\nX124")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (3, 5)\nX125")
        self.assertEqual(ws.cell(row=8, column=19).value, "P6 (3, 6)\nX126")
        self.assertEqual(ws.cell(row=9, column=1).value, "a008")
        self.assertEqual(ws.cell(row=9, column=2).value, "P4 (4, 1)\nX127")
        self.assertEqual(ws.cell(row=9, column=3).value, "P4 (4, 2)\nX128")
        self.assertEqual(ws.cell(row=9, column=4).value, "P4 (4, 3)\nX129")
        self.assertEqual(ws.cell(row=9, column=5).value, "P4 (4, 4)\nX130")
        self.assertEqual(ws.cell(row=9, column=6).value, "P4 (4, 5)\nX131")
        self.assertEqual(ws.cell(row=9, column=7).value, "P4 (4, 6)\nX132")
        self.assertEqual(ws.cell(row=9, column=8).value, "P5 (4, 1)\nX133")
        self.assertEqual(ws.cell(row=9, column=9).value, "P5 (4, 2)\nX134")
        self.assertEqual(ws.cell(row=9, column=10).value, "P5 (4, 3)\nX135")
        self.assertEqual(ws.cell(row=9, column=11).value, "P5 (4, 4)\nX136")
        self.assertEqual(ws.cell(row=9, column=12).value, "P5 (4, 5)\nX137")
        self.assertEqual(ws.cell(row=9, column=13).value, "P5 (4, 6)\nX138")
        self.assertEqual(ws.cell(row=9, column=14).value, "P6 (4, 1)\nX139")
        self.assertEqual(ws.cell(row=9, column=15).value, "P6 (4, 2)\nX140")
        self.assertEqual(ws.cell(row=9, column=16).value, "P6 (4, 3)\nX141")
        self.assertEqual(ws.cell(row=9, column=17).value, "P6 (4, 4)\nX142")
        self.assertEqual(ws.cell(row=9, column=18).value, "P6 (4, 5)\nX143")
        self.assertEqual(ws.cell(row=9, column=19).value, "P6 (4, 6)\nX144")
        self.assertEqual(ws.cell(row=11, column=1).value,
                         u"Add 10.00µL of Sugar to media")

    def test_save_rep_setup_instructions_cells_and_inducer(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.total_media_vol = 80000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_initial_od600 = 1e-5

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.shot_vol = 5.
        iptg.set_gradient(min=1e-6, max=1e-3, n=18, scale='log')
        p.apply_inducer(iptg, apply_to='rows')

        # Create second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.shot_vol = 5.
        atc.set_gradient(min=0.5, max=50, n=8, scale='log')
        p.apply_inducer(atc, apply_to='cols')

        # Create inducer for plate wells
        xyl = platedesign.inducer.ChemicalInducer(
            name='Xylose',
            units=u'%')
        xyl.set_gradient(min=1e-6, max=1e-3, n=144, scale='log')
        p.apply_inducer(xyl, apply_to='wells')

        # Create inducer for plate
        sugar = platedesign.inducer.ChemicalInducer(
            name='Sugar',
            units=u'ng/µL')
        sugar.shot_vol = 10.
        sugar.concentrations = [2]
        p.apply_inducer(sugar, apply_to='media')

        # Run save_rep_setup_instructions
        p.save_rep_setup_instructions(file_name=os.path.join(self.temp_dir,
                                                             'plate_rep.xlsx'))
        # Load spreadsheet
        wb = openpyxl.load_workbook(filename=os.path.join(self.temp_dir,
                                                          'plate_rep.xlsx'))
        # Check that sheet exists in spreadsheet
        self.assertTrue("Inducers for Plate Array A1" in wb.sheetnames)
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, None)
        self.assertEqual(ws.cell(row=1, column=2).value, "I001")
        self.assertEqual(ws.cell(row=1, column=3).value, "I002")
        self.assertEqual(ws.cell(row=1, column=4).value, "I003")
        self.assertEqual(ws.cell(row=1, column=5).value, "I004")
        self.assertEqual(ws.cell(row=1, column=6).value, "I005")
        self.assertEqual(ws.cell(row=1, column=7).value, "I006")
        self.assertEqual(ws.cell(row=1, column=8).value, "I007")
        self.assertEqual(ws.cell(row=1, column=9).value, "I008")
        self.assertEqual(ws.cell(row=1, column=10).value, "I009")
        self.assertEqual(ws.cell(row=1, column=11).value, "I010")
        self.assertEqual(ws.cell(row=1, column=12).value, "I011")
        self.assertEqual(ws.cell(row=1, column=13).value, "I012")
        self.assertEqual(ws.cell(row=1, column=14).value, "I013")
        self.assertEqual(ws.cell(row=1, column=15).value, "I014")
        self.assertEqual(ws.cell(row=1, column=16).value, "I015")
        self.assertEqual(ws.cell(row=1, column=17).value, "I016")
        self.assertEqual(ws.cell(row=1, column=18).value, "I017")
        self.assertEqual(ws.cell(row=1, column=19).value, "I018")
        self.assertEqual(ws.cell(row=2, column=1).value, "a001")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (1, 1)\nX001")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (1, 2)\nX002")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (1, 3)\nX003")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (1, 4)\nX004")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (1, 5)\nX005")
        self.assertEqual(ws.cell(row=2, column=7).value, "P1 (1, 6)\nX006")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (1, 1)\nX007")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (1, 2)\nX008")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (1, 3)\nX009")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (1, 4)\nX010")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (1, 5)\nX011")
        self.assertEqual(ws.cell(row=2, column=13).value, "P2 (1, 6)\nX012")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (1, 1)\nX013")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (1, 2)\nX014")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (1, 3)\nX015")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (1, 4)\nX016")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (1, 5)\nX017")
        self.assertEqual(ws.cell(row=2, column=19).value, "P3 (1, 6)\nX018")
        self.assertEqual(ws.cell(row=3, column=1).value, "a002")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (2, 1)\nX019")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (2, 2)\nX020")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (2, 3)\nX021")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (2, 4)\nX022")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (2, 5)\nX023")
        self.assertEqual(ws.cell(row=3, column=7).value, "P1 (2, 6)\nX024")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (2, 1)\nX025")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (2, 2)\nX026")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (2, 3)\nX027")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (2, 4)\nX028")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (2, 5)\nX029")
        self.assertEqual(ws.cell(row=3, column=13).value, "P2 (2, 6)\nX030")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (2, 1)\nX031")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (2, 2)\nX032")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (2, 3)\nX033")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (2, 4)\nX034")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (2, 5)\nX035")
        self.assertEqual(ws.cell(row=3, column=19).value, "P3 (2, 6)\nX036")
        self.assertEqual(ws.cell(row=4, column=1).value, "a003")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (3, 1)\nX037")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (3, 2)\nX038")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (3, 3)\nX039")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (3, 4)\nX040")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (3, 5)\nX041")
        self.assertEqual(ws.cell(row=4, column=7).value, "P1 (3, 6)\nX042")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (3, 1)\nX043")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (3, 2)\nX044")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (3, 3)\nX045")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (3, 4)\nX046")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (3, 5)\nX047")
        self.assertEqual(ws.cell(row=4, column=13).value, "P2 (3, 6)\nX048")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (3, 1)\nX049")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (3, 2)\nX050")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (3, 3)\nX051")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (3, 4)\nX052")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (3, 5)\nX053")
        self.assertEqual(ws.cell(row=4, column=19).value, "P3 (3, 6)\nX054")
        self.assertEqual(ws.cell(row=5, column=1).value, "a004")
        self.assertEqual(ws.cell(row=5, column=2).value, "P1 (4, 1)\nX055")
        self.assertEqual(ws.cell(row=5, column=3).value, "P1 (4, 2)\nX056")
        self.assertEqual(ws.cell(row=5, column=4).value, "P1 (4, 3)\nX057")
        self.assertEqual(ws.cell(row=5, column=5).value, "P1 (4, 4)\nX058")
        self.assertEqual(ws.cell(row=5, column=6).value, "P1 (4, 5)\nX059")
        self.assertEqual(ws.cell(row=5, column=7).value, "P1 (4, 6)\nX060")
        self.assertEqual(ws.cell(row=5, column=8).value, "P2 (4, 1)\nX061")
        self.assertEqual(ws.cell(row=5, column=9).value, "P2 (4, 2)\nX062")
        self.assertEqual(ws.cell(row=5, column=10).value, "P2 (4, 3)\nX063")
        self.assertEqual(ws.cell(row=5, column=11).value, "P2 (4, 4)\nX064")
        self.assertEqual(ws.cell(row=5, column=12).value, "P2 (4, 5)\nX065")
        self.assertEqual(ws.cell(row=5, column=13).value, "P2 (4, 6)\nX066")
        self.assertEqual(ws.cell(row=5, column=14).value, "P3 (4, 1)\nX067")
        self.assertEqual(ws.cell(row=5, column=15).value, "P3 (4, 2)\nX068")
        self.assertEqual(ws.cell(row=5, column=16).value, "P3 (4, 3)\nX069")
        self.assertEqual(ws.cell(row=5, column=17).value, "P3 (4, 4)\nX070")
        self.assertEqual(ws.cell(row=5, column=18).value, "P3 (4, 5)\nX071")
        self.assertEqual(ws.cell(row=5, column=19).value, "P3 (4, 6)\nX072")
        self.assertEqual(ws.cell(row=6, column=1).value, "a005")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (1, 1)\nX073")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (1, 2)\nX074")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (1, 3)\nX075")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (1, 4)\nX076")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (1, 5)\nX077")
        self.assertEqual(ws.cell(row=6, column=7).value, "P4 (1, 6)\nX078")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (1, 1)\nX079")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (1, 2)\nX080")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (1, 3)\nX081")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (1, 4)\nX082")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (1, 5)\nX083")
        self.assertEqual(ws.cell(row=6, column=13).value, "P5 (1, 6)\nX084")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (1, 1)\nX085")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (1, 2)\nX086")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (1, 3)\nX087")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (1, 4)\nX088")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (1, 5)\nX089")
        self.assertEqual(ws.cell(row=6, column=19).value, "P6 (1, 6)\nX090")
        self.assertEqual(ws.cell(row=7, column=1).value, "a006")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (2, 1)\nX091")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (2, 2)\nX092")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (2, 3)\nX093")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (2, 4)\nX094")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (2, 5)\nX095")
        self.assertEqual(ws.cell(row=7, column=7).value, "P4 (2, 6)\nX096")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (2, 1)\nX097")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (2, 2)\nX098")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (2, 3)\nX099")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (2, 4)\nX100")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (2, 5)\nX101")
        self.assertEqual(ws.cell(row=7, column=13).value, "P5 (2, 6)\nX102")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (2, 1)\nX103")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (2, 2)\nX104")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (2, 3)\nX105")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (2, 4)\nX106")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (2, 5)\nX107")
        self.assertEqual(ws.cell(row=7, column=19).value, "P6 (2, 6)\nX108")
        self.assertEqual(ws.cell(row=8, column=1).value, "a007")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (3, 1)\nX109")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (3, 2)\nX110")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (3, 3)\nX111")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (3, 4)\nX112")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (3, 5)\nX113")
        self.assertEqual(ws.cell(row=8, column=7).value, "P4 (3, 6)\nX114")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (3, 1)\nX115")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (3, 2)\nX116")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (3, 3)\nX117")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (3, 4)\nX118")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (3, 5)\nX119")
        self.assertEqual(ws.cell(row=8, column=13).value, "P5 (3, 6)\nX120")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (3, 1)\nX121")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (3, 2)\nX122")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (3, 3)\nX123")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (3, 4)\nX124")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (3, 5)\nX125")
        self.assertEqual(ws.cell(row=8, column=19).value, "P6 (3, 6)\nX126")
        self.assertEqual(ws.cell(row=9, column=1).value, "a008")
        self.assertEqual(ws.cell(row=9, column=2).value, "P4 (4, 1)\nX127")
        self.assertEqual(ws.cell(row=9, column=3).value, "P4 (4, 2)\nX128")
        self.assertEqual(ws.cell(row=9, column=4).value, "P4 (4, 3)\nX129")
        self.assertEqual(ws.cell(row=9, column=5).value, "P4 (4, 4)\nX130")
        self.assertEqual(ws.cell(row=9, column=6).value, "P4 (4, 5)\nX131")
        self.assertEqual(ws.cell(row=9, column=7).value, "P4 (4, 6)\nX132")
        self.assertEqual(ws.cell(row=9, column=8).value, "P5 (4, 1)\nX133")
        self.assertEqual(ws.cell(row=9, column=9).value, "P5 (4, 2)\nX134")
        self.assertEqual(ws.cell(row=9, column=10).value, "P5 (4, 3)\nX135")
        self.assertEqual(ws.cell(row=9, column=11).value, "P5 (4, 4)\nX136")
        self.assertEqual(ws.cell(row=9, column=12).value, "P5 (4, 5)\nX137")
        self.assertEqual(ws.cell(row=9, column=13).value, "P5 (4, 6)\nX138")
        self.assertEqual(ws.cell(row=9, column=14).value, "P6 (4, 1)\nX139")
        self.assertEqual(ws.cell(row=9, column=15).value, "P6 (4, 2)\nX140")
        self.assertEqual(ws.cell(row=9, column=16).value, "P6 (4, 3)\nX141")
        self.assertEqual(ws.cell(row=9, column=17).value, "P6 (4, 4)\nX142")
        self.assertEqual(ws.cell(row=9, column=18).value, "P6 (4, 5)\nX143")
        self.assertEqual(ws.cell(row=9, column=19).value, "P6 (4, 6)\nX144")
        self.assertEqual(ws.cell(row=11, column=1).value,
                         u"Add 10.00µL of Sugar to media")

        # Check that sheet exists in spreadsheet
        self.assertTrue("Cells for Plate Array A1" in wb.sheetnames)
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertTrue('A2:C2' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=2, column=1).value, "Predilution")
        self.assertEqual(ws.cell(row=3, column=1).value, "Predilution factor")
        self.assertEqual(ws.cell(row=3, column=2).value, 100)
        self.assertEqual(ws.cell(row=3, column=3).value, "x")
        self.assertEqual(ws.cell(row=4, column=1).value, "Media volume")
        self.assertEqual(ws.cell(row=4, column=2).value, 990)
        self.assertEqual(ws.cell(row=4, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=5, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=5, column=2).value, 10)
        self.assertEqual(ws.cell(row=5, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=6, column=1).value, "Predilution OD600")
        self.assertEqual(ws.cell(row=6, column=2).value, None)
        self.assertEqual(ws.cell(row=6, column=3).value, None)
        self.assertTrue('A7:C7' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=7, column=1).value, "Inoculation")
        self.assertEqual(ws.cell(row=8, column=1).value, "Target OD600")
        self.assertEqual(ws.cell(row=8, column=2).value, 1e-5)
        self.assertEqual(ws.cell(row=8, column=3).value, None)
        self.assertEqual(ws.cell(row=9, column=1).value, "Predilution volume")
        self.assertEqual(ws.cell(row=9, column=2).value, "=0.8/B6")
        self.assertEqual(ws.cell(row=9, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=10, column=1).value, "Add into 80.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_cells_and_inducer_workbook(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.total_media_vol = 80000.
        # Add some information for cell setup
        p.cell_strain_name = 'Test strain 1'
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_initial_od600 = 1e-5

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.shot_vol = 5.
        iptg.set_gradient(min=1e-6, max=1e-3, n=18, scale='log')
        p.apply_inducer(iptg, apply_to='rows')

        # Create second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.shot_vol = 5.
        atc.set_gradient(min=0.5, max=50, n=8, scale='log')
        p.apply_inducer(atc, apply_to='cols')

        # Create inducer for plate wells
        xyl = platedesign.inducer.ChemicalInducer(
            name='Xylose',
            units=u'%')
        xyl.set_gradient(min=1e-6, max=1e-3, n=144, scale='log')
        p.apply_inducer(xyl, apply_to='wells')

        # Create inducer for plate
        sugar = platedesign.inducer.ChemicalInducer(
            name='Sugar',
            units=u'ng/µL')
        sugar.shot_vol = 10.
        sugar.concentrations = [2]
        p.apply_inducer(sugar, apply_to='media')

        # Create new spreadsheet
        wb = openpyxl.Workbook()
        # Save instructions on existing spreadsheet
        p.save_rep_setup_instructions(workbook=wb)

        # Check that sheet exists in spreadsheet
        self.assertTrue("Inducers for Plate Array A1" in wb.sheetnames)
        # Check inducer inoculation instructions
        ws = wb.get_sheet_by_name("Inducers for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, '')
        self.assertEqual(ws.cell(row=1, column=2).value, "I001")
        self.assertEqual(ws.cell(row=1, column=3).value, "I002")
        self.assertEqual(ws.cell(row=1, column=4).value, "I003")
        self.assertEqual(ws.cell(row=1, column=5).value, "I004")
        self.assertEqual(ws.cell(row=1, column=6).value, "I005")
        self.assertEqual(ws.cell(row=1, column=7).value, "I006")
        self.assertEqual(ws.cell(row=1, column=8).value, "I007")
        self.assertEqual(ws.cell(row=1, column=9).value, "I008")
        self.assertEqual(ws.cell(row=1, column=10).value, "I009")
        self.assertEqual(ws.cell(row=1, column=11).value, "I010")
        self.assertEqual(ws.cell(row=1, column=12).value, "I011")
        self.assertEqual(ws.cell(row=1, column=13).value, "I012")
        self.assertEqual(ws.cell(row=1, column=14).value, "I013")
        self.assertEqual(ws.cell(row=1, column=15).value, "I014")
        self.assertEqual(ws.cell(row=1, column=16).value, "I015")
        self.assertEqual(ws.cell(row=1, column=17).value, "I016")
        self.assertEqual(ws.cell(row=1, column=18).value, "I017")
        self.assertEqual(ws.cell(row=1, column=19).value, "I018")
        self.assertEqual(ws.cell(row=2, column=1).value, "a001")
        self.assertEqual(ws.cell(row=2, column=2).value, "P1 (1, 1)\nX001")
        self.assertEqual(ws.cell(row=2, column=3).value, "P1 (1, 2)\nX002")
        self.assertEqual(ws.cell(row=2, column=4).value, "P1 (1, 3)\nX003")
        self.assertEqual(ws.cell(row=2, column=5).value, "P1 (1, 4)\nX004")
        self.assertEqual(ws.cell(row=2, column=6).value, "P1 (1, 5)\nX005")
        self.assertEqual(ws.cell(row=2, column=7).value, "P1 (1, 6)\nX006")
        self.assertEqual(ws.cell(row=2, column=8).value, "P2 (1, 1)\nX007")
        self.assertEqual(ws.cell(row=2, column=9).value, "P2 (1, 2)\nX008")
        self.assertEqual(ws.cell(row=2, column=10).value, "P2 (1, 3)\nX009")
        self.assertEqual(ws.cell(row=2, column=11).value, "P2 (1, 4)\nX010")
        self.assertEqual(ws.cell(row=2, column=12).value, "P2 (1, 5)\nX011")
        self.assertEqual(ws.cell(row=2, column=13).value, "P2 (1, 6)\nX012")
        self.assertEqual(ws.cell(row=2, column=14).value, "P3 (1, 1)\nX013")
        self.assertEqual(ws.cell(row=2, column=15).value, "P3 (1, 2)\nX014")
        self.assertEqual(ws.cell(row=2, column=16).value, "P3 (1, 3)\nX015")
        self.assertEqual(ws.cell(row=2, column=17).value, "P3 (1, 4)\nX016")
        self.assertEqual(ws.cell(row=2, column=18).value, "P3 (1, 5)\nX017")
        self.assertEqual(ws.cell(row=2, column=19).value, "P3 (1, 6)\nX018")
        self.assertEqual(ws.cell(row=3, column=1).value, "a002")
        self.assertEqual(ws.cell(row=3, column=2).value, "P1 (2, 1)\nX019")
        self.assertEqual(ws.cell(row=3, column=3).value, "P1 (2, 2)\nX020")
        self.assertEqual(ws.cell(row=3, column=4).value, "P1 (2, 3)\nX021")
        self.assertEqual(ws.cell(row=3, column=5).value, "P1 (2, 4)\nX022")
        self.assertEqual(ws.cell(row=3, column=6).value, "P1 (2, 5)\nX023")
        self.assertEqual(ws.cell(row=3, column=7).value, "P1 (2, 6)\nX024")
        self.assertEqual(ws.cell(row=3, column=8).value, "P2 (2, 1)\nX025")
        self.assertEqual(ws.cell(row=3, column=9).value, "P2 (2, 2)\nX026")
        self.assertEqual(ws.cell(row=3, column=10).value, "P2 (2, 3)\nX027")
        self.assertEqual(ws.cell(row=3, column=11).value, "P2 (2, 4)\nX028")
        self.assertEqual(ws.cell(row=3, column=12).value, "P2 (2, 5)\nX029")
        self.assertEqual(ws.cell(row=3, column=13).value, "P2 (2, 6)\nX030")
        self.assertEqual(ws.cell(row=3, column=14).value, "P3 (2, 1)\nX031")
        self.assertEqual(ws.cell(row=3, column=15).value, "P3 (2, 2)\nX032")
        self.assertEqual(ws.cell(row=3, column=16).value, "P3 (2, 3)\nX033")
        self.assertEqual(ws.cell(row=3, column=17).value, "P3 (2, 4)\nX034")
        self.assertEqual(ws.cell(row=3, column=18).value, "P3 (2, 5)\nX035")
        self.assertEqual(ws.cell(row=3, column=19).value, "P3 (2, 6)\nX036")
        self.assertEqual(ws.cell(row=4, column=1).value, "a003")
        self.assertEqual(ws.cell(row=4, column=2).value, "P1 (3, 1)\nX037")
        self.assertEqual(ws.cell(row=4, column=3).value, "P1 (3, 2)\nX038")
        self.assertEqual(ws.cell(row=4, column=4).value, "P1 (3, 3)\nX039")
        self.assertEqual(ws.cell(row=4, column=5).value, "P1 (3, 4)\nX040")
        self.assertEqual(ws.cell(row=4, column=6).value, "P1 (3, 5)\nX041")
        self.assertEqual(ws.cell(row=4, column=7).value, "P1 (3, 6)\nX042")
        self.assertEqual(ws.cell(row=4, column=8).value, "P2 (3, 1)\nX043")
        self.assertEqual(ws.cell(row=4, column=9).value, "P2 (3, 2)\nX044")
        self.assertEqual(ws.cell(row=4, column=10).value, "P2 (3, 3)\nX045")
        self.assertEqual(ws.cell(row=4, column=11).value, "P2 (3, 4)\nX046")
        self.assertEqual(ws.cell(row=4, column=12).value, "P2 (3, 5)\nX047")
        self.assertEqual(ws.cell(row=4, column=13).value, "P2 (3, 6)\nX048")
        self.assertEqual(ws.cell(row=4, column=14).value, "P3 (3, 1)\nX049")
        self.assertEqual(ws.cell(row=4, column=15).value, "P3 (3, 2)\nX050")
        self.assertEqual(ws.cell(row=4, column=16).value, "P3 (3, 3)\nX051")
        self.assertEqual(ws.cell(row=4, column=17).value, "P3 (3, 4)\nX052")
        self.assertEqual(ws.cell(row=4, column=18).value, "P3 (3, 5)\nX053")
        self.assertEqual(ws.cell(row=4, column=19).value, "P3 (3, 6)\nX054")
        self.assertEqual(ws.cell(row=5, column=1).value, "a004")
        self.assertEqual(ws.cell(row=5, column=2).value, "P1 (4, 1)\nX055")
        self.assertEqual(ws.cell(row=5, column=3).value, "P1 (4, 2)\nX056")
        self.assertEqual(ws.cell(row=5, column=4).value, "P1 (4, 3)\nX057")
        self.assertEqual(ws.cell(row=5, column=5).value, "P1 (4, 4)\nX058")
        self.assertEqual(ws.cell(row=5, column=6).value, "P1 (4, 5)\nX059")
        self.assertEqual(ws.cell(row=5, column=7).value, "P1 (4, 6)\nX060")
        self.assertEqual(ws.cell(row=5, column=8).value, "P2 (4, 1)\nX061")
        self.assertEqual(ws.cell(row=5, column=9).value, "P2 (4, 2)\nX062")
        self.assertEqual(ws.cell(row=5, column=10).value, "P2 (4, 3)\nX063")
        self.assertEqual(ws.cell(row=5, column=11).value, "P2 (4, 4)\nX064")
        self.assertEqual(ws.cell(row=5, column=12).value, "P2 (4, 5)\nX065")
        self.assertEqual(ws.cell(row=5, column=13).value, "P2 (4, 6)\nX066")
        self.assertEqual(ws.cell(row=5, column=14).value, "P3 (4, 1)\nX067")
        self.assertEqual(ws.cell(row=5, column=15).value, "P3 (4, 2)\nX068")
        self.assertEqual(ws.cell(row=5, column=16).value, "P3 (4, 3)\nX069")
        self.assertEqual(ws.cell(row=5, column=17).value, "P3 (4, 4)\nX070")
        self.assertEqual(ws.cell(row=5, column=18).value, "P3 (4, 5)\nX071")
        self.assertEqual(ws.cell(row=5, column=19).value, "P3 (4, 6)\nX072")
        self.assertEqual(ws.cell(row=6, column=1).value, "a005")
        self.assertEqual(ws.cell(row=6, column=2).value, "P4 (1, 1)\nX073")
        self.assertEqual(ws.cell(row=6, column=3).value, "P4 (1, 2)\nX074")
        self.assertEqual(ws.cell(row=6, column=4).value, "P4 (1, 3)\nX075")
        self.assertEqual(ws.cell(row=6, column=5).value, "P4 (1, 4)\nX076")
        self.assertEqual(ws.cell(row=6, column=6).value, "P4 (1, 5)\nX077")
        self.assertEqual(ws.cell(row=6, column=7).value, "P4 (1, 6)\nX078")
        self.assertEqual(ws.cell(row=6, column=8).value, "P5 (1, 1)\nX079")
        self.assertEqual(ws.cell(row=6, column=9).value, "P5 (1, 2)\nX080")
        self.assertEqual(ws.cell(row=6, column=10).value, "P5 (1, 3)\nX081")
        self.assertEqual(ws.cell(row=6, column=11).value, "P5 (1, 4)\nX082")
        self.assertEqual(ws.cell(row=6, column=12).value, "P5 (1, 5)\nX083")
        self.assertEqual(ws.cell(row=6, column=13).value, "P5 (1, 6)\nX084")
        self.assertEqual(ws.cell(row=6, column=14).value, "P6 (1, 1)\nX085")
        self.assertEqual(ws.cell(row=6, column=15).value, "P6 (1, 2)\nX086")
        self.assertEqual(ws.cell(row=6, column=16).value, "P6 (1, 3)\nX087")
        self.assertEqual(ws.cell(row=6, column=17).value, "P6 (1, 4)\nX088")
        self.assertEqual(ws.cell(row=6, column=18).value, "P6 (1, 5)\nX089")
        self.assertEqual(ws.cell(row=6, column=19).value, "P6 (1, 6)\nX090")
        self.assertEqual(ws.cell(row=7, column=1).value, "a006")
        self.assertEqual(ws.cell(row=7, column=2).value, "P4 (2, 1)\nX091")
        self.assertEqual(ws.cell(row=7, column=3).value, "P4 (2, 2)\nX092")
        self.assertEqual(ws.cell(row=7, column=4).value, "P4 (2, 3)\nX093")
        self.assertEqual(ws.cell(row=7, column=5).value, "P4 (2, 4)\nX094")
        self.assertEqual(ws.cell(row=7, column=6).value, "P4 (2, 5)\nX095")
        self.assertEqual(ws.cell(row=7, column=7).value, "P4 (2, 6)\nX096")
        self.assertEqual(ws.cell(row=7, column=8).value, "P5 (2, 1)\nX097")
        self.assertEqual(ws.cell(row=7, column=9).value, "P5 (2, 2)\nX098")
        self.assertEqual(ws.cell(row=7, column=10).value, "P5 (2, 3)\nX099")
        self.assertEqual(ws.cell(row=7, column=11).value, "P5 (2, 4)\nX100")
        self.assertEqual(ws.cell(row=7, column=12).value, "P5 (2, 5)\nX101")
        self.assertEqual(ws.cell(row=7, column=13).value, "P5 (2, 6)\nX102")
        self.assertEqual(ws.cell(row=7, column=14).value, "P6 (2, 1)\nX103")
        self.assertEqual(ws.cell(row=7, column=15).value, "P6 (2, 2)\nX104")
        self.assertEqual(ws.cell(row=7, column=16).value, "P6 (2, 3)\nX105")
        self.assertEqual(ws.cell(row=7, column=17).value, "P6 (2, 4)\nX106")
        self.assertEqual(ws.cell(row=7, column=18).value, "P6 (2, 5)\nX107")
        self.assertEqual(ws.cell(row=7, column=19).value, "P6 (2, 6)\nX108")
        self.assertEqual(ws.cell(row=8, column=1).value, "a007")
        self.assertEqual(ws.cell(row=8, column=2).value, "P4 (3, 1)\nX109")
        self.assertEqual(ws.cell(row=8, column=3).value, "P4 (3, 2)\nX110")
        self.assertEqual(ws.cell(row=8, column=4).value, "P4 (3, 3)\nX111")
        self.assertEqual(ws.cell(row=8, column=5).value, "P4 (3, 4)\nX112")
        self.assertEqual(ws.cell(row=8, column=6).value, "P4 (3, 5)\nX113")
        self.assertEqual(ws.cell(row=8, column=7).value, "P4 (3, 6)\nX114")
        self.assertEqual(ws.cell(row=8, column=8).value, "P5 (3, 1)\nX115")
        self.assertEqual(ws.cell(row=8, column=9).value, "P5 (3, 2)\nX116")
        self.assertEqual(ws.cell(row=8, column=10).value, "P5 (3, 3)\nX117")
        self.assertEqual(ws.cell(row=8, column=11).value, "P5 (3, 4)\nX118")
        self.assertEqual(ws.cell(row=8, column=12).value, "P5 (3, 5)\nX119")
        self.assertEqual(ws.cell(row=8, column=13).value, "P5 (3, 6)\nX120")
        self.assertEqual(ws.cell(row=8, column=14).value, "P6 (3, 1)\nX121")
        self.assertEqual(ws.cell(row=8, column=15).value, "P6 (3, 2)\nX122")
        self.assertEqual(ws.cell(row=8, column=16).value, "P6 (3, 3)\nX123")
        self.assertEqual(ws.cell(row=8, column=17).value, "P6 (3, 4)\nX124")
        self.assertEqual(ws.cell(row=8, column=18).value, "P6 (3, 5)\nX125")
        self.assertEqual(ws.cell(row=8, column=19).value, "P6 (3, 6)\nX126")
        self.assertEqual(ws.cell(row=9, column=1).value, "a008")
        self.assertEqual(ws.cell(row=9, column=2).value, "P4 (4, 1)\nX127")
        self.assertEqual(ws.cell(row=9, column=3).value, "P4 (4, 2)\nX128")
        self.assertEqual(ws.cell(row=9, column=4).value, "P4 (4, 3)\nX129")
        self.assertEqual(ws.cell(row=9, column=5).value, "P4 (4, 4)\nX130")
        self.assertEqual(ws.cell(row=9, column=6).value, "P4 (4, 5)\nX131")
        self.assertEqual(ws.cell(row=9, column=7).value, "P4 (4, 6)\nX132")
        self.assertEqual(ws.cell(row=9, column=8).value, "P5 (4, 1)\nX133")
        self.assertEqual(ws.cell(row=9, column=9).value, "P5 (4, 2)\nX134")
        self.assertEqual(ws.cell(row=9, column=10).value, "P5 (4, 3)\nX135")
        self.assertEqual(ws.cell(row=9, column=11).value, "P5 (4, 4)\nX136")
        self.assertEqual(ws.cell(row=9, column=12).value, "P5 (4, 5)\nX137")
        self.assertEqual(ws.cell(row=9, column=13).value, "P5 (4, 6)\nX138")
        self.assertEqual(ws.cell(row=9, column=14).value, "P6 (4, 1)\nX139")
        self.assertEqual(ws.cell(row=9, column=15).value, "P6 (4, 2)\nX140")
        self.assertEqual(ws.cell(row=9, column=16).value, "P6 (4, 3)\nX141")
        self.assertEqual(ws.cell(row=9, column=17).value, "P6 (4, 4)\nX142")
        self.assertEqual(ws.cell(row=9, column=18).value, "P6 (4, 5)\nX143")
        self.assertEqual(ws.cell(row=9, column=19).value, "P6 (4, 6)\nX144")
        self.assertEqual(ws.cell(row=11, column=1).value,
                         u"Add 10.00µL of Sugar to media")

        # Check that sheet exists in spreadsheet
        self.assertTrue("Cells for Plate Array A1" in wb.sheetnames)
        # Check cell inoculation instructions
        ws = wb.get_sheet_by_name("Cells for Plate Array A1")
        self.assertEqual(ws.cell(row=1, column=1).value, "Strain Name")
        self.assertEqual(ws.cell(row=1, column=2).value, "Test strain 1")
        self.assertTrue('A2:C2' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=2, column=1).value, "Predilution")
        self.assertEqual(ws.cell(row=3, column=1).value, "Predilution factor")
        self.assertEqual(ws.cell(row=3, column=2).value, 100)
        self.assertEqual(ws.cell(row=3, column=3).value, "x")
        self.assertEqual(ws.cell(row=4, column=1).value, "Media volume")
        self.assertEqual(ws.cell(row=4, column=2).value, 990)
        self.assertEqual(ws.cell(row=4, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=5, column=1).value, "Preculture volume")
        self.assertEqual(ws.cell(row=5, column=2).value, 10)
        self.assertEqual(ws.cell(row=5, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=6, column=1).value, "Predilution OD600")
        self.assertEqual(ws.cell(row=6, column=2).value, None)
        self.assertEqual(ws.cell(row=6, column=3).value, None)
        self.assertTrue('A7:C7' in ws.merged_cell_ranges)
        self.assertEqual(ws.cell(row=7, column=1).value, "Inoculation")
        self.assertEqual(ws.cell(row=8, column=1).value, "Target OD600")
        self.assertEqual(ws.cell(row=8, column=2).value, 1e-5)
        self.assertEqual(ws.cell(row=8, column=3).value, None)
        self.assertEqual(ws.cell(row=9, column=1).value, "Predilution volume")
        self.assertEqual(ws.cell(row=9, column=2).value, "=0.8/B6")
        self.assertEqual(ws.cell(row=9, column=3).value, u"µL")
        self.assertEqual(ws.cell(row=10, column=1).value, "Add into 80.00mL "
            "media, and distribute into plate wells.")

    def test_save_rep_setup_instructions_argument_error(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Run save_rep_setup_instructions with no arguments
        self.assertRaises(ValueError, p.save_rep_setup_instructions)

    def test_save_rep_setup_files(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Check that save_rep_setup_files runs successfully
        p.save_rep_setup_files()
        # save_exp_setup_files does not do anything in Plate. There is no need
        # to check for results.

    def test_add_inducer_setup_instructions_sheet_error(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create workbook
        wb = openpyxl.Workbook()
        # Add sheet
        wb.create_sheet("Test sheet")
        # Attempt to add cell setup instructions
        self.assertRaises(ValueError,
                          p.add_inducer_setup_instructions,
                          wb,
                          "Test sheet")

    def test_add_cell_setup_instructions_sheet_error(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        # Create workbook
        wb = openpyxl.Workbook()
        # Add sheet
        wb.create_sheet("Test sheet")
        # Attempt to add cell setup instructions
        self.assertRaises(ValueError,
                          p.add_cell_setup_instructions,
                          wb,
                          "Test sheet")

    def test_close_plates_no_inducer_no_cell_inoculation(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
            # Check well info
            well_info = pandas.DataFrame()
            well_info['Measure'] = [True]*24
            pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_metadata(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'
        # Add some metadata
        p.metadata['Meta 1'] = 'Value 1'
        p.metadata['Meta 2'] = 'Value 2'
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 4)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
            self.assertTrue('Meta 1' in cp.plate_info)
            self.assertEqual(cp.plate_info['Meta 1'], 'Value 1')
            self.assertTrue('Meta 2' in cp.plate_info)
            self.assertEqual(cp.plate_info['Meta 2'], 'Value 2')
            # Check well info
            well_info = pandas.DataFrame()
            well_info['Measure'] = [True]*24
            pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_cell_inoculation_1(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'
        # Add cell inoculation info
        p.cell_setup_method = 'fixed_volume'
        p.cell_shot_vol = 5
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 4)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
            self.assertTrue('Preculture Dilution' in cp.plate_info)
            self.assertEqual(cp.plate_info['Preculture Dilution'], 1)
            self.assertTrue('Cell Inoculated Vol.' in cp.plate_info)
            self.assertEqual(cp.plate_info['Cell Inoculated Vol.'], 5)
            # Check well info
            well_info = pandas.DataFrame()
            well_info['Measure'] = [True]*24
            pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_cell_inoculation_2(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'
        # Add cell inoculation info
        p.cell_setup_method = 'fixed_volume'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_shot_vol = 5
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 4)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
            self.assertTrue('Preculture Dilution' in cp.plate_info)
            self.assertEqual(cp.plate_info['Preculture Dilution'], 100)
            self.assertTrue('Cell Inoculated Vol.' in cp.plate_info)
            self.assertEqual(cp.plate_info['Cell Inoculated Vol.'], 5)
            # Check well info
            well_info = pandas.DataFrame()
            well_info['Measure'] = [True]*24
            pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_cell_inoculation_3(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'
        # Add cell inoculation info
        p.cell_setup_method = 'fixed_od600'
        p.cell_initial_od600 = 1e-5
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 4)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
            self.assertTrue('Preculture Dilution' in cp.plate_info)
            self.assertEqual(cp.plate_info['Preculture Dilution'], 1)
            self.assertTrue('Initial OD600' in cp.plate_info)
            self.assertEqual(cp.plate_info['Initial OD600'], 1e-5)
            # Check well info
            well_info = pandas.DataFrame()
            well_info['Measure'] = [True]*24
            pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_cell_inoculation_4(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'
        # Add cell inoculation info
        p.cell_setup_method = 'fixed_od600'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_initial_od600 = 1e-5
        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 4)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
            self.assertTrue('Preculture Dilution' in cp.plate_info)
            self.assertEqual(cp.plate_info['Preculture Dilution'], 100)
            self.assertTrue('Initial OD600' in cp.plate_info)
            self.assertEqual(cp.plate_info['Initial OD600'], 1e-5)
            # Check well info
            well_info = pandas.DataFrame()
            well_info['Measure'] = [True]*24
            pandas.util.testing.assert_frame_equal(cp.well_info, well_info)

    def test_close_plates_inducer_row_1(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(18) + 3
        p.apply_inducer(iptg, apply_to='rows')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.]
        well_info_1['Measure'] = [True]*24

        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,]
        well_info_2['Measure'] = [True]*24

        well_info_3 = pandas.DataFrame()
        well_info_3[u'IPTG Concentration (µM)'] = \
            [15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,]
        well_info_3['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_3)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_3)

    def test_close_plates_inducer_row_2(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(18) + 3
        p.apply_inducer(iptg, apply_to='rows')

        # Second inducer for plate rows
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = (numpy.arange(18.) + 1)/10.
        p.apply_inducer(atc, apply_to='rows')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.]
        well_info_1[u'aTc Concentration (ng/µL)'] = \
            [0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
             0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
             0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
             0.1, 0.2, 0.3, 0.4, 0.5, 0.6,]
        well_info_1['Measure'] = [True]*24

        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,]
        well_info_2[u'aTc Concentration (ng/µL)'] = \
            [0.7, 0.8, 0.9, 1.0, 1.1, 1.2,
             0.7, 0.8, 0.9, 1.0, 1.1, 1.2,
             0.7, 0.8, 0.9, 1.0, 1.1, 1.2,
             0.7, 0.8, 0.9, 1.0, 1.1, 1.2,]
        well_info_2['Measure'] = [True]*24

        well_info_3 = pandas.DataFrame()
        well_info_3[u'IPTG Concentration (µM)'] = \
            [15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,]
        well_info_3[u'aTc Concentration (ng/µL)'] = \
            [1.3, 1.4, 1.5, 1.6, 1.7, 1.8,
             1.3, 1.4, 1.5, 1.6, 1.7, 1.8,
             1.3, 1.4, 1.5, 1.6, 1.7, 1.8,
             1.3, 1.4, 1.5, 1.6, 1.7, 1.8,]
        well_info_3['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_3)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_3)

    def test_close_plates_inducer_row_3(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(18) + 3
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='rows')

        # Second inducer for plate rows
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = (numpy.arange(18.) + 1)/10.
        p.apply_inducer(atc, apply_to='rows')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [14., 18., 20., 4., 13., 7.,
             14., 18., 20., 4., 13., 7.,
             14., 18., 20., 4., 13., 7.,
             14., 18., 20., 4., 13., 7.,]
        well_info_1[u'aTc Concentration (ng/µL)'] = \
            [0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
             0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
             0.1, 0.2, 0.3, 0.4, 0.5, 0.6,
             0.1, 0.2, 0.3, 0.4, 0.5, 0.6,]
        well_info_1['Measure'] = [True]*24

        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [19., 16., 12., 3., 11., 10.,
             19., 16., 12., 3., 11., 10.,
             19., 16., 12., 3., 11., 10.,
             19., 16., 12., 3., 11., 10.,]
        well_info_2[u'aTc Concentration (ng/µL)'] = \
            [0.7, 0.8, 0.9, 1.0, 1.1, 1.2,
             0.7, 0.8, 0.9, 1.0, 1.1, 1.2,
             0.7, 0.8, 0.9, 1.0, 1.1, 1.2,
             0.7, 0.8, 0.9, 1.0, 1.1, 1.2,]
        well_info_2['Measure'] = [True]*24

        well_info_3 = pandas.DataFrame()
        well_info_3[u'IPTG Concentration (µM)'] = \
            [8., 9., 6., 15., 17., 5.,
             8., 9., 6., 15., 17., 5.,
             8., 9., 6., 15., 17., 5.,
             8., 9., 6., 15., 17., 5.,]
        well_info_3[u'aTc Concentration (ng/µL)'] = \
            [1.3, 1.4, 1.5, 1.6, 1.7, 1.8,
             1.3, 1.4, 1.5, 1.6, 1.7, 1.8,
             1.3, 1.4, 1.5, 1.6, 1.7, 1.8,
             1.3, 1.4, 1.5, 1.6, 1.7, 1.8,]
        well_info_3['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_3)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_3)

    def test_close_plates_inducer_col_1(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(8) + 3
        p.apply_inducer(iptg, apply_to='cols')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [3., 3., 3., 3., 3., 3.,
             4., 4., 4., 4., 4., 4.,
             5., 5., 5., 5., 5., 5.,
             6., 6., 6., 6., 6., 6.]
        well_info_1['Measure'] = [True]*24
        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [7., 7., 7., 7., 7., 7.,
             8., 8., 8., 8., 8., 8.,
             9., 9., 9., 9., 9., 9.,
             10., 10., 10., 10., 10., 10.]
        well_info_2['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_2)

    def test_close_plates_inducer_col_2(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(8) + 3
        p.apply_inducer(iptg, apply_to='cols')

        # Second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = (numpy.arange(8.) + 1)/10.
        p.apply_inducer(atc, apply_to='cols')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [3., 3., 3., 3., 3., 3.,
             4., 4., 4., 4., 4., 4.,
             5., 5., 5., 5., 5., 5.,
             6., 6., 6., 6., 6., 6.]
        well_info_1[u'aTc Concentration (ng/µL)'] = \
            [0.1, 0.1, 0.1, 0.1, 0.1, 0.1,
             0.2, 0.2, 0.2, 0.2, 0.2, 0.2,
             0.3, 0.3, 0.3, 0.3, 0.3, 0.3,
             0.4, 0.4, 0.4, 0.4, 0.4, 0.4,]
        well_info_1['Measure'] = [True]*24

        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [7., 7., 7., 7., 7., 7.,
             8., 8., 8., 8., 8., 8.,
             9., 9., 9., 9., 9., 9.,
             10., 10., 10., 10., 10., 10.]
        well_info_2[u'aTc Concentration (ng/µL)'] = \
            [0.5, 0.5, 0.5, 0.5, 0.5, 0.5,
             0.6, 0.6, 0.6, 0.6, 0.6, 0.6,
             0.7, 0.7, 0.7, 0.7, 0.7, 0.7,
             0.8, 0.8, 0.8, 0.8, 0.8, 0.8,]
        well_info_2['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_2)

    def test_close_plates_inducer_col_3(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(8) + 3
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='cols')

        # Second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = (numpy.arange(8.) + 1)/10.
        p.apply_inducer(atc, apply_to='cols')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [3., 3., 3., 3., 3., 3.,
             5., 5., 5., 5., 5., 5.,
             6., 6., 6., 6., 6., 6.,
             9., 9., 9., 9., 9., 9.,]
        well_info_1[u'aTc Concentration (ng/µL)'] = \
            [0.1, 0.1, 0.1, 0.1, 0.1, 0.1,
             0.2, 0.2, 0.2, 0.2, 0.2, 0.2,
             0.3, 0.3, 0.3, 0.3, 0.3, 0.3,
             0.4, 0.4, 0.4, 0.4, 0.4, 0.4,]
        well_info_1['Measure'] = [True]*24

        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [10., 10., 10., 10., 10., 10.,
              7., 7., 7., 7., 7., 7.,
              8., 8., 8., 8., 8., 8.,
              4., 4., 4., 4., 4., 4.,]
        well_info_2[u'aTc Concentration (ng/µL)'] = \
            [0.5, 0.5, 0.5, 0.5, 0.5, 0.5,
             0.6, 0.6, 0.6, 0.6, 0.6, 0.6,
             0.7, 0.7, 0.7, 0.7, 0.7, 0.7,
             0.8, 0.8, 0.8, 0.8, 0.8, 0.8,]
        well_info_2['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_2)

    def test_close_plates_inducer_wells_1(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(144) + 3
        p.apply_inducer(iptg, apply_to='wells')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [  3.,   4.,   5.,   6.,   7.,   8.,
              21.,  22.,  23.,  24.,  25.,  26.,
              39.,  40.,  41.,  42.,  43.,  44.,
              57.,  58.,  59.,  60.,  61.,  62.]
        well_info_1['Measure'] = [True]*24

        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [  9.,  10.,  11.,  12.,  13.,  14.,
              27.,  28.,  29.,  30.,  31.,  32.,
              45.,  46.,  47.,  48.,  49.,  50.,
              63.,  64.,  65.,  66.,  67.,  68.]
        well_info_2['Measure'] = [True]*24

        well_info_3 = pandas.DataFrame()
        well_info_3[u'IPTG Concentration (µM)'] = \
            [ 15.,  16.,  17.,  18.,  19.,  20.,
              33.,  34.,  35.,  36.,  37.,  38.,
              51.,  52.,  53.,  54.,  55.,  56.,
              69.,  70.,  71.,  72.,  73.,  74.]
        well_info_3['Measure'] = [True]*24

        well_info_4 = pandas.DataFrame()
        well_info_4[u'IPTG Concentration (µM)'] = \
            [  75.,   76.,   77.,   78.,   79.,   80.,
               93.,   94.,   95.,   96.,   97.,   98.,
              111.,  112.,  113.,  114.,  115.,  116.,
              129.,  130.,  131.,  132.,  133.,  134.]
        well_info_4['Measure'] = [True]*24

        well_info_5 = pandas.DataFrame()
        well_info_5[u'IPTG Concentration (µM)'] = \
            [  81.,   82.,   83.,   84.,   85.,   86.,
               99.,  100.,  101.,  102.,  103.,  104.,
              117.,  118.,  119.,  120.,  121.,  122.,
              135.,  136.,  137.,  138.,  139.,  140.]
        well_info_5['Measure'] = [True]*24

        well_info_6 = pandas.DataFrame()
        well_info_6[u'IPTG Concentration (µM)'] = \
            [  87.,   88.,   89.,   90.,   91.,   92.,
              105.,  106.,  107.,  108.,  109.,  110.,
              123.,  124.,  125.,  126.,  127.,  128.,
              141.,  142.,  143.,  144.,  145.,  146.]
        well_info_6['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_3)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_4)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_5)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_6)

    def test_close_plates_inducer_wells_2(self):
        # Create plate
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(144) + 3
        p.apply_inducer(iptg, apply_to='wells')

        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = (numpy.arange(144) + 1)/10.
        p.apply_inducer(atc, apply_to='wells')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [  3.,   4.,   5.,   6.,   7.,   8.,
              21.,  22.,  23.,  24.,  25.,  26.,
              39.,  40.,  41.,  42.,  43.,  44.,
              57.,  58.,  59.,  60.,  61.,  62.]
        well_info_1[u'aTc Concentration (ng/µL)'] = \
            [ 0.1,  0.2,  0.3,  0.4,  0.5,  0.6,
              1.9,  2. ,  2.1,  2.2,  2.3,  2.4,
              3.7,  3.8,  3.9,  4. ,  4.1,  4.2,
              5.5,  5.6,  5.7,  5.8,  5.9,  6. ]
        well_info_1['Measure'] = [True]*24

        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [  9.,  10.,  11.,  12.,  13.,  14.,
              27.,  28.,  29.,  30.,  31.,  32.,
              45.,  46.,  47.,  48.,  49.,  50.,
              63.,  64.,  65.,  66.,  67.,  68.]
        well_info_2[u'aTc Concentration (ng/µL)'] = \
            [ 0.7,  0.8,  0.9,  1. ,  1.1,  1.2,
              2.5,  2.6,  2.7,  2.8,  2.9,  3. ,
              4.3,  4.4,  4.5,  4.6,  4.7,  4.8,
              6.1,  6.2,  6.3,  6.4,  6.5,  6.6]
        well_info_2['Measure'] = [True]*24

        well_info_3 = pandas.DataFrame()
        well_info_3[u'IPTG Concentration (µM)'] = \
            [ 15.,  16.,  17.,  18.,  19.,  20.,
              33.,  34.,  35.,  36.,  37.,  38.,
              51.,  52.,  53.,  54.,  55.,  56.,
              69.,  70.,  71.,  72.,  73.,  74.]
        well_info_3[u'aTc Concentration (ng/µL)'] = \
            [ 1.3,  1.4,  1.5,  1.6,  1.7,  1.8,
              3.1,  3.2,  3.3,  3.4,  3.5,  3.6,
              4.9,  5. ,  5.1,  5.2,  5.3,  5.4,
              6.7,  6.8,  6.9,  7. ,  7.1,  7.2]
        well_info_3['Measure'] = [True]*24

        well_info_4 = pandas.DataFrame()
        well_info_4[u'IPTG Concentration (µM)'] = \
            [  75.,   76.,   77.,   78.,   79.,   80.,
               93.,   94.,   95.,   96.,   97.,   98.,
              111.,  112.,  113.,  114.,  115.,  116.,
              129.,  130.,  131.,  132.,  133.,  134.]
        well_info_4[u'aTc Concentration (ng/µL)'] = \
            [  7.3,   7.4,   7.5,   7.6,   7.7,   7.8,
               9.1,   9.2,   9.3,   9.4,   9.5,   9.6,
              10.9,  11. ,  11.1,  11.2,  11.3,  11.4,
              12.7,  12.8,  12.9,  13. ,  13.1,  13.2]
        well_info_4['Measure'] = [True]*24

        well_info_5 = pandas.DataFrame()
        well_info_5[u'IPTG Concentration (µM)'] = \
            [  81.,   82.,   83.,   84.,   85.,   86.,
               99.,  100.,  101.,  102.,  103.,  104.,
              117.,  118.,  119.,  120.,  121.,  122.,
              135.,  136.,  137.,  138.,  139.,  140.]
        well_info_5[u'aTc Concentration (ng/µL)'] = \
            [  7.9,   8. ,   8.1,   8.2,   8.3,   8.4,
               9.7,   9.8,   9.9,  10. ,  10.1,  10.2,
              11.5,  11.6,  11.7,  11.8,  11.9,  12. ,
              13.3,  13.4,  13.5,  13.6,  13.7,  13.8]
        well_info_5['Measure'] = [True]*24

        well_info_6 = pandas.DataFrame()
        well_info_6[u'IPTG Concentration (µM)'] = \
            [  87.,   88.,   89.,   90.,   91.,   92.,
              105.,  106.,  107.,  108.,  109.,  110.,
              123.,  124.,  125.,  126.,  127.,  128.,
              141.,  142.,  143.,  144.,  145.,  146.]
        well_info_6[u'aTc Concentration (ng/µL)'] = \
            [  8.5,   8.6,   8.7,   8.8,   8.9,   9. ,
              10.3,  10.4,  10.5,  10.6,  10.7,  10.8,
              12.1,  12.2,  12.3,  12.4,  12.5,  12.6,
              13.9,  14. ,  14.1,  14.2,  14.3,  14.4]
        well_info_6['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_3)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_4)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_5)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_6)

    def test_close_plates_inducer_wells_3(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(144) + 3
        iptg.shuffle()
        p.apply_inducer(iptg, apply_to='wells')

        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = (numpy.arange(144) + 1)/10.
        p.apply_inducer(atc, apply_to='wells')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [  20.,  118.,   83.,   51.,   64.,   61.,
               25.,  112.,    4.,   63.,   87.,    9.,
              104.,   16.,   21.,   23.,   11.,   26.,
               79.,   18.,   70.,  141.,   59.,   33.]
        well_info_1[u'aTc Concentration (ng/µL)'] = \
            [ 0.1,  0.2,  0.3,  0.4,  0.5,  0.6,
              1.9,  2. ,  2.1,  2.2,  2.3,  2.4,
              3.7,  3.8,  3.9,  4. ,  4.1,  4.2,
              5.5,  5.6,  5.7,  5.8,  5.9,  6. ]
        well_info_1['Measure'] = [True]*24

        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [  74.,   13.,   46.,  125.,   77.,   35.,
               85.,  100.,   81.,  129.,   19.,    8.,
               40.,  102.,  107.,  119.,  136.,   41.,
               86.,   56.,   68.,  101.,   97.,  134.]
        well_info_2[u'aTc Concentration (ng/µL)'] = \
            [ 0.7,  0.8,  0.9,  1. ,  1.1,  1.2,
              2.5,  2.6,  2.7,  2.8,  2.9,  3. ,
              4.3,  4.4,  4.5,  4.6,  4.7,  4.8,
              6.1,  6.2,  6.3,  6.4,  6.5,  6.6]
        well_info_2['Measure'] = [True]*24

        well_info_3 = pandas.DataFrame()
        well_info_3[u'IPTG Concentration (µM)'] = \
            [  50.,  131.,   12.,   30.,    7.,   45.,
               10.,  105.,   58.,   78.,   89.,   67.,
               76.,   47.,   44.,  142.,  133.,  113.,
               14.,  115.,   99.,   88.,   52.,   90.]
        well_info_3[u'aTc Concentration (ng/µL)'] = \
            [ 1.3,  1.4,  1.5,  1.6,  1.7,  1.8,
              3.1,  3.2,  3.3,  3.4,  3.5,  3.6,
              4.9,  5. ,  5.1,  5.2,  5.3,  5.4,
              6.7,  6.8,  6.9,  7. ,  7.1,  7.2]
        well_info_3['Measure'] = [True]*24

        well_info_4 = pandas.DataFrame()
        well_info_4[u'IPTG Concentration (µM)'] = \
            [ 139.,   39.,   98.,  110.,  109.,   42.,
               80.,   84.,   57.,   31.,   66.,   82.,
                5.,   34.,   54.,   27.,  128.,   29.,
              123.,   32.,   96.,  135.,    3.,  103.]
        well_info_4[u'aTc Concentration (ng/µL)'] = \
            [  7.3,   7.4,   7.5,   7.6,   7.7,   7.8,
               9.1,   9.2,   9.3,   9.4,   9.5,   9.6,
              10.9,  11. ,  11.1,  11.2,  11.3,  11.4,
              12.7,  12.8,  12.9,  13. ,  13.1,  13.2]
        well_info_4['Measure'] = [True]*24

        well_info_5 = pandas.DataFrame()
        well_info_5[u'IPTG Concentration (µM)'] = \
            [  37.,  130.,   95.,  145.,  140.,   17.,
               43.,   94.,   73.,   75.,   36.,  138.,
              132.,  121.,  122.,  126.,   53.,   28.,
               60.,  114.,    6.,   15.,  144.,   92.]
        well_info_5[u'aTc Concentration (ng/µL)'] = \
            [  7.9,   8. ,   8.1,   8.2,   8.3,   8.4,
               9.7,   9.8,   9.9,  10. ,  10.1,  10.2,
              11.5,  11.6,  11.7,  11.8,  11.9,  12. ,
              13.3,  13.4,  13.5,  13.6,  13.7,  13.8]
        well_info_5['Measure'] = [True]*24

        well_info_6 = pandas.DataFrame()
        well_info_6[u'IPTG Concentration (µM)'] = \
            [ 143.,  108.,   24.,  120.,   55.,   48.,
               91.,  106.,  146.,   71.,   62.,   93.,
               49.,  117.,   69.,  127.,  137.,  116.,
               65.,   72.,   38.,  111.,  124.,   22.]
        well_info_6[u'aTc Concentration (ng/µL)'] = \
            [  8.5,   8.6,   8.7,   8.8,   8.9,   9. ,
              10.3,  10.4,  10.5,  10.6,  10.7,  10.8,
              12.1,  12.2,  12.3,  12.4,  12.5,  12.6,
              13.9,  14. ,  14.1,  14.2,  14.3,  14.4]
        well_info_6['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_3)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_4)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_5)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_6)

    def test_close_plates_inducer_wells_4(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.samples_to_measure = 80
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(80) + 3
        p.apply_inducer(iptg, apply_to='wells')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [  3.,   4.,   5.,   6.,   7.,   8.,
              21.,  22.,  23.,  24.,  25.,  26.,
              39.,  40.,  41.,  42.,  43.,  44.,
              57.,  58.,  59.,  60.,  61.,  62.]
        well_info_1['Measure'] = [True]*24

        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [  9.,  10.,  11.,  12.,  13.,  14.,
              27.,  28.,  29.,  30.,  31.,  32.,
              45.,  46.,  47.,  48.,  49.,  50.,
              63.,  64.,  65.,  66.,  67.,  68.]
        well_info_2['Measure'] = [True]*24

        well_info_3 = pandas.DataFrame()
        well_info_3[u'IPTG Concentration (µM)'] = \
            [ 15.,  16.,  17.,  18.,  19.,  20.,
              33.,  34.,  35.,  36.,  37.,  38.,
              51.,  52.,  53.,  54.,  55.,  56.,
              69.,  70.,  71.,  72.,  73.,  74.]
        well_info_3['Measure'] = [True]*24

        well_info_4 = pandas.DataFrame()
        well_info_4[u'IPTG Concentration (µM)'] = \
            [       75.,        76.,        77.,        78.,        79.,        80.,
              numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
              numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
              numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan]
        well_info_4['Measure'] = [True]*6 + [False]* 18

        well_info_5 = pandas.DataFrame()
        well_info_5[u'IPTG Concentration (µM)'] = \
            [      81.,        82.,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan]
        well_info_5['Measure'] = [True]*2 + [False]* 22

        well_info_6 = pandas.DataFrame()
        well_info_6[u'IPTG Concentration (µM)'] = \
            [numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan]
        well_info_6['Measure'] = [False]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_3)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_4)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_5)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_6)

    def test_close_plates_inducer_media_1(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3]
        p.apply_inducer(iptg, apply_to='media')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [  3.,   3.,   3.,   3.,   3.,   3.,
               3.,   3.,   3.,   3.,   3.,   3.,
               3.,   3.,   3.,   3.,   3.,   3.,
               3.,   3.,   3.,   3.,   3.,   3.]
        well_info_1['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_1)

    def test_close_plates_inducer_media_2(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3]
        p.apply_inducer(iptg, apply_to='media')

        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = [5]
        p.apply_inducer(atc, apply_to='media')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
        [  3.,   3.,   3.,   3.,   3.,   3.,
           3.,   3.,   3.,   3.,   3.,   3.,
           3.,   3.,   3.,   3.,   3.,   3.,
           3.,   3.,   3.,   3.,   3.,   3.]
        well_info_1[u'aTc Concentration (ng/µL)'] = \
        [  5.,   5.,   5.,   5.,   5.,   5.,
           5.,   5.,   5.,   5.,   5.,   5.,
           5.,   5.,   5.,   5.,   5.,   5.,
           5.,   5.,   5.,   5.,   5.,   5.]
        well_info_1['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_1)

    def test_close_plates_inducer_media_3(self):
        # Create plate
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.samples_to_measure = 80
        p.cell_strain_name = 'Test strain 1'

        # Create inducer for plate
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = [3]
        p.apply_inducer(iptg, apply_to='media')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)
        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 2)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [  3.,   3.,   3.,   3.,   3.,   3.,
               3.,   3.,   3.,   3.,   3.,   3.,
               3.,   3.,   3.,   3.,   3.,   3.,
               3.,   3.,   3.,   3.,   3.,   3.]
        well_info_1['Measure'] = [True]*24

        well_info_4 = pandas.DataFrame()
        well_info_4[u'IPTG Concentration (µM)'] = \
            [        3.,         3.,         3.,         3.,         3.,         3.,
              numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
              numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
              numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan]
        well_info_4['Measure'] = [True]*6 + [False]* 18

        well_info_5 = pandas.DataFrame()
        well_info_5[u'IPTG Concentration (µM)'] = \
            [       3.,         3.,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan]
        well_info_5['Measure'] = [True]*2 + [False]* 22

        well_info_6 = pandas.DataFrame()
        well_info_6[u'IPTG Concentration (µM)'] = \
            [numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,
             numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan,  numpy.nan]
        well_info_6['Measure'] = [False]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_4)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_5)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_6)

    def test_close_plates_combined(self):
        p = platedesign.plate.PlateArray(name='A1',
                                         array_n_rows=2,
                                         array_n_cols=3,
                                         plate_names=['P{}'.format(i+1)
                                                      for i in range(6)])
        p.cell_strain_name = 'Test strain 1'
        # Add some metadata
        p.metadata['Meta 1'] = 'Value 1'
        p.metadata['Meta 2'] = 'Value 2'
        # Add cell inoculation info
        p.cell_setup_method = 'fixed_volume'
        p.cell_predilution = 100
        p.cell_predilution_vol = 1000
        p.cell_shot_vol = 5

        # Create inducer for plate rows
        iptg = platedesign.inducer.ChemicalInducer(
            name='IPTG',
            units=u'µM')
        iptg.concentrations = numpy.arange(18) + 3
        p.apply_inducer(iptg, apply_to='rows')

        # Second inducer for plate columns
        atc = platedesign.inducer.ChemicalInducer(
            name='aTc',
            units=u'ng/µL')
        atc.concentrations = (numpy.arange(8.) + 1)/10.
        p.apply_inducer(atc, apply_to='cols')

        # Call close plates and check length of output
        cps = p.close_plates()
        self.assertEqual(len(cps), 6)

        # Check basic properties
        self.assertEqual(cps[0].name, 'P1')
        self.assertEqual(cps[1].name, 'P2')
        self.assertEqual(cps[2].name, 'P3')
        self.assertEqual(cps[3].name, 'P4')
        self.assertEqual(cps[4].name, 'P5')
        self.assertEqual(cps[5].name, 'P6')
        for cp in cps:
            self.assertEqual(cp.n_rows, 4)
            self.assertEqual(cp.n_cols, 6)
            # Check plate info
            self.assertEqual(len(cp.plate_info), 6)
            self.assertTrue('Plate Array' in cp.plate_info)
            self.assertEqual(cp.plate_info['Plate Array'], 'A1')
            self.assertTrue('Strain' in cp.plate_info)
            self.assertEqual(cp.plate_info['Strain'], 'Test strain 1')
            self.assertTrue('Meta 1' in cp.plate_info)
            self.assertEqual(cp.plate_info['Meta 1'], 'Value 1')
            self.assertTrue('Meta 2' in cp.plate_info)
            self.assertEqual(cp.plate_info['Meta 2'], 'Value 2')
            self.assertTrue('Preculture Dilution' in cp.plate_info)
            self.assertEqual(cp.plate_info['Preculture Dilution'], 100)
            self.assertTrue('Cell Inoculated Vol.' in cp.plate_info)
            self.assertEqual(cp.plate_info['Cell Inoculated Vol.'], 5)

        # Check well info
        well_info_1 = pandas.DataFrame()
        well_info_1[u'IPTG Concentration (µM)'] = \
            [3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.]
        well_info_1[u'aTc Concentration (ng/µL)'] = \
            [0.1, 0.1, 0.1, 0.1, 0.1, 0.1,
             0.2, 0.2, 0.2, 0.2, 0.2, 0.2,
             0.3, 0.3, 0.3, 0.3, 0.3, 0.3,
             0.4, 0.4, 0.4, 0.4, 0.4, 0.4,]
        well_info_1['Measure'] = [True]*24

        well_info_2 = pandas.DataFrame()
        well_info_2[u'IPTG Concentration (µM)'] = \
            [9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,]
        well_info_2[u'aTc Concentration (ng/µL)'] = \
            [0.1, 0.1, 0.1, 0.1, 0.1, 0.1,
             0.2, 0.2, 0.2, 0.2, 0.2, 0.2,
             0.3, 0.3, 0.3, 0.3, 0.3, 0.3,
             0.4, 0.4, 0.4, 0.4, 0.4, 0.4,]
        well_info_2['Measure'] = [True]*24

        well_info_3 = pandas.DataFrame()
        well_info_3[u'IPTG Concentration (µM)'] = \
            [15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,]
        well_info_3[u'aTc Concentration (ng/µL)'] = \
            [0.1, 0.1, 0.1, 0.1, 0.1, 0.1,
             0.2, 0.2, 0.2, 0.2, 0.2, 0.2,
             0.3, 0.3, 0.3, 0.3, 0.3, 0.3,
             0.4, 0.4, 0.4, 0.4, 0.4, 0.4,]
        well_info_3['Measure'] = [True]*24

        well_info_4 = pandas.DataFrame()
        well_info_4[u'IPTG Concentration (µM)'] = \
            [3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.,
             3., 4., 5., 6., 7., 8.]
        well_info_4[u'aTc Concentration (ng/µL)'] = \
            [0.5, 0.5, 0.5, 0.5, 0.5, 0.5,
             0.6, 0.6, 0.6, 0.6, 0.6, 0.6,
             0.7, 0.7, 0.7, 0.7, 0.7, 0.7,
             0.8, 0.8, 0.8, 0.8, 0.8, 0.8,]
        well_info_4['Measure'] = [True]*24

        well_info_5 = pandas.DataFrame()
        well_info_5[u'IPTG Concentration (µM)'] = \
            [9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,
             9., 10., 11., 12., 13., 14.,]
        well_info_5[u'aTc Concentration (ng/µL)'] = \
            [0.5, 0.5, 0.5, 0.5, 0.5, 0.5,
             0.6, 0.6, 0.6, 0.6, 0.6, 0.6,
             0.7, 0.7, 0.7, 0.7, 0.7, 0.7,
             0.8, 0.8, 0.8, 0.8, 0.8, 0.8,]
        well_info_5['Measure'] = [True]*24

        well_info_6 = pandas.DataFrame()
        well_info_6[u'IPTG Concentration (µM)'] = \
            [15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,
             15., 16., 17., 18., 19., 20,]
        well_info_6[u'aTc Concentration (ng/µL)'] = \
            [0.5, 0.5, 0.5, 0.5, 0.5, 0.5,
             0.6, 0.6, 0.6, 0.6, 0.6, 0.6,
             0.7, 0.7, 0.7, 0.7, 0.7, 0.7,
             0.8, 0.8, 0.8, 0.8, 0.8, 0.8,]
        well_info_6['Measure'] = [True]*24

        pandas.util.testing.assert_frame_equal(cps[0].well_info, well_info_1)
        pandas.util.testing.assert_frame_equal(cps[1].well_info, well_info_2)
        pandas.util.testing.assert_frame_equal(cps[2].well_info, well_info_3)
        pandas.util.testing.assert_frame_equal(cps[3].well_info, well_info_4)
        pandas.util.testing.assert_frame_equal(cps[4].well_info, well_info_5)
        pandas.util.testing.assert_frame_equal(cps[5].well_info, well_info_6)

class TestClosedPlate(unittest.TestCase):
    """
    Tests for the ClosedPlate class

    """
    def test_create(self):
        # Create
        p = platedesign.plate.ClosedPlate(name='P1')

    def test_attributes_no_optional_parameters(self):
        # Create
        p = platedesign.plate.ClosedPlate(name='P1')
        # Check all default attributes
        self.assertEqual(p.name, 'P1')
        self.assertEqual(p.n_rows, 4)
        self.assertEqual(p.n_cols, 6)
        self.assertIsNone(p.plate_info)
        self.assertIsNone(p.well_info)
        # Check samples table
        exp_samples_table = pandas.DataFrame(
            {'Plate': ['P1']*24,
             'Row': [1,1,1,1,1,1,
                     2,2,2,2,2,2,
                     3,3,3,3,3,3,
                     4,4,4,4,4,4],
             'Column': [1,2,3,4,5,6]*4})
        exp_samples_table = exp_samples_table[['Plate', 'Row', 'Column']]
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)

    def test_attributes_non_default_size(self):
        # Create
        p = platedesign.plate.ClosedPlate(name='P1', n_rows=8, n_cols=12)
        # Check all default attributes
        self.assertEqual(p.name, 'P1')
        self.assertEqual(p.n_rows, 8)
        self.assertEqual(p.n_cols, 12)
        self.assertIsNone(p.plate_info)
        self.assertIsNone(p.well_info)
        # Check samples table
        exp_samples_table = pandas.DataFrame(
            {'Plate': ['P1']*96,
             'Row': [1,1,1,1,1,1,1,1,1,1,1,1,
                     2,2,2,2,2,2,2,2,2,2,2,2,
                     3,3,3,3,3,3,3,3,3,3,3,3,
                     4,4,4,4,4,4,4,4,4,4,4,4,
                     5,5,5,5,5,5,5,5,5,5,5,5,
                     6,6,6,6,6,6,6,6,6,6,6,6,
                     7,7,7,7,7,7,7,7,7,7,7,7,
                     8,8,8,8,8,8,8,8,8,8,8,8],
             'Column': [1,2,3,4,5,6,7,8,9,10,11,12]*8})
        exp_samples_table = exp_samples_table[['Plate', 'Row', 'Column']]
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)

    def test_attributes_non_default_plate_info(self):
        # Create
        plate_info = collections.OrderedDict()
        plate_info['Cell strain'] = 'Strain 1'
        plate_info['Initial OD'] = 5e-5
        p = platedesign.plate.ClosedPlate(name='P1', plate_info=plate_info)
        # Check all default attributes
        self.assertEqual(p.name, 'P1')
        self.assertEqual(p.n_rows, 4)
        self.assertEqual(p.n_cols, 6)
        self.assertEqual(p.plate_info, plate_info)
        self.assertIsNone(p.well_info)
        # Check samples table
        exp_samples_table = pandas.DataFrame(
            {'Plate': ['P1']*24,
             'Cell strain': ['Strain 1']*24,
             'Initial OD': [5e-5]*24,
             'Row': [1,1,1,1,1,1,
                     2,2,2,2,2,2,
                     3,3,3,3,3,3,
                     4,4,4,4,4,4],
             'Column': [1,2,3,4,5,6]*4})
        exp_samples_table = exp_samples_table[['Plate',
                                               'Cell strain',
                                               'Initial OD',
                                               'Row',
                                               'Column']]
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)

    def test_attributes_non_default_well_info(self):
        # Create
        well_info = pandas.DataFrame()
        well_info['IPTG'] = [1,2,1,4,7,8,4,7,9,6,5,8,4,2,1,4,7,0,3,2,1,5,3,1]
        well_info['Measure'] = [True]*10 + [False]*14
        p = platedesign.plate.ClosedPlate(name='P1', well_info=well_info)
        # Check all default attributes
        self.assertEqual(p.name, 'P1')
        self.assertEqual(p.n_rows, 4)
        self.assertEqual(p.n_cols, 6)
        self.assertIsNone(p.plate_info)
        pandas.util.testing.assert_frame_equal(p.well_info, well_info)
        # Check samples table
        exp_samples_table = pandas.DataFrame(
            {'Plate': ['P1']*24,
             'Row': [1,1,1,1,1,1,
                     2,2,2,2,2,2,
                     3,3,3,3,3,3,
                     4,4,4,4,4,4],
             'Column': [1,2,3,4,5,6]*4,
             'IPTG':[1,2,1,4,7,8,4,7,9,6,5,8,4,2,1,4,7,0,3,2,1,5,3,1],
             'Measure':[True]*10 + [False]*14})
        exp_samples_table = exp_samples_table[['Plate',
                                               'Row',
                                               'Column',
                                               'IPTG',
                                               'Measure']]
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)

    def test_attributes_non_default_info(self):
        # Create
        plate_info = collections.OrderedDict()
        plate_info['Cell strain'] = 'Strain 1'
        plate_info['Initial OD'] = 5e-5
        well_info = pandas.DataFrame()
        well_info['IPTG'] = [1,2,1,4,7,8,4,7,9,6,5,8,4,2,1,4,7,0,3,2,1,5,3,1]
        well_info['Measure'] = [True]*10 + [False]*14
        p = platedesign.plate.ClosedPlate(name='P1',
                                          plate_info=plate_info,
                                          well_info=well_info)
        # Check all default attributes
        self.assertEqual(p.name, 'P1')
        self.assertEqual(p.n_rows, 4)
        self.assertEqual(p.n_cols, 6)
        self.assertEqual(p.plate_info, plate_info)
        pandas.util.testing.assert_frame_equal(p.well_info, well_info)
        # Check samples table
        exp_samples_table = pandas.DataFrame(
            {'Plate': ['P1']*24,
             'Cell strain': ['Strain 1']*24,
             'Initial OD': [5e-5]*24,
             'Row': [1,1,1,1,1,1,
                     2,2,2,2,2,2,
                     3,3,3,3,3,3,
                     4,4,4,4,4,4],
             'Column': [1,2,3,4,5,6]*4,
             'IPTG':[1,2,1,4,7,8,4,7,9,6,5,8,4,2,1,4,7,0,3,2,1,5,3,1],
             'Measure':[True]*10 + [False]*14})
        exp_samples_table = exp_samples_table[['Plate',
                                               'Cell strain',
                                               'Initial OD',
                                               'Row',
                                               'Column',
                                               'IPTG',
                                               'Measure']]
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)

    def test_modify_plate_info(self):
        # Create
        plate_info = collections.OrderedDict()
        plate_info['Cell strain'] = 'Strain 1'
        plate_info['Initial OD'] = 5e-5
        well_info = pandas.DataFrame()
        well_info['IPTG'] = [1,2,1,4,7,8,4,7,9,6,5,8,4,2,1,4,7,0,3,2,1,5,3,1]
        well_info['Measure'] = [True]*10 + [False]*14
        p = platedesign.plate.ClosedPlate(name='P1',
                                          plate_info=plate_info,
                                          well_info=well_info)
        # Check samples table
        exp_samples_table = pandas.DataFrame(
            {'Plate': ['P1']*24,
             'Cell strain': ['Strain 1']*24,
             'Initial OD': [5e-5]*24,
             'Row': [1,1,1,1,1,1,
                     2,2,2,2,2,2,
                     3,3,3,3,3,3,
                     4,4,4,4,4,4],
             'Column': [1,2,3,4,5,6]*4,
             'IPTG':[1,2,1,4,7,8,4,7,9,6,5,8,4,2,1,4,7,0,3,2,1,5,3,1],
             'Measure':[True]*10 + [False]*14})
        exp_samples_table = exp_samples_table[['Plate',
                                               'Cell strain',
                                               'Initial OD',
                                               'Row',
                                               'Column',
                                               'IPTG',
                                               'Measure']]
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)

        # Add field to plate_info
        p.plate_info['Predilution'] = 10
        # Samples table should be the same as before
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)
        # Update and check that samples table has changed
        p.update_samples_table()
        exp_samples_table = pandas.DataFrame(
            {'Plate': ['P1']*24,
             'Cell strain': ['Strain 1']*24,
             'Initial OD': [5e-5]*24,
             'Predilution': [10]*24,
             'Row': [1,1,1,1,1,1,
                     2,2,2,2,2,2,
                     3,3,3,3,3,3,
                     4,4,4,4,4,4],
             'Column': [1,2,3,4,5,6]*4,
             'IPTG':[1,2,1,4,7,8,4,7,9,6,5,8,4,2,1,4,7,0,3,2,1,5,3,1],
             'Measure':[True]*10 + [False]*14})
        exp_samples_table = exp_samples_table[['Plate',
                                               'Cell strain',
                                               'Initial OD',
                                               'Predilution',
                                               'Row',
                                               'Column',
                                               'IPTG',
                                               'Measure']]
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)


    def test_modify_well_info(self):
        # Create
        plate_info = collections.OrderedDict()
        plate_info['Cell strain'] = 'Strain 1'
        plate_info['Initial OD'] = 5e-5
        well_info = pandas.DataFrame()
        well_info['IPTG'] = [1,2,1,4,7,8,4,7,9,6,5,8,4,2,1,4,7,0,3,2,1,5,3,1]
        well_info['Measure'] = [True]*10 + [False]*14
        p = platedesign.plate.ClosedPlate(name='P1',
                                          plate_info=plate_info,
                                          well_info=well_info)
        pandas.util.testing.assert_frame_equal(p.well_info, well_info)
        # Check samples table
        exp_samples_table = pandas.DataFrame(
            {'Plate': ['P1']*24,
             'Cell strain': ['Strain 1']*24,
             'Initial OD': [5e-5]*24,
             'Row': [1,1,1,1,1,1,
                     2,2,2,2,2,2,
                     3,3,3,3,3,3,
                     4,4,4,4,4,4],
             'Column': [1,2,3,4,5,6]*4,
             'IPTG':[1,2,1,4,7,8,4,7,9,6,5,8,4,2,1,4,7,0,3,2,1,5,3,1],
             'Measure':[True]*10 + [False]*14})
        exp_samples_table = exp_samples_table[['Plate',
                                               'Cell strain',
                                               'Initial OD',
                                               'Row',
                                               'Column',
                                               'IPTG',
                                               'Measure']]
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)

        # Change one of the columns in well_info
        p.well_info['Measure'] = [False]*8 + [True]*16
        # Samples table should be the same as before
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)
        # Update and check that samples table has changed
        p.update_samples_table()
        exp_samples_table = pandas.DataFrame(
            {'Plate': ['P1']*24,
             'Cell strain': ['Strain 1']*24,
             'Initial OD': [5e-5]*24,
             'Row': [1,1,1,1,1,1,
                     2,2,2,2,2,2,
                     3,3,3,3,3,3,
                     4,4,4,4,4,4],
             'Column': [1,2,3,4,5,6]*4,
             'IPTG':[1,2,1,4,7,8,4,7,9,6,5,8,4,2,1,4,7,0,3,2,1,5,3,1],
             'Measure':[False]*8 + [True]*16})
        exp_samples_table = exp_samples_table[['Plate',
                                               'Cell strain',
                                               'Initial OD',
                                               'Row',
                                               'Column',
                                               'IPTG',
                                               'Measure']]
        pandas.util.testing.assert_frame_equal(exp_samples_table,
                                               p.samples_table)
